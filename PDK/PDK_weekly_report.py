import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import logging
import win32com.client as win32
import time
import sys
from datetime import datetime
from typing import Optional
from dataclasses import dataclass

# 设置控制台编码为 UTF-8
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

@dataclass
class EmailConfig:
    """邮件配置类"""
    retry_times: int = 3
    retry_delay: int = 2

@dataclass
class PathConfig:
    """路径配置类"""
    source_path: str
    target_path: str
    output_path: str

class PDKReportConfig:
    """PDK报告配置管理类"""
    def __init__(self):
        self.email = EmailConfig()
        self.paths = PathConfig(
            source_path=r'\\atdfile3.ch.intel.com\PRO\Reports',
            target_path=r'C:\\Users\\qibaiche\\OneDrive - Intel Corporation\\Documents\\PDK',
            output_path=r'C:\\Users\\qibaiche\\OneDrive - Intel Corporation\\Documents\\PDK'
        )
        self.relevant_columns = ['ProgramName', 'ModelName', 'CCB', 'Package', 'Device', 'Revision', 'Stepping', 'DieCodeName', 'TestName']
        self.pivot_columns = ['DieCodeName', 'ProgramName', 'CCB', 'Package', 'Device', 'Revision', 'Stepping']

# 全局配置实例
CONFIG = PDKReportConfig()

# 兼容性：保持原有的CONFIG字典格式（移除不再需要的字段）
CONFIG_DICT = {
    'source_path': CONFIG.paths.source_path,
    'target_path': CONFIG.paths.target_path,
    'output_path': CONFIG.paths.output_path,
    'relevant_columns': CONFIG.relevant_columns,
    'pivot_columns': CONFIG.pivot_columns
}


def setup_logging(output_path: str) -> str:
    """设置日志记录，优化了编码处理"""
    log_dir = os.path.join(output_path, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(log_dir, f'comparison_log_{timestamp}.txt')
    
    # 清除已有的处理器，避免重复日志
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)
    
    # 配置日志
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ],
        force=True  # 强制重新配置
    )
    
    return log_file

class OutlookManager:
    """Outlook邮件管理器，单例模式避免重复启动"""
    _instance = None
    _outlook = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance
    
    def get_outlook(self):
        """获取Outlook应用实例"""
        if self._outlook is None:
            try:
                self._outlook = win32.Dispatch('Outlook.Application')
            except:
                os.system('start outlook')
                time.sleep(5)
                self._outlook = win32.Dispatch('Outlook.Application')
        return self._outlook
    
    def send_email_with_retry(self, send_func, max_retries: int = 3, delay: int = 2) -> bool:
        """带重试的邮件发送"""
        for attempt in range(max_retries):
            try:
                send_func()
                return True
            except Exception as e:
                if attempt < max_retries - 1:
                    logging.warning(f"邮件发送失败，第{attempt + 1}次重试: {str(e)}")
                    time.sleep(delay)
                else:
                    logging.error(f"邮件发送最终失败: {str(e)}")
                    return False
        return False

class HTMLEmailTemplate:
    """HTML邮件模板管理器"""
    
    @staticmethod
    def get_common_styles() -> str:
        """获取通用样式"""
        return """
        <style>
            body { font-family: Arial, sans-serif; }
            table { 
                border-collapse: collapse; 
                margin-top: 10px; 
                margin-bottom: 20px; 
            }
            th, td { 
                border: 1px solid #ddd; 
                padding: 2px 4px; 
                text-align: left;
                line-height: 1;
                font-size: 12px;
            }
            th { background-color: #f2f2f2; }
            .added { color: green; font-weight: bold; }
            .removed { color: red; font-weight: bold; }
            .summary { margin-bottom: 15px; }
            .dataframe { margin: 0; border-spacing: 0; }
            .dataframe tr { height: 18px; line-height: 1; }
            .dataframe td, .dataframe th { height: auto; }
            .sharepoint-link { margin-top: 15px; margin-bottom: 15px; }
            .important-text { 
                color: red; 
                background-color: yellow; 
                padding: 2px 4px; 
                font-weight: bold; 
            }
            .highlight { background-color: #ffffcc; }
        </style>
        """
    
    @staticmethod
    def create_weekly_report_body(added_pivot, removed_pivot) -> str:
        """创建周报邮件正文"""
        added_count = len(added_pivot) if not added_pivot.empty else 0
        removed_count = len(removed_pivot) if not removed_pivot.empty else 0
        
        body = f"""
        <html>
        <head>{HTMLEmailTemplate.get_common_styles()}</head>
        <body>
            <p>Hi TPO,</p>
            <p>The following outlines the changes to the PDK model since the last update.</p>
            <p>Kindly review whether these changes are relevant to your product or align with the requirements you previously submitted.</p>
            <p><span class="important-text">If you want to get a specific notification of your prod, please write related information at this file</span> <a href="https://intel-my.sharepoint.com/:x:/p/qibai_chen/EeMwoE_DGjRPrBUaK-zBcrABwIWUCqjG3KMjcFhGfb5NUg?e=qQtOnB">here</a>.</p>
            <div class="summary">
                <p><b>Summary:</b></p>
                <ul>
                    <li>Added items: <span class="added">{added_count}</span></li>
                    <li>Removed items: <span class="removed">{removed_count}</span></li>
                </ul>
            </div>
        """
        
        # 添加数据表格
        if not added_pivot.empty:
            body += '<div class="added">Added items:</div>'
            body += added_pivot.to_html(index=False, classes='dataframe')
        else:
            body += '<div class="added">No items added</div>'
            
        if not removed_pivot.empty:
            body += '<div class="removed">Removed items:</div>'
            body += removed_pivot.to_html(index=False, classes='dataframe')
        else:
            body += '<div class="removed">No items removed</div>'
            
        body += """
            <p>For detailed information, please check the attachment.</p>
            <p>If you have any questions, please feel free to contact me.</p>
            <p>Thank you!</p>
            <p>Best Regards</p>
            <p>Qibai Chen (陈麒百)</p>
            <p>CDG CD PDE</p>
        </body>
        </html>
        """
        
        return body

# 已移除按产品负责人通知相关函数（get_product_owners、send_owner_notification_optimized、notify_product_owners_optimized）

def _rotate_pdk_files(target_dir: str):
    """将 PDK_Programs_old.csv -> PDK_Programs_bak.csv，并将 PDK_Programs.csv -> PDK_Programs_old.csv。
    缺失文件将被忽略，目标存在则用 os.replace 覆盖。
    """
    try:
        old_path = os.path.join(target_dir, 'PDK_Programs_old.csv')
        bak_path = os.path.join(target_dir, 'PDK_Programs_bak.csv')
        cur_path = os.path.join(target_dir, 'PDK_Programs.csv')
        # old -> bak
        if os.path.exists(old_path):
            try:
                os.replace(old_path, bak_path)
            except Exception as e:
                logging.warning(f'无法重命名旧文件到备份: {e}')
        # current -> old
        if os.path.exists(cur_path):
            try:
                os.replace(cur_path, old_path)
            except Exception as e:
                logging.warning(f'无法将当前文件重命名为旧文件: {e}')
    except Exception as e:
        logging.error(f'轮转PDK文件时出错: {str(e)}', exc_info=True)

def compare_csv_files(path_a: str, path_b: str, output_path: str) -> Optional[str]:
    """比较两个CSV文件并生成报告
    
    Args:
        path_a: 源CSV文件路径
        path_b: 目标CSV文件路径
        output_path: 输出结果的路径
    
    Returns:
        生成的主要报告文件路径，如果失败则返回None
    """
    # 设置日志
    log_file = setup_logging(output_path)
    logging.info('开始比较过程')

    try:
        # Step 1: 从路径A复制最新的CSV文件到路径B
        file_a = os.path.join(path_a, 'PDK_Programs.csv')
        file_b = os.path.join(path_b, 'PDK_Programs_old.csv')
        copied_file = os.path.join(path_b, 'PDK_Programs.csv')
        
        if not os.path.exists(file_a):
            logging.error(f'源文件不存在: {file_a}')
            return None
            
        shutil.copy(file_a, copied_file)
        logging.info(f'已复制 {file_a} 到 {copied_file}')
        
        # Step 2: 将两个CSV文件读入DataFrame
        try:
            df_a = pd.read_csv(copied_file, low_memory=False)
            df_b = pd.read_csv(file_b, low_memory=False)
        except Exception as e:
            logging.error(f'读取CSV文件失败: {str(e)}')
            return None
            
        logging.info(f'已加载CSV文件到DataFrame，新文件行数: {len(df_a)}，旧文件行数: {len(df_b)}')
        
        # Step 3: 只保留与比较相关的列
        relevant_columns = CONFIG.relevant_columns
        
        # 检查列是否存在
        missing_cols_a = [col for col in relevant_columns if col not in df_a.columns]
        missing_cols_b = [col for col in relevant_columns if col not in df_b.columns]
        
        if missing_cols_a or missing_cols_b:
            logging.warning(f'新文件缺少列: {missing_cols_a}, 旧文件缺少列: {missing_cols_b}')
            # 只保留两个文件都有的列
            available_columns = [col for col in relevant_columns if col in df_a.columns and col in df_b.columns]
            if not available_columns:
                logging.error('没有可用于比较的列')
                return None
            relevant_columns = available_columns
        
        df_a = df_a[relevant_columns]
        df_b = df_b[relevant_columns]
        logging.info(f'保留用于比较的相关列: {relevant_columns}')
        
        # Step 4: 对两个DataFrame进行排序，确保行顺序不影响比较
        df_a_sorted = df_a.sort_values(by=df_a.columns.tolist()).reset_index(drop=True)
        df_b_sorted = df_b.sort_values(by=df_b.columns.tolist()).reset_index(drop=True)
        logging.info('已排序DataFrame以确保一致的行顺序')
        
        # Step 5: 找出添加和删除的行
        df_added = pd.merge(df_a_sorted, df_b_sorted, how='outer', indicator=True).query("_merge == 'left_only'").drop('_merge', axis=1)
        df_removed = pd.merge(df_b_sorted, df_a_sorted, how='outer', indicator=True).query("_merge == 'left_only'").drop('_merge', axis=1)
        logging.info(f'找到 {len(df_added)} 个添加的行和 {len(df_removed)} 个删除的行')
        
        # 如果没有变化，提前返回
        if df_added.empty and df_removed.empty:
            logging.info('没有发现变化，无需生成报告')
            print('没有发现变化，无需生成报告')
            # 仍执行文件轮转
            _rotate_pdk_files(path_b)
            return None
        
        # 添加一列来指示状态并将其移到第一列
        df_added.insert(0, 'Modify_Flag', 'Added')
        df_removed.insert(0, 'Modify_Flag', 'Removed')
        logging.info('添加Modify_Flag列以指示添加和删除的行')
        
        # Step 6: 合并添加和删除的DataFrame
        df_result = pd.concat([df_added, df_removed], ignore_index=True)
        logging.info('合并添加和删除的DataFrame')
        
        # Step 7: 将结果保存到新的Excel文件
        timestamp = datetime.now().strftime('%Y%m%d')
        output_file = os.path.join(output_path, f'what_changed_{timestamp}.xlsx')
        df_result.to_excel(output_file, index=False)
        logging.info(f'将比较结果保存到 {output_file}')
        
        # Step 8: 对Excel文件应用格式
        apply_excel_formatting(output_file)
        logging.info('对比较结果Excel文件应用了格式')
        
        # Step 9: 为添加和删除的行创建数据透视表
        pivot_columns = [col for col in CONFIG.pivot_columns if col in relevant_columns]
        
        # 添加行的数据透视表
        df_added_pivot = pd.DataFrame()
        if not df_added.empty and pivot_columns:
            df_added_pivot = df_added[pivot_columns + ['Modify_Flag']].pivot_table(
                index=pivot_columns, aggfunc='size').reset_index(name='Count')
            added_pivot_output_file = os.path.join(output_path, f'added_pivot_result_{timestamp}.xlsx')
            df_added_pivot.to_excel(added_pivot_output_file, index=False)
            logging.info(f'将添加的数据透视表保存到 {added_pivot_output_file}')
            
        # 删除行的数据透视表
        df_removed_pivot = pd.DataFrame()
        if not df_removed.empty and pivot_columns:
            df_removed_pivot = df_removed[pivot_columns + ['Modify_Flag']].pivot_table(
                index=pivot_columns, aggfunc='size').reset_index(name='Count')
            removed_pivot_output_file = os.path.join(output_path, f'removed_pivot_result_{timestamp}.xlsx')
            df_removed_pivot.to_excel(removed_pivot_output_file, index=False)
            logging.info(f'将删除的数据透视表保存到 {removed_pivot_output_file}')
        
        # Step 10: 通过电子邮件发送数据透视表
        send_email_with_outlook_optimized(
            df_added_pivot.iloc[:, :7] if not df_added_pivot.empty else pd.DataFrame(), 
            df_removed_pivot.iloc[:, :7] if not df_removed_pivot.empty else pd.DataFrame(), 
            output_file
        )
        logging.info('比较完成并发送了电子邮件')
        
        # 文件轮转：old -> bak，current -> old
        _rotate_pdk_files(path_b)
        
        return output_file
            
    except Exception as e:
        logging.error(f'比较过程中出错: {str(e)}', exc_info=True)
        print(f'错误: {str(e)}')
        return None

def apply_excel_formatting(excel_file):
    """对Excel文件应用格式
    
    Args:
        excel_file: 要格式化的Excel文件路径
    """
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # 定义填充和字体样式
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF')
        
        # 设置列宽
        for col in range(1, sheet.max_column + 1 if sheet.max_column is not None else 1):
            sheet.column_dimensions[get_column_letter(col)].width = 15
            
        # 设置标题行格式
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 根据修改标志应用行格式
        for row in range(2, sheet.max_row + 1):
            modify_flag = sheet.cell(row=row, column=1).value
            if modify_flag == 'Added':
                cell = sheet.cell(row=row, column=1)
                cell.fill = green_fill
            elif modify_flag == 'Removed':
                cell = sheet.cell(row=row, column=1)
                cell.fill = red_fill
                cell.font = white_font
                
        workbook.save(excel_file)
    except Exception as e:
        logging.error(f'应用Excel格式时出错: {str(e)}')

def send_email_with_outlook_optimized(added_pivot, removed_pivot, result_file: str) -> bool:
    """发送周报邮件：固定仅向指定邮箱发送"""
    try:
        # 固定收件人，仅向本人发送
        recipients = ["qibai.chen@intel.com"]
        
        outlook_manager = OutlookManager()
        outlook = outlook_manager.get_outlook()
        
        def send_mail():
            mail = outlook.CreateItem(0)
            # 设置单一收件人
            mail.To = recipients[0]
            
            # 获取当前日期用于邮件主题
            current_date = datetime.now().strftime('%Y-%m-%d')
            mail.Subject = f'Weekly PDK Update Report - {current_date}'
            
            # 使用模板创建邮件正文
            mail.HTMLBody = HTMLEmailTemplate.create_weekly_report_body(added_pivot, removed_pivot)
            
            # 附加结果文件
            if os.path.exists(result_file):
                mail.Attachments.Add(Source=result_file)
            
            # 发送电子邮件
            mail.Send()
        
        # 使用重试机制发送邮件
        success = outlook_manager.send_email_with_retry(send_mail, CONFIG.email.retry_times, CONFIG.email.retry_delay)
        if success:
            logging.info('已发送电子邮件至 qibai.chen@intel.com')
        return success
        
    except Exception as e:
        logging.error(f'发送电子邮件失败: {str(e)}', exc_info=True)
        return False

if __name__ == "__main__":
    try:
        print("开始PDK周报生成...")
        # 运行比较
        result_file = compare_csv_files(CONFIG.paths.source_path, CONFIG.paths.target_path, CONFIG.paths.output_path)
        if result_file:
            print(f"PDK周报生成完成！主要报告文件: {result_file}")
        else:
            print("没有发现变化或处理失败")
    except Exception as e:
        print(f"程序执行过程中出现错误: {str(e)}")
        logging.error(f"程序执行过程中出现错误: {str(e)}", exc_info=True)