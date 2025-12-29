import logging
from datetime import datetime
from pathlib import Path
from typing import Tuple, List, Optional

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config_loader import AppConfig


LOGGER = logging.getLogger(__name__)


def _assemble_pareto_table(
    quantity: pd.DataFrame,
    percentages: pd.DataFrame,
    row_totals: pd.Series,
    total_percentage: pd.Series,
) -> pd.DataFrame:
    """组装Pareto表格，使用多级列索引，并在底部添加summary行"""
    import logging
    LOGGER = logging.getLogger(__name__)
    
    # 检查 quantity 是否为空
    if quantity.empty:
        LOGGER.warning("⚠️ 警告：quantity 表为空，生成的 Pareto 表将只包含 Grand Total 行（值为 0）")
        LOGGER.warning("   可能的原因：")
        LOGGER.warning("   1. 数据被过滤条件全部过滤掉了")
        LOGGER.warning("   2. 数据中没有有效的 functional_bin 或 devrevstep 值")
        LOGGER.warning("   3. 列名不匹配")
    
    data = {}
    for dev in quantity.columns:
        data[(dev, "Quantity")] = quantity[dev]
        data[(dev, "Percentage")] = percentages[dev]
    data[("Grand Total", "Quantity")] = row_totals
    data[("Grand Total", "Percentage")] = total_percentage
    table = pd.DataFrame(data)
    table.index.name = quantity.index.name if not quantity.empty else None
    if not quantity.empty:
        table = table.sort_values(("Grand Total", "Quantity"), ascending=False)
    
    # 在表格底部添加summary行（Grand Total行）
    summary_row = {}
    for dev in quantity.columns:
        # 计算每个devrevstep的总计
        col_total_qty = quantity[dev].sum()
        col_total_pct = 1.0 if col_total_qty > 0 else 0.0  # 总计行的百分比应该是100%
        summary_row[(dev, "Quantity")] = col_total_qty
        summary_row[(dev, "Percentage")] = col_total_pct
    
    # Grand Total列的总计
    grand_total_qty = row_totals.sum()
    grand_total_pct = 1.0 if grand_total_qty > 0 else 0.0
    summary_row[("Grand Total", "Quantity")] = grand_total_qty
    summary_row[("Grand Total", "Percentage")] = grand_total_pct
    
    # 将summary行添加到表格底部
    summary_df = pd.DataFrame([summary_row], index=["Grand Total"])
    table = pd.concat([table, summary_df])
    
    return table


def _write_pareto_sheet_with_style(
    writer: pd.ExcelWriter,
    pareto_table: pd.DataFrame,
    sheet_name: str,
    config: AppConfig,
) -> None:
    """写入Pareto表格到Excel，并应用样式"""
    # 写入数据（从第6行开始，前5行留给标题和自定义表头）
    pareto_table.to_excel(writer, sheet_name=sheet_name, startrow=5, index=True)
    
    workbook = writer.book
    ws = workbook[sheet_name]
    
    # 获取阈值
    red_threshold = config.processing.percentage_thresholds.get("red", 0.05)
    yellow_threshold = config.processing.percentage_thresholds.get("yellow", 0.02)
    
    # 设置样式
    _format_pareto_sheet(ws, pareto_table, sheet_name, red_threshold, yellow_threshold)


def _format_pareto_sheet(ws, pareto_table: pd.DataFrame, sheet_name: str, red_threshold: float, yellow_threshold: float) -> None:
    """格式化Pareto表格，使其像Excel透视表"""
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 清除所有合并单元格
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    
    # 删除pandas写入的表头行（第6-8行，即第6、7、8行）
    # 注意：pandas从第6行开始写入（startrow=5），所以第6、7、8行是pandas的表头
    ws.delete_rows(6, 3)  # 从第6行开始删除3行
    
    # 清除前5行的内容（自定义表头区域）
    for row in range(1, 6):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
    
    # 定义样式
    # 标题样式（浅粉色背景，深紫色字体）
    title_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 浅粉色/灰色
    title_font = Font(bold=True, color="7030A0", size=14)  # 深紫色
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 第1-2行：标题行
    title_text = sheet_name  # 例如 "Interface_bin Pareto" 或 "Functional_bin Pareto"
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title_text
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center_align
    # 合并标题单元格（跨所有列，跨2行）
    end_col_letter = get_column_letter(max_col)
    ws.merge_cells(f"A1:{end_col_letter}2")
    # 设置第2行也为标题样式
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = title_fill
        cell.border = border
    
    # 处理多级表头（第3、4、5行）
    if isinstance(pareto_table.columns, pd.MultiIndex):
        # 定义行号
        row3 = 3
        row4 = 4
        row5 = 5
        col = 2  # 从B列开始（A列是索引）
        
        # 获取所有唯一的devrevstep值
        devrevsteps = [col[0] for col in pareto_table.columns if col[0] != "Grand Total"]
        devrevsteps = list(dict.fromkeys(devrevsteps))  # 去重保持顺序
        
        # 写入"Devrevstep"主表头（合并所有devrevstep列，只跨第3行）
        devrevstep_start_col = col
        devrevstep_end_col = col
        for dev in devrevsteps:
            sub_cols = [c for c in pareto_table.columns if c[0] == dev]
            devrevstep_end_col += len(sub_cols)
        
        if devrevstep_end_col > devrevstep_start_col:
            cell = ws.cell(row=row3, column=devrevstep_start_col)
            cell.value = "Devrevstep"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            end_col_letter = get_column_letter(devrevstep_end_col - 1)
            ws.merge_cells(f"{cell.coordinate}:{end_col_letter}{row3}")  # 只跨第3行
        
        # Grand Total列（第3行，跨第3-4行，跨所有 Grand Total 列）
        gt_cols = [c for c in pareto_table.columns if c[0] == "Grand Total"]
        if gt_cols:
            cell = ws.cell(row=row3, column=devrevstep_end_col)
            cell.value = "Grand Total"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            end_col_letter = get_column_letter(devrevstep_end_col + len(gt_cols) - 1)
            # 跨第3-4行，跨所有 Grand Total 列
            ws.merge_cells(f"{cell.coordinate}:{end_col_letter}{row4}")
        
        # 第4行：子表头（具体的devrevstep值和Grand Total，以及索引列标题）
        col = 2
        
        # 索引列标题（第4行，跨第3-5行）
        index_name = pareto_table.index.name or "Index"
        cell_a4 = ws.cell(row=row4, column=1)
        cell_a4.value = index_name
        cell_a4.fill = header_fill
        cell_a4.font = header_font
        cell_a4.alignment = center_align
        cell_a4.border = border
        ws.merge_cells(f"A3:A5")  # 索引列跨第3-5行
        
        # 写入具体的devrevstep值（第4行）
        for dev in devrevsteps:
            sub_cols = [c for c in pareto_table.columns if c[0] == dev]
            span = len(sub_cols)
            
            if span > 0:
                cell = ws.cell(row=row4, column=col)
                cell.value = dev
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
                
                # 合并单元格（只在第4行合并，不跨第5行）
                if span > 1:
                    end_col_letter = get_column_letter(col + span - 1)
                    ws.merge_cells(f"{cell.coordinate}:{end_col_letter}{row4}")  # 只跨第4行
                # 如果只有一列，不需要合并
                
                col += span
        
        # Grand Total列（第4行，已经被第3行合并了，不需要单独处理）
        # 第3行的 "Grand Total" 已经跨第3-4行，跨所有 Grand Total 列（Quantity 和 Percentage）
        # 合并后的单元格会自动应用样式和居中，不需要再单独处理第4行
        
        # 第5行：子表头（Quantity和Percentage）
        col = 2
        for col_tuple in pareto_table.columns:
            cell = ws.cell(row=row5, column=col)
            cell.value = col_tuple[1]  # Quantity 或 Percentage
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = center_align
            cell.border = border
            col += 1
        
        # 索引列已经在上面合并了（A3:A5），不需要单独处理第5行
        
        data_start_row = 6
    else:
        # 单级表头的情况
        data_start_row = 2
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
    
    # 格式化数据行
    # 检查最后一行是否是"Grand Total"（summary行）
    # summary行使用与Devrevstep相同的蓝色背景
    summary_row_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 蓝色背景
    summary_row_font = Font(bold=True, color="FFFFFF", size=11)  # 白色字体
    
    for row in range(data_start_row, max_row + 1):
        # 检查是否是summary行（最后一行，且索引为"Grand Total"）
        is_summary_row = False
        if row == max_row:
            try:
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and "Grand Total" in str(cell_value):
                    is_summary_row = True
            except:
                pass
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 1:  # 索引列
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = center_align
            
            # 如果是summary行，应用特殊样式
            if is_summary_row:
                cell.fill = summary_row_fill
                cell.font = summary_row_font
    
    # 删除summary行下面的空白行（如果有的话）
    # 检查summary行是否是最后一行
    summary_row_num = None
    for row in range(data_start_row, max_row + 1):
        try:
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Grand Total" in str(cell_value):
                summary_row_num = row
                break
        except:
            pass
    
    if summary_row_num:
        # 删除summary行后面的所有空白行
        rows_to_delete = max_row - summary_row_num
        if rows_to_delete > 0:
            ws.delete_rows(summary_row_num + 1, rows_to_delete)
            max_row = summary_row_num  # 更新max_row
    
    # 格式化百分比列
    percentage_columns: List[int] = []
    col = 2
    for col_tuple in pareto_table.columns:
        if isinstance(col_tuple, tuple) and col_tuple[1] == "Percentage":
            percentage_columns.append(col)
        elif isinstance(col_tuple, str) and "Percentage" in str(col_tuple):
            percentage_columns.append(col)
        col += 1
    
    for col in percentage_columns:
        for row in range(data_start_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = "0.00%"
    
    # 识别需要特殊处理的 bin（functional bin 100 和 interface bin 1）
    # 判断是 functional bin 还是 interface bin
    is_functional_bin = "Func" in sheet_name or "Functional" in sheet_name
    is_interface_bin = "Intf" in sheet_name or "Interface" in sheet_name
    
    # 需要排除的行（bin 100 用于 functional，bin 1 用于 interface）
    excluded_rows = set()
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # 遍历数据行，识别需要特殊处理的行
    for row in range(data_start_row, max_row):
        try:
            bin_value = ws.cell(row=row, column=1).value  # A列是 bin 值
            if bin_value is not None:
                bin_str = str(bin_value).strip()
                # functional bin 100 或 interface bin 1
                if (is_functional_bin and bin_str == "100") or (is_interface_bin and bin_str == "1"):
                    excluded_rows.add(row)
                    # 直接填充绿色（对所有百分比列）
                    for col in percentage_columns:
                        cell = ws.cell(row=row, column=col)
                        cell.fill = green_fill
        except:
            pass
    
    # 应用条件格式（仅对百分比列，排除summary行和特殊bin行）
    if percentage_columns:
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
        
        # 确定数据行的范围（排除summary行）
        data_end_row = max_row - 1  # 排除最后一行（summary行）
        if data_end_row < data_start_row:
            data_end_row = data_start_row  # 如果没有数据行，至少包含起始行
        
        for col in percentage_columns:
            col_letter = get_column_letter(col)
            
            # 构建不包含排除行的范围列表
            ranges_to_format = []
            current_start = None
            
            for row in range(data_start_row, data_end_row + 1):
                if row not in excluded_rows:
                    if current_start is None:
                        current_start = row
                else:
                    # 遇到排除行，如果之前有连续范围，添加到列表
                    if current_start is not None:
                        if current_start == row - 1:
                            # 单个单元格
                            ranges_to_format.append(f"{col_letter}{current_start}")
                        else:
                            # 范围
                            ranges_to_format.append(f"{col_letter}{current_start}:{col_letter}{row - 1}")
                        current_start = None
            
            # 处理最后一段
            if current_start is not None:
                if current_start == data_end_row:
                    ranges_to_format.append(f"{col_letter}{current_start}")
                else:
                    ranges_to_format.append(f"{col_letter}{current_start}:{col_letter}{data_end_row}")
            
            # 对每个范围应用条件格式
            for range_ref in ranges_to_format:
                # 红色：大于red_threshold
                ws.conditional_formatting.add(
                    range_ref,
                    CellIsRule(operator="greaterThan", formula=[str(red_threshold)], fill=red_fill),
                )
                # 黄色：在yellow_threshold和red_threshold之间
                ws.conditional_formatting.add(
                    range_ref,
                    CellIsRule(operator="between", formula=[str(yellow_threshold), str(red_threshold)], fill=yellow_fill),
                )
                # 绿色：小于等于yellow_threshold
                ws.conditional_formatting.add(
                    range_ref,
                    CellIsRule(operator="lessThanOrEqual", formula=[str(yellow_threshold)], fill=green_fill),
                )
    
    # 设置列宽
    ws.column_dimensions["A"].width = 15
    for col in range(2, max_col + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 12


def save_report(
    report_dir: Path,
    df: pd.DataFrame,
    functional_pareto: pd.DataFrame,
    interface_pareto: Optional[pd.DataFrame],
    retest_pareto: Optional[pd.DataFrame],
    exceptions: pd.DataFrame,
    config: AppConfig,
    functional_pareto_by_step: Optional[dict] = None,
    interface_pareto_by_step: Optional[dict] = None,
    retest_pareto_by_step: Optional[dict] = None,
) -> Path:
    """保存报告到Excel文件"""
    report_name = f"Bin_Pareto_Report_{datetime.now():%Y-%m-%d}.xlsx"
    report_path = report_dir / report_name
    LOGGER.info("生成报表：%s", report_path)

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        # Functional_bin Pareto（总的，如果存在）
        if functional_pareto is not None and not functional_pareto.empty:
            _write_pareto_sheet_with_style(
                writer, functional_pareto, "Functional_bin Pareto", config
            )
        
        # 按 process_step 分组的 Functional_bin Pareto（如果存在）
        if functional_pareto_by_step:
            LOGGER.info(f"生成 {len(functional_pareto_by_step)} 个按 process_step 分组的 Functional_bin Pareto 表")
            for step_name, pareto_table in functional_pareto_by_step.items():
                if not pareto_table.empty:
                    # Excel sheet 名称限制为 31 个字符
                    sheet_name = f"FuncBin_{step_name}"[:31]
                    _write_pareto_sheet_with_style(
                        writer, pareto_table, sheet_name, config
                    )
                    LOGGER.info(f"✅ 已生成 {step_name} 的 Functional_bin Pareto 表")
        
        # Interface_bin Pareto（总的，如果存在）
        if interface_pareto is not None and not interface_pareto.empty:
            _write_pareto_sheet_with_style(
                writer, interface_pareto, "Interface_bin Pareto", config
            )
        
        # 按 process_step 分组的 Interface_bin Pareto（如果存在）
        if interface_pareto_by_step:
            LOGGER.info(f"生成 {len(interface_pareto_by_step)} 个按 process_step 分组的 Interface_bin Pareto 表")
            for step_name, pareto_table in interface_pareto_by_step.items():
                if not pareto_table.empty:
                    # Excel sheet 名称限制为 31 个字符
                    sheet_name = f"IntfBin_{step_name}"[:31]
                    _write_pareto_sheet_with_style(
                        writer, pareto_table, sheet_name, config
                    )
                    LOGGER.info(f"✅ 已生成 {step_name} 的 Interface_bin Pareto 表")
        
        # Retest Bin Pareto（总的，如果存在）
        if retest_pareto is not None and not retest_pareto.empty:
            _write_pareto_sheet_with_style(
                writer, retest_pareto, "Retest Bin Pareto", config
            )
        
        # 按 process_step 分组的 Retest Bin Pareto（如果存在）
        if retest_pareto_by_step:
            LOGGER.info(f"生成 {len(retest_pareto_by_step)} 个按 process_step 分组的 Retest Bin Pareto 表")
            for step_name, pareto_table in retest_pareto_by_step.items():
                if not pareto_table.empty:
                    # Excel sheet 名称限制为 31 个字符
                    sheet_name = f"RetestBin_{step_name}"[:31]
                    _write_pareto_sheet_with_style(
                        writer, pareto_table, sheet_name, config
                    )
                    LOGGER.info(f"✅ 已生成 {step_name} 的 Retest Bin Pareto 表")
        
        # Details
        df.to_excel(writer, sheet_name="Details", index=False)
        
        # Exceptions
        exceptions.to_excel(writer, sheet_name="Exceptions", index=False)
        
        # Correlation_UnitLevel
        correlation_table = build_correlation_unitlevel_table(df, config)
        if correlation_table is not None and not correlation_table.empty:
            correlation_table.to_excel(writer, sheet_name="Correlation_UnitLevel", index=False)
            # 格式化Correlation_UnitLevel表
            workbook = writer.book
            ws = workbook["Correlation_UnitLevel"]
            _format_correlation_sheet(ws, correlation_table)
            LOGGER.info("✅ Correlation_UnitLevel表已生成并格式化")
        else:
            LOGGER.warning("⚠️ Correlation_UnitLevel表为空，跳过生成")

    return report_path


def build_pareto_table(
    quantity: pd.DataFrame,
    percentages: pd.DataFrame,
    row_totals: pd.Series,
    total_percentage: pd.Series,
) -> pd.DataFrame:
    """构建Pareto表格"""
    return _assemble_pareto_table(quantity, percentages, row_totals, total_percentage)


def build_correlation_unitlevel_table(df: pd.DataFrame, config: AppConfig) -> Optional[pd.DataFrame]:
    """
    构建Correlation_UnitLevel表
    按process_step分组，每个process_step有独立的列、BINSWITCH和COMMENTS
    
    BINSWITCH三种情况：
    1. NO（绿色）：任意ECG与任意CCG比较，有相同的functional bin
    2. YES（黄色）：没有相同的functional bin 且没有任意ECG或CCG出现Bin 100
    3. YES（红色）：没有相同的functional bin 且有任意ECG或CCG出现Bin 100
    """
    # 检查必要的列
    required_cols = [config.fields.visual_id, config.fields.functional_bin, config.fields.lot, config.fields.process_step]
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        LOGGER.warning(f"⚠️ 缺少必要的列来生成Correlation_UnitLevel表: {missing_cols}")
        return None
    
    # 检查ECG_OR_CCG列是否存在
    ecg_col = None
    possible_ecg_cols = ["ECG_OR_CCG", "ecg_or_ccg", "ECG", "CCG"]
    for col_name in possible_ecg_cols:
        if col_name in df.columns:
            ecg_col = col_name
            break
    
    if ecg_col is None:
        LOGGER.warning("⚠️ 未找到ECG_OR_CCG列，尝试查找包含'ECG'或'CCG'的列")
        for col in df.columns:
            if 'ECG' in str(col).upper() or 'CCG' in str(col).upper():
                ecg_col = col
                LOGGER.info(f"使用列: {ecg_col} 作为ECG_OR_CCG")
                break
    
    if ecg_col is None:
        LOGGER.error("❌ 无法找到ECG_OR_CCG相关列，跳过Correlation_UnitLevel表生成")
        return None
    
    LOGGER.info(f"开始生成Correlation_UnitLevel表，使用ECG列: {ecg_col}")
    
    # 按process_step分组处理
    process_step_col = config.fields.process_step
    visual_id_col = config.fields.visual_id
    functional_bin_col = config.fields.functional_bin
    lot_col = config.fields.lot
    
    # 获取所有process_step，并按照指定顺序排序
    all_process_steps = df[process_step_col].dropna().unique().tolist()
    process_step_priority = {'CLASSHOT': 1, 'CLASSCOLD': 2}
    
    def get_process_step_sort_key(step):
        step_str = str(step).upper()
        priority = process_step_priority.get(step_str, 999)
        return (priority, step_str)
    
    process_steps = sorted(all_process_steps, key=get_process_step_sort_key)
    
    if len(process_steps) == 0:
        LOGGER.warning("⚠️ 没有有效的process_step数据，无法生成Correlation_UnitLevel表")
        return None
    
    # 获取所有唯一的VID
    all_vids = sorted(df[visual_id_col].dropna().unique())
    
    if len(all_vids) == 0:
        LOGGER.warning("⚠️ 没有有效的VID数据，无法生成Correlation_UnitLevel表")
        return None
    
    # 为每个process_step收集数据
    process_step_data = {}
    
    for step in process_steps:
        step_df = df[df[process_step_col] == step].copy()
        if step_df.empty:
            continue
        
        LOGGER.info(f"处理process_step: {step}，数据行数: {len(step_df)}")
        
        # 创建ECG_OR_CCG + Lot的组合列
        step_df['_ecg_lot'] = step_df[ecg_col].astype(str) + '_' + step_df[lot_col].astype(str)
        
        # 获取所有唯一的ECG_Lot组合
        all_ecg_lots = sorted(step_df['_ecg_lot'].dropna().unique())
        
        if len(all_ecg_lots) == 0:
            LOGGER.warning(f"process_step {step} 没有有效的ECG_Lot组合，跳过")
            continue
        
        # 为每个VID收集数据
        vid_data_dict = {}
        for vid in all_vids:
            vid_bins = {}  # 存储每个ECG_Lot对应的functional_bin值
            
            vid_df = step_df[step_df[visual_id_col] == vid]
            
            for _, row in vid_df.iterrows():
                ecg_lot = row['_ecg_lot']
                func_bin = row[functional_bin_col]
                if ecg_lot not in vid_bins:
                    vid_bins[ecg_lot] = func_bin
            
            # 分离CCG和ECG开头的列
            ccg_lots = [ecg_lot for ecg_lot in all_ecg_lots if str(ecg_lot).upper().startswith('CCG_')]
            ecg_lots = [ecg_lot for ecg_lot in all_ecg_lots if str(ecg_lot).upper().startswith('ECG_')]
            
            # 获取CCG和ECG列的bin值（排除空值和NaN）
            ccg_bin_values = []
            for ecg_lot in ccg_lots:
                bin_val = vid_bins.get(ecg_lot, '')
                # 检查是否为空值或NaN
                if bin_val != '' and not (isinstance(bin_val, float) and pd.isna(bin_val)):
                    ccg_bin_values.append(bin_val)
            
            ecg_bin_values = []
            for ecg_lot in ecg_lots:
                bin_val = vid_bins.get(ecg_lot, '')
                # 检查是否为空值或NaN
                if bin_val != '' and not (isinstance(bin_val, float) and pd.isna(bin_val)):
                    ecg_bin_values.append(bin_val)
            
            # 如果CCG或ECG列中任何一个为空，无法比较，设为NO
            if len(ccg_bin_values) == 0 or len(ecg_bin_values) == 0:
                binswitch = 'NO'
                LOGGER.debug(f"BINSWITCH=NO: VID={vid}, step={step}, 无数据比较")
            else:
                # 检查CCG的bin值是否与ECG的bin值有相同的
                has_match = False
                matched_pairs = []
                
                # 调试日志：显示所有bin值
                LOGGER.debug(f"VID={vid}, step={step}")
                LOGGER.debug(f"  CCG bin值: {ccg_bin_values} (类型: {[type(v).__name__ for v in ccg_bin_values]})")
                LOGGER.debug(f"  ECG bin值: {ecg_bin_values} (类型: {[type(v).__name__ for v in ecg_bin_values]})")
                
                for ccg_bin in ccg_bin_values:
                    for ecg_bin in ecg_bin_values:
                        # 统一转换为字符串进行比较
                        ccg_str = str(ccg_bin).strip()
                        ecg_str = str(ecg_bin).strip()
                        
                        LOGGER.debug(f"    比较: CCG='{ccg_str}' vs ECG='{ecg_str}'")
                        
                        if ccg_str == ecg_str:
                            has_match = True
                            matched_pairs.append((ccg_str, ecg_str))
                            LOGGER.debug(f"    ✓ 找到匹配!")
                            break
                    if has_match:
                        break
                
                if has_match:
                    # 情况1：有相同的functional bin → NO（绿色）
                    binswitch = 'NO'
                    LOGGER.info(f"✓ BINSWITCH=NO: VID={vid}, step={step}, 匹配的bin值: {matched_pairs}")
                else:
                    # 没有匹配，检查是否有Bin 100
                    has_bin_100 = any(str(bin_val).strip() == '100' for bin_val in ccg_bin_values + ecg_bin_values)
                    
                    if has_bin_100:
                        # 情况3：没有匹配且有Bin 100 → YES_RED（红色）
                        binswitch = 'YES_RED'
                        LOGGER.warning(f"✗ BINSWITCH=YES_RED: VID={vid}, step={step}")
                        LOGGER.warning(f"    CCG bin值: {ccg_bin_values}")
                        LOGGER.warning(f"    ECG bin值: {ecg_bin_values}")
                        LOGGER.warning(f"    无匹配，但有Bin 100")
                    else:
                        # 情况2：没有匹配且没有Bin 100 → YES_YELLOW（黄色）
                        binswitch = 'YES_YELLOW'
                        LOGGER.info(f"BINSWITCH=YES_YELLOW: VID={vid}, step={step}, 无匹配且无Bin 100")
            
            vid_data_dict[vid] = {
                'bins': vid_bins,
                'binswitch': binswitch
            }
        
        process_step_data[step] = {
            'ecg_lots': all_ecg_lots,
            'vid_data': vid_data_dict
        }
        LOGGER.info(f"✅ process_step {step} 数据处理完成，ECG_Lot列数: {len(all_ecg_lots)}")
    
    if not process_step_data:
        LOGGER.warning("⚠️ 所有process_step的数据都为空，无法生成Correlation_UnitLevel表")
        return None
    
    # 构建最终的表结构
    # 列顺序：VID, [process_step1的列], BINSWITCH(step1), COMMENTS(step1), ...
    result_data = []
    ordered_columns = ['VID']
    # 创建列名到step的映射，避免同一个ECG_Lot在多个step中时取错数据
    col_to_step = {}
    # 创建唯一列名到原始列名的映射（用于格式化Excel时显示）
    unique_col_to_original = {}
    
    # 为每个process_step添加列
    for step in process_steps:
        if step not in process_step_data:
            continue
        
        step_info = process_step_data[step]
        ecg_lots = step_info['ecg_lots']
        
        # 添加该process_step的所有ECG_Lot列，并记录它们属于哪个step
        # 为了支持重复的列名，我们为每个列添加step前缀使其唯一
        for ecg_lot in ecg_lots:
            unique_col_name = f"{step}_{ecg_lot}"  # 创建唯一列名
            ordered_columns.append(unique_col_name)
            col_to_step[unique_col_name] = step  # 记录每个列属于哪个step
            unique_col_to_original[unique_col_name] = ecg_lot  # 记录原始列名
        
        # 添加该process_step的BINSWITCH列
        ordered_columns.append(f'BINSWITCH_{step}')
        # 添加该process_step的COMMENTS列
        ordered_columns.append(f'COMMENTS_{step}')
    
    # 构建数据行
    for vid in all_vids:
        row_data = {}
        
        for col_name in ordered_columns:
            if col_name == 'VID':
                row_data[col_name] = vid
            elif col_name.startswith('COMMENTS_'):
                row_data[col_name] = ''
            elif col_name.startswith('BINSWITCH_'):
                step = col_name.replace('BINSWITCH_', '')
                if step in process_step_data:
                    vid_info = process_step_data[step]['vid_data'].get(vid, {'binswitch': 'NO'})
                    row_data[col_name] = vid_info['binswitch']
                else:
                    row_data[col_name] = 'NO'
            else:
                # 这是ECG_Lot列（可能是带step前缀的唯一列名），使用映射找到正确的step和原始列名
                if col_name in col_to_step:
                    step = col_to_step[col_name]
                    original_col_name = unique_col_to_original.get(col_name, col_name)
                    if step in process_step_data:
                        vid_info = process_step_data[step]['vid_data'].get(vid, {'bins': {}})
                        bin_value = vid_info['bins'].get(original_col_name, '')
                        row_data[col_name] = bin_value
                    else:
                        row_data[col_name] = ''
                else:
                    row_data[col_name] = ''
        
        result_data.append(row_data)
    
    # 创建DataFrame
    result_table = pd.DataFrame(result_data, columns=ordered_columns)
    
    LOGGER.info(f"✅ Correlation_UnitLevel表生成完成，总行数: {len(result_table)}，总列数: {len(result_table.columns)}")
    return result_table


def _format_correlation_sheet(ws, correlation_table: pd.DataFrame) -> None:
    """格式化Correlation_UnitLevel表，支持多级表头"""
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    process_step_colors = {
        'CLASSHOT': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'CLASSCOLD': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    }
    default_process_step_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    ecg_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    ccg_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # 三种填充颜色
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    
    # 分析列结构
    columns = correlation_table.columns.tolist()
    
    # 识别BINSWITCH列和COMMENTS列，重新构建process_step_groups
    binswitch_cols = {}
    comments_cols = {}
    process_step_groups = {}
    
    # 第一遍：找到所有BINSWITCH和COMMENTS列的位置
    for idx, col_name in enumerate(columns):
        col_name_str = str(col_name)
        
        if col_name_str.startswith('BINSWITCH_'):
            process_step = col_name_str.replace('BINSWITCH_', '')
            binswitch_col = idx + 1
            binswitch_cols[binswitch_col] = process_step
        
        elif col_name_str.startswith('COMMENTS_'):
            process_step = col_name_str.replace('COMMENTS_', '')
            comments_col = idx + 1
            comments_cols[comments_col] = process_step
    
    # 第二遍：根据BINSWITCH列的位置确定每个process_step的列范围
    prev_end_col = 1  # VID列
    for binswitch_col in sorted(binswitch_cols.keys()):
        process_step = binswitch_cols[binswitch_col]
        start_col = prev_end_col + 1
        end_col = binswitch_col - 1
        
        if start_col <= end_col:
            process_step_groups[process_step] = [start_col, end_col]
        
        # 更新prev_end_col为当前process_step的COMMENTS列
        # 找到对应的COMMENTS列
        for comments_col in comments_cols.keys():
            if comments_cols[comments_col] == process_step:
                prev_end_col = comments_col
                break
    
    # 删除pandas写入的表头，然后插入两行新表头
    ws.delete_rows(1, 1)
    ws.insert_rows(1, 2)  # 在第1行位置插入2行
    
    # 创建两行表头
    row1 = 1
    row2 = 2
    max_row = ws.max_row  # 更新max_row
    
    # VID列（跨2行）
    vid_cell = ws.cell(row=row1, column=1)
    vid_cell.value = "VID"
    vid_cell.fill = header_fill
    vid_cell.font = header_font
    vid_cell.alignment = center_align
    vid_cell.border = border
    ws.merge_cells(f"A{row1}:A{row2}")
    
    # 获取process_step顺序
    process_step_order = []
    for idx, col_name in enumerate(columns):
        if str(col_name).startswith('BINSWITCH_'):
            process_step = str(col_name).replace('BINSWITCH_', '')
            if process_step not in process_step_order:
                process_step_order.append(process_step)
    
    # 为每个process_step创建表头
    for process_step in process_step_order:
        if process_step not in process_step_groups:
            continue
        
        start_col, end_col = process_step_groups[process_step]
        
        # process_step表头（第1行）
        step_fill = process_step_colors.get(process_step.upper(), default_process_step_fill)
        step_cell = ws.cell(row=row1, column=start_col)
        step_cell.value = process_step
        step_cell.fill = step_fill
        step_cell.font = Font(bold=True, color="000000", size=11)
        step_cell.alignment = center_align
        step_cell.border = border
        
        # 合并process_step单元格
        if start_col < end_col:
            ws.merge_cells(f"{get_column_letter(start_col)}{row1}:{get_column_letter(end_col)}{row1}")
        
        # 找到对应的BINSWITCH列和COMMENTS列
        binswitch_col = None
        comments_col = None
        for col_idx, col_name in enumerate(columns):
            if str(col_name) == f'BINSWITCH_{process_step}':
                binswitch_col = col_idx + 1
            elif str(col_name) == f'COMMENTS_{process_step}':
                comments_col = col_idx + 1
        
        # BINSWITCH列（跨2行）
        if binswitch_col:
            binswitch_cell = ws.cell(row=row1, column=binswitch_col)
            binswitch_cell.value = "BINSWITCH"
            binswitch_cell.fill = step_fill
            binswitch_cell.font = Font(bold=True, color="000000", size=11)
            binswitch_cell.alignment = center_align
            binswitch_cell.border = border
            ws.merge_cells(f"{get_column_letter(binswitch_col)}{row1}:{get_column_letter(binswitch_col)}{row2}")
        
        # COMMENTS列（跨2行）
        if comments_col:
            comments_cell = ws.cell(row=row1, column=comments_col)
            comments_cell.value = "COMMENTS"
            comments_cell.fill = header_fill
            comments_cell.font = header_font
            comments_cell.alignment = center_align
            comments_cell.border = border
            ws.merge_cells(f"{get_column_letter(comments_col)}{row1}:{get_column_letter(comments_col)}{row2}")
    
    # 创建第2行：具体列名
    for col_idx, col_name in enumerate(columns):
        col_num = col_idx + 1
        cell = ws.cell(row=row2, column=col_num)
        
        if col_name == 'VID' or str(col_name).startswith('BINSWITCH_') or str(col_name).startswith('COMMENTS_'):
            continue
        else:
            # ECG_Lot列名（可能是带step前缀的唯一列名，需要提取原始列名）
            col_name_str = str(col_name)
            # 检查是否是带step前缀的列名（格式：STEP_ECG_LOT）
            original_col_name = col_name_str
            for step in process_step_order:
                if col_name_str.startswith(f"{step}_"):
                    original_col_name = col_name_str[len(f"{step}_"):]
                    break
            
            cell.value = original_col_name
            original_col_name_upper = original_col_name.upper()
            if original_col_name_upper.startswith('ECG_'):
                cell.fill = ecg_header_fill
            elif original_col_name_upper.startswith('CCG_'):
                cell.fill = ccg_header_fill
            else:
                cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
    
    # 格式化数据行
    data_start_row = 3
    for row in range(data_start_row, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
            if col == 1:  # VID列
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            
            # 如果是BINSWITCH列，根据值设置颜色
            if col in binswitch_cols:
                cell_value = str(cell.value).strip().upper() if cell.value else ''
                if cell_value == 'NO':
                    cell.fill = green_fill
                    cell.value = 'NO'
                elif cell_value == 'YES_YELLOW':
                    cell.fill = yellow_fill
                    cell.value = 'YES'
                elif cell_value == 'YES_RED':
                    cell.fill = red_fill
                    cell.value = 'YES'
    
    # 设置列宽
    ws.column_dimensions["A"].width = 20
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 冻结第一列和前两行
    ws.freeze_panes = "B3"
