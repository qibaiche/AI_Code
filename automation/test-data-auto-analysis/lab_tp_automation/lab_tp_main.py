"""
Lab TP Performance 自动化主程序
"""
import sys
from pathlib import Path

# 添加父目录到路径，以便导入 prd_lot_automation 模块
sys.path.insert(0, str(Path(__file__).parent.parent))

import logging
import pandas as pd
from datetime import datetime

from prd_lot_automation.config_loader import load_config
from prd_lot_automation.lot_reader import read_lots
from prd_lot_automation.spf_runner import SQLPathFinderRunner
from prd_lot_automation.close_sqlpathfinder import close_sqlpathfinder


LOGGER = logging.getLogger(__name__)


def configure_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / "lab_tp_automation.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(log_file, encoding="utf-8"),
        ],
    )


def process_lab_data(csv_path: Path, config) -> pd.DataFrame:
    """处理 Lab TP 数据"""
    LOGGER.info("读取 Lab TP 数据：%s", csv_path)
    df = pd.read_csv(csv_path)
    
    if df.empty:
        raise ValueError(f"{csv_path} 为空")
    
    LOGGER.info("数据集形状: %s", df.shape)
    LOGGER.debug("列名: %s", df.columns.tolist())
    
    # 根据 VG2 文件，实际的列名应该是：
    # facility, operation, sub_flow_step, devrevstep, program_name,
    # Total_Tested, Teste_Good, Yield, TTG, ETT, RCS, Recovery_Rate
    
    # 创建列名映射（处理大小写和下划线差异）
    col_mapping = {}
    for col in df.columns:
        col_lower = col.lower()
        if 'facility' in col_lower:
            col_mapping[col] = 'Facility'
        elif 'operation' in col_lower:
            col_mapping[col] = 'Operation'
        elif 'sub_flow' in col_lower or 'sub flow' in col_lower:
            col_mapping[col] = 'Sub Flow Step'
        elif 'devrevstep' in col_lower:
            col_mapping[col] = 'Devrevstep'
        elif 'program' in col_lower and 'name' in col_lower:
            col_mapping[col] = 'Program Name'
        elif 'total' in col_lower and 'test' in col_lower:
            col_mapping[col] = 'Total Tested'
        elif 'teste' in col_lower and 'good' in col_lower:
            col_mapping[col] = 'Tested Good'
        elif col_lower == 'yield':
            col_mapping[col] = 'Yield'
        elif col_lower == 'ttg':
            col_mapping[col] = 'TTG'
        elif col_lower == 'ett':
            col_mapping[col] = 'ETT'
        elif col_lower == 'rcs':
            col_mapping[col] = 'RCS'
        elif 'recovery' in col_lower:
            col_mapping[col] = 'Recovery Rate'
    
    LOGGER.info(f"列名映射: {col_mapping}")
    
    # 选择需要的列
    target_cols = ['Facility', 'Operation', 'Sub Flow Step', 'Devrevstep', 'Program Name',
                   'Total Tested', 'Tested Good', 'Yield', 'TTG', 'ETT', 'RCS', 'Recovery Rate']
    
    # 检查是否所有目标列都有映射
    missing_targets = [col for col in target_cols if col not in col_mapping.values()]
    if missing_targets:
        LOGGER.error(f"缺少目标列: {missing_targets}")
        LOGGER.error(f"实际列名: {df.columns.tolist()}")
        LOGGER.error(f"列名映射: {col_mapping}")
        raise KeyError(f"无法找到必需的列: {missing_targets}")
    
    # 反向映射：目标列 -> 原始列
    reverse_mapping = {v: k for k, v in col_mapping.items()}
    
    # 选择并重命名列
    result_df = df[[reverse_mapping[col] for col in target_cols]].copy()
    result_df.columns = target_cols
    
    # 去重（因为SQL可能返回重复行）
    result_df = result_df.drop_duplicates()
    
    # 按 Sub Flow Step, Devrevstep, Program Name 排序
    result_df = result_df.sort_values(['Sub Flow Step', 'Devrevstep', 'Program Name'])
    
    LOGGER.info("处理后数据形状: %s", result_df.shape)
    
    return result_df


def save_report(report_dir: Path, df: pd.DataFrame) -> Path:
    """保存报告到 Excel"""
    report_name = f"Lab_TP_Performance_{datetime.now():%Y%m%d}.xlsx"
    report_path = report_dir / report_name
    LOGGER.info("生成报表：%s", report_path)
    
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Lot Level Performance", index=False)
        
        # 格式化表格
        workbook = writer.book
        ws = workbook["Lot Level Performance"]
        
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 定义样式
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 绿色
        header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        
        # 格式化表头
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # 格式化数据行
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center_align
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10  # Facility
        ws.column_dimensions['B'].width = 10  # Operation
        ws.column_dimensions['C'].width = 15  # Sub Flow Step
        ws.column_dimensions['D'].width = 15  # Devrevstep
        ws.column_dimensions['E'].width = 25  # Program Name
        ws.column_dimensions['F'].width = 12  # Total Tested
        ws.column_dimensions['G'].width = 12  # Tested Good
        ws.column_dimensions['H'].width = 10  # Yield
        ws.column_dimensions['I'].width = 10  # TTG
        ws.column_dimensions['J'].width = 10  # ETT
        ws.column_dimensions['K'].width = 10  # RCS
        ws.column_dimensions['L'].width = 15  # Recovery Rate
    
    return report_path


def run_lab_tp_pipeline(config_path: Path) -> Path:
    """运行 Lab TP Performance 流程"""
    config = load_config(config_path)
    configure_logging(config.paths.log_dir)
    LOGGER.info("Lab TP Performance 自动化开始")
    
    # 读取 LOT 列表
    lots = read_lots(config.paths.lots_file)
    LOGGER.info(f"读取到 {len(lots)} 个 LOT")
    
    # 运行 SQLPathFinder
    runner = SQLPathFinderRunner(config)
    runner.execute(lots)
    csv_path = runner.wait_for_output()
    
    # 处理数据
    df = process_lab_data(csv_path, config)
    
    # 保存报告
    report_path = save_report(config.paths.report_dir, df)
    
    LOGGER.info("流程完成：%s", report_path)
    
    # 关闭 SQLPathFinder
    LOGGER.info("关闭 SQLPathFinder...")
    close_sqlpathfinder(config.ui.main_window_title)
    
    return report_path


def main():
    config_path = Path(__file__).parent / "config.yaml"
    run_lab_tp_pipeline(config_path)


if __name__ == "__main__":
    main()

