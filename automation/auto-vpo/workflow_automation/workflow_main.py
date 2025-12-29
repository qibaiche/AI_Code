"""主工作流控制器"""
import logging
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Optional
import pandas as pd

try:
    import win32gui
except ImportError:
    win32gui = None

from .config_loader import load_config, WorkflowConfig
from .data_reader import read_excel_file, save_result_excel, validate_data
from .mole_submitter import MoleSubmitter
from .spark_submitter import SparkSubmitter
from .gts_submitter import GTSSubmitter
from .mole_config_ui import show_mole_config_ui
from .utils.keyboard_listener import start_global_listener, is_esc_pressed, stop_global_listener

LOGGER = logging.getLogger(__name__)


class WorkflowError(Exception):
    """工作流异常"""
    pass


class WorkflowController:
    """工作流控制器"""
    
    def __init__(self, config: WorkflowConfig):
        self.config = config
        # 创建本次运行的工作目录（带时间戳）
        self.work_dir = self._create_work_directory()
        # 使用 MIR 目录作为 debug 目录（简化结构）
        debug_dir = self.work_subdirs.get('mir', self.work_dir / '01_MIR')
        self.mole_submitter = MoleSubmitter(config.mole, debug_dir=debug_dir)
        self.spark_submitter = SparkSubmitter(config.spark, debug_dir=debug_dir)
        self.gts_submitter = GTSSubmitter(config.gts, debug_dir=debug_dir)
        self.results = []
        self.errors = []
        self.last_mir_result_file = None
        self.unit_comparison_details = []  # 存储详细的unit对比信息
        self.available_units_export_file = None  # 存储第一步验证的导出文件路径
        self.units_validation_comparison_file = None  # 存储验证比较文件路径
        # 创建合并Excel文件路径（在生成时就写入，而不是最后合并）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.merged_validation_file = self.work_subdirs['mir'] / f"Merged_Validation_Table_{timestamp}.xlsx"
        self.summary_table_file = None  # 存储汇总表文件路径（只包含 Source Lot, Part Type, Quantity, MIR）
        self.merged_excel_writer = None  # 用于保持Excel文件打开状态
        
        # 启动键盘监听器（ESC 键停止）
        def on_escape():
            """ESC 键按下时的处理"""
            LOGGER.warning("⚠️ 检测到 ESC 键，正在停止程序...")
            self._cleanup_on_exit()
        
        start_global_listener(on_escape)
        LOGGER.info("💡 提示：按 ESC 键可随时停止程序")
    
    def _cleanup_on_exit(self):
        """退出时的清理工作"""
        try:
            # 关闭浏览器
            if hasattr(self.spark_submitter, '_driver') and self.spark_submitter._driver:
                try:
                    self.spark_submitter._close_driver()
                except:
                    pass
            if hasattr(self.gts_submitter, 'driver') and self.gts_submitter.driver:
                try:
                    self.gts_submitter._close_browser()
                except:
                    pass
        except:
            pass
        finally:
            # 真正退出程序
            import sys
            LOGGER.warning("程序正在退出...")
            sys.exit(0)
    
    def _create_work_directory(self) -> Path:
        """
        创建工作目录，直接在output目录下创建三个子文件夹
        
        Returns:
            工作目录路径（output_dir）
        """
        work_dir = self.config.paths.output_dir
        
        # 直接在output目录下创建子目录（01_MIR、02_SPARK、03_GTS）
        subdirs = {
            'mir': work_dir / '01_MIR',
            'spark': work_dir / '02_SPARK',
            'gts': work_dir / '03_GTS'
        }
        
        for subdir in subdirs.values():
            subdir.mkdir(parents=True, exist_ok=True)
        
        self.work_subdirs = subdirs
        LOGGER.info(f"📁 工作目录: {work_dir}")
        LOGGER.info(f"   子目录结构:")
        LOGGER.info(f"     - 01_MIR: {subdirs['mir']}")
        LOGGER.info(f"     - 02_SPARK: {subdirs['spark']}")
        LOGGER.info(f"     - 03_GTS: {subdirs['gts']}")
        
        return work_dir
    
    def run_workflow(self, excel_file_path: str | Path) -> Path | None:
        """
        运行完整的工作流
        
        Args:
            excel_file_path: Excel文件路径
        
        Returns:
            输出Excel文件路径
        
        Raises:
            WorkflowError: 如果工作流执行失败
        """
        excel_file_path = Path(excel_file_path)
        
        LOGGER.info("=" * 80)
        LOGGER.info("开始执行自动化工作流")
        LOGGER.info(f"输入文件: {excel_file_path}")
        LOGGER.info("=" * 80)
        
        start_time = datetime.now()
        
        try:
            # 检查 ESC 键
            if is_esc_pressed():
                LOGGER.warning("⚠️ 程序已停止（ESC 键）")
                return None
            
            # 步骤0: 预先启动Mole工具
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 0/5: 启动Mole工具")
            LOGGER.info("=" * 80)
            self._step_start_mole()
            
            # 检查 ESC 键
            if is_esc_pressed():
                LOGGER.warning("⚠️ 程序已停止（ESC 键）")
                return None
            
            # 步骤1: 读取文件（Excel或CSV）
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 1/5: 读取source lot文件")
            LOGGER.info("=" * 80)
            df = self._step_read_excel(excel_file_path)
            
            # 检查 ESC 键
            if is_esc_pressed():
                LOGGER.warning("⚠️ 程序已停止（ESC 键）")
                return None
            
            # 步骤2: 提交MIR数据到Mole工具（循环处理所有行）
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 2/5: 提交MIR数据到Mole工具（循环处理所有行）")
            LOGGER.info("=" * 80)
            # 检查 ESC 键
            if is_esc_pressed():
                LOGGER.warning("⚠️ 程序已停止（ESC 键）")
                return None
            # 获取Source Lot文件路径
            source_lot_file_path = self._get_source_lot_file_path(excel_file_path)
            self._step_submit_to_mole(df, source_lot_file_path)
            
            # 检查 ESC 键
            if is_esc_pressed():
                LOGGER.warning("⚠️ 程序已停止（ESC 键）")
                return None
            
            # 步骤3: 提交VPO数据到Spark网页
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 3/5: 提交VPO数据到Spark网页")
            LOGGER.info("=" * 80)
            self._step_submit_to_spark(df)
            
            # 检查 ESC 键
            if is_esc_pressed():
                LOGGER.warning("⚠️ 程序已停止（ESC 键）")
                return None
            
            # 步骤4: 生成GTS填充文件
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 4/5: 生成GTS填充文件")
            LOGGER.info("=" * 80)
            self._step_generate_gts_file()
            
            # 检查 ESC 键
            if is_esc_pressed():
                LOGGER.warning("⚠️ 程序已停止（ESC 键）")
                return None
            
            # 步骤5: 自动填充并提交GTS
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 5/5: 自动填充并提交GTS")
            LOGGER.info("=" * 80)
            self._step_submit_to_gts()
            
            # 保存结果
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("保存处理结果")
            LOGGER.info("=" * 80)
            output_path = self._step_save_results(df)
            
            # 计算执行时间
            elapsed_time = (datetime.now() - start_time).total_seconds()
            
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("✅ 完整工作流执行成功！")
            LOGGER.info("=" * 80)
            LOGGER.info("已完成所有步骤:")
            LOGGER.info("  ✓ Mole: MIR 数据已提交")
            LOGGER.info("  ✓ Spark: VPO 数据已提交")
            LOGGER.info("  ✓ GTS: 填充文件已生成并提交")
            LOGGER.info(f"执行时间: {elapsed_time:.2f} 秒")
            if output_path:
                LOGGER.info(f"输出文件: {output_path}")
            else:
                LOGGER.info("注意: MIR结果已保存为CSV文件")
            LOGGER.info(f"📁 本次运行的所有文件保存在: {self.work_dir}")
            LOGGER.info("   文件分类:")
            LOGGER.info(f"     - MIR结果: {self.work_subdirs['mir'].name}")
            LOGGER.info(f"     - Spark文件: {self.work_subdirs['spark'].name}")
            LOGGER.info(f"     - GTS文件: {self.work_subdirs['gts'].name}")
            LOGGER.info("=" * 80)
            
            return output_path
            
        except Exception as e:
            elapsed_time = (datetime.now() - start_time).total_seconds()
            error_msg = f"工作流执行失败: {e}"
            LOGGER.error("\n" + "=" * 80)
            LOGGER.error(f"❌ {error_msg}")
            LOGGER.error(f"执行时间: {elapsed_time:.2f} 秒")
            LOGGER.error("=" * 80)
            LOGGER.error(traceback.format_exc())
            raise WorkflowError(error_msg) from e
    
    def run_mole_only(self, excel_file_path: str | Path) -> Path | None:
        """
        仅运行Mole步骤（不执行Spark/GTS）
        
        Args:
            excel_file_path: source lot文件路径
        
        Returns:
            最新的MIR结果文件路径（如果生成）
        """
        excel_file_path = Path(excel_file_path)
        
        LOGGER.info("=" * 80)
        LOGGER.info("开始执行Mole-only工作流（Spark/GTS已跳过）")
        LOGGER.info(f"输入文件: {excel_file_path}")
        LOGGER.info("=" * 80)
        
        start_time = datetime.now()
        
        try:
            # 步骤-1: 显示配置UI
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 -1/3: 配置Mole参数")
            LOGGER.info("=" * 80)
            
            # 获取config.yaml路径
            config_path = Path(__file__).parent / "config.yaml"
            
            # 显示配置UI
            LOGGER.info("显示Mole配置界面...")
            ui_config = show_mole_config_ui(config_path)
            
            if ui_config is None:
                LOGGER.warning("用户取消了配置，工作流终止")
                return None
            
            LOGGER.info(f"用户配置: {ui_config}")
            
            # 更新mole_submitter的配置
            self.mole_submitter.config.search_mode = ui_config.get('search_mode', 'vpos')
            self.mole_submitter.config.ui_config = ui_config
            
            # 保存UI配置到实例变量，供Spark提交时使用
            self.ui_config = ui_config
            
            # 步骤0: 预先启动Mole工具
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 0/3: 启动Mole工具")
            LOGGER.info("=" * 80)
            self._step_start_mole()
            
            # 步骤1: 读取文件（Excel或CSV）
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 1/3: 读取source lot文件")
            LOGGER.info("=" * 80)
            df = self._step_read_excel(excel_file_path)
            
            # 步骤2: 提交MIR数据到Mole工具（循环处理所有行）
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("步骤 2/3: 提交MIR数据到Mole工具（循环处理所有行）")
            LOGGER.info("=" * 80)
            source_lot_file_path = self._get_source_lot_file_path(excel_file_path)
            self._step_submit_to_mole(df, source_lot_file_path, ui_config)
            
            # 步骤2.5: 生成合并文件（MIR结果 + For Spark.csv）
            if self.last_mir_result_file and self.last_mir_result_file.exists():
                LOGGER.info("\n" + "=" * 80)
                LOGGER.info("步骤 2.5/3: 生成合并文件（MIR结果 + For Spark.csv）")
                LOGGER.info("=" * 80)
                try:
                    # 读取MIR结果文件
                    mir_df = read_excel_file(self.last_mir_result_file)
                    
                    # 查找 For Spark.csv 文件
                    base_dir = Path(__file__).parent.parent
                    possible_spark_config_paths = [
                        base_dir / "input" / "For Spark.csv",
                        base_dir / "For Spark.csv",
                        self.config.paths.input_dir / "For Spark.csv"
                    ]
                    
                    spark_config_file = None
                    for path in possible_spark_config_paths:
                        if path.exists():
                            spark_config_file = path
                            break
                    
                    # 合并文件
                    merged_df = self._merge_mir_with_spark_config(mir_df, spark_config_file)
                    
                    # 保存合并文件到 output 根目录
                    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                    merged_file = self.work_subdirs['spark'] / f"MIR_Results_For_Spark_{date_str}.xlsx"
                    try:
                        merged_df.to_excel(merged_file, index=False, engine='openpyxl')
                        LOGGER.info(f"✅ 已生成合并文件: {merged_file}")
                        LOGGER.info(f"   包含 {len(merged_df)} 行数据")
                        LOGGER.info(f"   列: {merged_df.columns.tolist()}")
                    except Exception as e:
                        # 如果Excel保存失败，尝试保存为CSV
                        LOGGER.warning(f"保存Excel文件失败: {e}，尝试保存为CSV格式...")
                        merged_file = self.work_subdirs['spark'] / f"MIR_Results_For_Spark_{date_str}.csv"
                        merged_df.to_csv(merged_file, index=False, encoding='utf-8-sig')
                        LOGGER.info(f"✅ 已生成合并文件: {merged_file} (CSV格式)")
                except Exception as e:
                    LOGGER.warning(f"生成合并文件时出错: {e}，继续执行...")
                    LOGGER.debug(traceback.format_exc())
            
            elapsed_time = (datetime.now() - start_time).total_seconds()
            output_path = self.last_mir_result_file
            
            LOGGER.info("\n" + "=" * 80)
            LOGGER.info("✅ Mole-only工作流执行成功（已跳过Spark/GTS）")
            LOGGER.info(f"执行时间: {elapsed_time:.2f} 秒")
            if output_path:
                LOGGER.info(f"输出文件: {output_path}")
            else:
                LOGGER.info("注意: 未获取到MIR结果文件路径")
            
            # 查找并显示合并文件
            merged_files = sorted(self.work_subdirs['spark'].glob("MIR_Results_For_Spark_*.xlsx"), reverse=True)
            merged_files.extend(sorted(self.work_subdirs['spark'].glob("MIR_Results_For_Spark_*.csv"), reverse=True))
            if merged_files:
                latest_merged = merged_files[0]
                LOGGER.info(f"📊 合并文件（MIR + For Spark.csv）: {latest_merged}")
                LOGGER.info(f"   位置: {latest_merged.parent}")
            
            LOGGER.info(f"📁 本次运行的所有文件保存在: {self.work_dir}")
            LOGGER.info("=" * 80)
            
            return output_path
            
        except Exception as e:
            elapsed_time = (datetime.now() - start_time).total_seconds()
            error_msg = f"Mole-only工作流执行失败: {e}"
            LOGGER.error("\n" + "=" * 80)
            LOGGER.error(f"❌ {error_msg}")
            LOGGER.error(f"执行时间: {elapsed_time:.2f} 秒")
            LOGGER.error("=" * 80)
            LOGGER.error(traceback.format_exc())
            raise WorkflowError(error_msg) from e
    
    def _get_source_lot_file_path(self, excel_file_path: Path) -> Path:
        """获取Source Lot文件路径"""
        # 优先使用配置文件中的路径
        if self.config.paths.source_lot_file and self.config.paths.source_lot_file.exists():
            return self.config.paths.source_lot_file
        
        # 父目录（Auto VPO根目录）
        parent_dir = excel_file_path.parent
        
        # 可能的文件名列表
        possible_names = [
            "Source Lot.csv",
            "Source Lot.xlsx",
            "Source Lot.xls",
            "source lot.csv",
            "source lot.xlsx",
            "source lot.xls",
        ]
        
        # 优先在 input/ 目录下查找
        input_dir = parent_dir / "input"
        if input_dir.exists():
            for filename in possible_names:
                file_path = input_dir / filename
                if file_path.exists():
                    LOGGER.info(f"在input目录找到Source Lot文件: {file_path}")
                    return file_path
        
        # 在父目录（Auto VPO根目录）中查找
        for filename in possible_names:
            file_path = parent_dir / filename
            if file_path.exists():
                LOGGER.info(f"在根目录找到Source Lot文件: {file_path}")
                return file_path
        
        raise WorkflowError(f"未找到Source Lot文件。请确保文件存在于以下位置之一:\n  - {input_dir if input_dir.exists() else parent_dir / 'input'}\n  - {parent_dir}")
    
    def _save_all_mir_results(self, source_lot_file_path: Path, mir_results: list) -> None:
        """
        保存所有MIR结果到CSV文件
        
        Args:
            source_lot_file_path: Source Lot文件路径
            mir_results: MIR结果列表，每个元素是一个字典，包含原始行数据+MIR列
        """
        try:
            if not mir_results:
                LOGGER.warning("没有MIR结果需要保存")
                return
            
            # 创建DataFrame
            result_df = pd.DataFrame(mir_results)
            
            # 统一列名：将 Source 或 SourceLot 重命名为 Source Lot
            column_mapping = {}
            for col in result_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCE', 'SOURCELOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS'] and col != 'Source Lot':
                    column_mapping[col] = 'Source Lot'
                    LOGGER.info(f"将列 '{col}' 重命名为 'Source Lot'")
            
            if column_mapping:
                result_df = result_df.rename(columns=column_mapping)
            
            # 查找SourceLot列（不区分大小写）
            source_lot_col = None
            for col in result_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS', 'SOURCE']:
                    source_lot_col = col
                    break
            
            # 如果找到了 Source Lot 列，确保列名为 "Source Lot"
            if source_lot_col and source_lot_col != 'Source Lot':
                result_df = result_df.rename(columns={source_lot_col: 'Source Lot'})
                source_lot_col = 'Source Lot'
                LOGGER.info(f"已统一列名为 'Source Lot'")
            
            # 如果找到SourceLot列，按SourceLot分组，确保相同SourceLot的行放在一起
            # 同时保持source lot表的原始顺序（相同SourceLot首次出现的顺序）
            if source_lot_col:
                LOGGER.info(f"按SourceLot列 '{source_lot_col}' 分组，确保相同SourceLot的MIR放在一起...")
                
                # 添加原始索引列，用于保持原始顺序
                result_df['_original_index'] = range(len(result_df))
                
                # 记录每个SourceLot首次出现的索引
                first_occurrence = {}
                for idx, source_lot_value in enumerate(result_df[source_lot_col]):
                    if pd.notna(source_lot_value):
                        source_lot_str = str(source_lot_value).strip()
                        if source_lot_str and source_lot_str not in first_occurrence:
                            first_occurrence[source_lot_str] = idx
                
                # 创建分组键：SourceLot首次出现的索引 + SourceLot值
                def get_group_key(row):
                    source_lot_value = row[source_lot_col]
                    if pd.isna(source_lot_value):
                        return (float('inf'), '')  # NaN值放在最后
                    source_lot_str = str(source_lot_value).strip()
                    first_idx = first_occurrence.get(source_lot_str, float('inf'))
                    return (first_idx, source_lot_str)
                
                # 创建分组键列
                result_df['_group_key'] = result_df.apply(get_group_key, axis=1)
                
                # 按分组键和原始索引排序
                result_df = result_df.sort_values(by=['_group_key', '_original_index'], na_position='last')
                
                # 删除临时列
                result_df = result_df.drop(columns=['_original_index', '_group_key'])
                
                LOGGER.info(f"✅ 已按SourceLot分组，相同SourceLot的MIR已放在一起（保持source lot表的原始顺序）")
            else:
                LOGGER.warning("未找到SourceLot列，保持原始顺序")
            
            # 生成输出文件名（使用工作目录）
            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.work_subdirs['mir'] / f"MIR_Results_{date_str}.xlsx"
            
            # 保存到Excel（多个工作表）
            try:
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                wb.remove(wb.active)  # 删除默认sheet
                
                # Sheet 1: 按Source Lot的总结信息
                summary_sheet = wb.create_sheet("Summary by Source Lot")
                for r in dataframe_to_rows(result_df, index=False, header=True):
                    summary_sheet.append(r)
                
                # Sheet 2: 详细的Unit Name对比（如果有）
                if hasattr(self, 'unit_comparison_details') and self.unit_comparison_details:
                    unit_comparison_df = pd.DataFrame(self.unit_comparison_details)
                    
                    # 按Source Lot和Status排序
                    if 'Source Lot' in unit_comparison_df.columns:
                        unit_comparison_df = unit_comparison_df.sort_values(
                            by=['Source Lot', 'Status', 'Unit Name'],
                            ascending=[True, True, True]
                        )
                    
                    detail_sheet = wb.create_sheet("Unit Comparison Details")
                    for r in dataframe_to_rows(unit_comparison_df, index=False, header=True):
                        detail_sheet.append(r)
                    
                    # Sheet 3: 不匹配的Units（Missing和Extra）
                    unmatched_df = unit_comparison_df[unit_comparison_df['Status'].isin(['Missing', 'Extra'])].copy()
                    if not unmatched_df.empty:
                        unmatched_sheet = wb.create_sheet("Unmatched Units")
                        for r in dataframe_to_rows(unmatched_df, index=False, header=True):
                            unmatched_sheet.append(r)
                        LOGGER.info(f"   不匹配的Units: {len(unmatched_df)} 个")
                    else:
                        LOGGER.info(f"   ✅ 所有Units都匹配")
                    
                    LOGGER.info(f"   详细的Unit对比: {len(unit_comparison_df)} 行")
                
                wb.save(output_file)
                LOGGER.info(f"✅ 所有MIR结果已保存到: {output_file}")
                LOGGER.info(f"   共 {len(mir_results)} 个Source Lot的总结信息")
                if hasattr(self, 'unit_comparison_details') and self.unit_comparison_details:
                    total_units = len(self.unit_comparison_details)
                    matched_count = len([u for u in self.unit_comparison_details if u.get('Status') == 'Matched'])
                    missing_count = len([u for u in self.unit_comparison_details if u.get('Status') == 'Missing'])
                    extra_count = len([u for u in self.unit_comparison_details if u.get('Status') == 'Extra'])
                    LOGGER.info(f"   详细的Unit对比: 总计 {total_units} 个units")
                    LOGGER.info(f"     - 匹配: {matched_count} 个")
                    LOGGER.info(f"     - 缺失: {missing_count} 个")
                    LOGGER.info(f"     - 额外: {extra_count} 个")
                LOGGER.info(f"   文件格式: Excel (.xlsx)")
                LOGGER.info(f"   包含工作表: Summary by Source Lot" + 
                           (", Unit Comparison Details, Unmatched Units" if hasattr(self, 'unit_comparison_details') and self.unit_comparison_details else ""))
            except Exception as e:
                # 如果Excel保存失败，尝试保存为CSV
                LOGGER.warning(f"保存Excel文件失败: {e}，尝试保存为CSV格式...")
                csv_file = self.work_subdirs['mir'] / f"MIR_Results_{date_str}.csv"
                result_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
                output_file = csv_file
                LOGGER.info(f"✅ 所有MIR结果已保存到: {output_file} (CSV格式)")
            
            self.last_mir_result_file = output_file
            
            # 显示每行的详细信息（按排序后的顺序）
            for idx, (_, row) in enumerate(result_df.iterrows(), 1):
                source_lot = row.get(source_lot_col, 'N/A') if source_lot_col else 'N/A'
                mir = row.get('MIR', 'N/A')
                LOGGER.info(f"   第 {idx} 行: SourceLot={source_lot}, MIR={mir}")
            
            return output_file
            
        except Exception as e:
            LOGGER.error(f"保存MIR结果到CSV失败: {e}")
            import traceback
            LOGGER.error(traceback.format_exc())
    
    def _generate_summary_table_for_vpos(self, source_lot_file_path: Path, mir_results: list) -> None:
        """
        为 VPOs 模式生成汇总表（只包含 Source Lot, Part Type, Quantity, MIR）
        
        Args:
            source_lot_file_path: Source Lot文件路径
            mir_results: MIR结果列表
        """
        try:
            # 读取 Source Lot 文件获取 Part Type 和 Quantity
            if not source_lot_file_path or not source_lot_file_path.exists():
                LOGGER.warning(f"Source Lot 文件不存在: {source_lot_file_path}，无法生成汇总表")
                return
            
            LOGGER.info(f"读取 Source Lot 文件: {source_lot_file_path}")
            source_lot_df = read_excel_file(source_lot_file_path)
            LOGGER.info(f"  - 包含 {len(source_lot_df)} 行数据")
            
            # 查找列名
            source_lot_col = None
            for col in source_lot_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS', 'SOURCE']:
                    source_lot_col = col
                    LOGGER.info(f"  找到 Source Lot 列: '{col}'")
                    break
            
            part_type_col = None
            for col in source_lot_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE']:
                    part_type_col = col
                    LOGGER.info(f"  找到 Part Type 列: '{col}'")
                    break
            
            quantity_col = None
            for col in source_lot_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['QUANTITY', 'QTY', 'QTY.', 'COUNT']:
                    quantity_col = col
                    LOGGER.info(f"  找到 Quantity 列: '{col}'")
                    break
            
            if not source_lot_col:
                LOGGER.warning(f"未找到 Source Lot 列，无法生成汇总表。可用列: {source_lot_df.columns.tolist()}")
                return
            
            # 创建 MIR 映射（source lot -> MIR）
            mir_map = {}
            for result in mir_results:
                source_lot = result.get('Source Lot', '') or result.get(source_lot_col, '')
                mir = result.get('MIR', '')
                if source_lot and mir:
                    mir_map[str(source_lot).strip()] = str(mir).strip()
            
            # 创建汇总表
            summary_list = []
            for _, row in source_lot_df.iterrows():
                source_lot_value = str(row[source_lot_col]).strip() if pd.notna(row[source_lot_col]) else ''
                if not source_lot_value:
                    continue
                
                part_type = ''
                if part_type_col and pd.notna(row.get(part_type_col)):
                    part_type = str(row[part_type_col]).strip()
                
                quantity = ''
                if quantity_col and pd.notna(row.get(quantity_col)):
                    quantity = str(row[quantity_col]).strip()
                
                mir = mir_map.get(source_lot_value, '')
                
                summary_list.append({
                    'Source Lot': source_lot_value,
                    'Part Type': part_type,
                    'Quantity': quantity,
                    'MIR': mir
                })
            
            if not summary_list:
                LOGGER.warning("没有数据可生成汇总表")
                return
            
            summary_df = pd.DataFrame(summary_list)
            summary_df = summary_df.sort_values(by='Source Lot', ascending=True)
            
            # 保存汇总表
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            summary_output_file = self.work_subdirs['mir'] / f"Summary_Table_{timestamp}.xlsx"
            self.summary_table_file = summary_output_file
            
            summary_df.to_excel(summary_output_file, index=False, engine='openpyxl')
            LOGGER.info(f"✅ VPOs模式汇总表已保存到: {summary_output_file}")
            LOGGER.info(f"   包含列: Source Lot, Part Type, Quantity, MIR ({len(summary_df)} 行)")
            
            # 在写入 MIR 后，将 Summary_Table 重命名为 Spark 会调用的表的名字
            # 重命名为 MIR_Results_*.xlsx（使用当前时间戳），这样 Spark 步骤可以读取它
            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            spark_summary_file = self.work_subdirs['mir'] / f"MIR_Results_{date_str}.xlsx"
            
            try:
                # 如果目标文件已存在，先删除
                if spark_summary_file.exists():
                    spark_summary_file.unlink()
                
                # 重命名文件
                summary_output_file.rename(spark_summary_file)
                self.summary_table_file = spark_summary_file
                LOGGER.info(f"✅ 汇总表已重命名为: {spark_summary_file.name}（供 Spark 使用）")
            except Exception as e:
                LOGGER.warning(f"重命名汇总表失败: {e}，保持原文件名: {summary_output_file.name}")
            
        except Exception as e:
            LOGGER.error(f"生成 VPOs 模式汇总表失败: {e}")
            import traceback
            LOGGER.error(traceback.format_exc())
    
    def _merge_validation_tables_immediately(self, source_lot_file_path: Path | None, units_df: pd.DataFrame, source_lot_col: str, unit_name_col: str) -> None:
        """
        在生成两张表时立即合并 available_units_export 和 Units_Validation_Comparison 表
        生成包含 source lot, Part Type, quantity 的合并表（此时MIR还没有）
        
        Args:
            source_lot_file_path: Source Lot文件路径（可以为None，会从units_df中获取信息）
            units_df: Units DataFrame（包含 source lot 和 unit name 信息）
            source_lot_col: Source Lot 列名
            unit_name_col: Unit Name 列名
        """
        try:
            # 检查文件是否存在
            if not self.available_units_export_file or not self.available_units_export_file.exists():
                LOGGER.warning(f"available_units_export 文件不存在: {self.available_units_export_file}")
                return
            
            if not self.units_validation_comparison_file or not self.units_validation_comparison_file.exists():
                LOGGER.warning(f"Units_Validation_Comparison 文件不存在: {self.units_validation_comparison_file}")
                return
            
            LOGGER.info(f"读取 available_units_export 文件: {self.available_units_export_file}")
            available_units_df = read_excel_file(self.available_units_export_file)
            LOGGER.info(f"  - 包含 {len(available_units_df)} 行数据")
            
            LOGGER.info(f"读取 Units_Validation_Comparison 文件: {self.units_validation_comparison_file}")
            # 如果是Excel文件，读取 'All Units Comparison' sheet
            if self.units_validation_comparison_file.suffix.lower() == '.xlsx':
                try:
                    validation_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='All Units Comparison')
                except:
                    # 如果没有该sheet，读取第一个sheet
                    validation_df = read_excel_file(self.units_validation_comparison_file)
            else:
                validation_df = read_excel_file(self.units_validation_comparison_file)
            LOGGER.info(f"  - 包含 {len(validation_df)} 行数据")
            
            # 读取 Source Lot 文件获取 Part Type 和 quantity（如果文件存在）
            source_lot_df = None
            part_type_col = None
            quantity_col = None
            
            if source_lot_file_path and source_lot_file_path.exists():
                LOGGER.info(f"读取 Source Lot 文件: {source_lot_file_path}")
                source_lot_df = read_excel_file(source_lot_file_path)
                LOGGER.info(f"  - 包含 {len(source_lot_df)} 行数据")
                
                # 查找列名
                # Part Type 列
                for col in source_lot_df.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE']:
                        part_type_col = col
                        break
                
                # Quantity 列
                for col in source_lot_df.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper in ['QUANTITY', 'QTY', 'QTY.', 'COUNT']:
                        quantity_col = col
                        break
            else:
                # 如果没有 source_lot_file，尝试从 units_df 中获取 Part Type 和 Quantity
                LOGGER.info("未提供 Source Lot 文件，尝试从 units_df 中获取 Part Type 和 Quantity 信息...")
                for col in units_df.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE'] and part_type_col is None:
                        part_type_col = col
                    if col_upper in ['QUANTITY', 'QTY', 'QTY.', 'COUNT'] and quantity_col is None:
                        quantity_col = col
                
                if part_type_col or quantity_col:
                    LOGGER.info(f"  从 units_df 中找到列: Part Type={part_type_col}, Quantity={quantity_col}")
                    source_lot_df = units_df  # 使用 units_df 作为数据源
            
            # Unit Name 列（在 available_units_export 和 validation 中）
            unit_name_col_available = None
            for col in available_units_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['UNIT NAME', 'UNITNAME', 'UNIT_NAME', 'UNIT', 'UNITS', 'UNIT ID', 'UNITID']:
                    unit_name_col_available = col
                    break
            
            unit_name_col_validation = None
            for col in validation_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['UNIT NAME', 'UNITNAME', 'UNIT_NAME', 'UNIT', 'UNITS', 'UNIT ID', 'UNITID']:
                    unit_name_col_validation = col
                    break
            
            # 创建 Source Lot -> Part Type, Quantity 映射
            source_lot_info_map = {}
            if source_lot_df is not None:
                # 重新查找 Source Lot 列（在 source_lot_df 中）
                source_lot_col_in_file = None
                for col in source_lot_df.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper in ['SOURCE', 'SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS']:
                        source_lot_col_in_file = col
                        LOGGER.info(f"  在 source_lot_df 中找到 Source Lot 列: '{col}'")
                        break
                
                if source_lot_col_in_file:
                    for _, row in source_lot_df.iterrows():
                        sl = str(row[source_lot_col_in_file]).strip() if pd.notna(row[source_lot_col_in_file]) else ''
                        if sl:
                            info = {}
                            if part_type_col and pd.notna(row.get(part_type_col)):
                                info['Part Type'] = str(row[part_type_col]).strip()
                            if quantity_col and pd.notna(row.get(quantity_col)):
                                info['Quantity'] = str(row[quantity_col]).strip()
                            if info:
                                source_lot_info_map[sl] = info
                else:
                    LOGGER.warning(f"在 source_lot_df 中未找到 Source Lot 列。可用列: {source_lot_df.columns.tolist()}")
            
            # 合并数据 - 使用 Unit Name 作为键来合并两张表的所有列
            # 首先创建 available_units_df 的索引（以 Unit Name 为键）
            available_units_dict = {}
            if unit_name_col_available:
                for _, avail_row in available_units_df.iterrows():
                    unit_name = str(avail_row[unit_name_col_available]).strip() if pd.notna(avail_row.get(unit_name_col_available)) else ''
                    if unit_name:
                        # 存储该 unit 的所有信息
                        available_units_dict[unit_name] = avail_row.to_dict()
            
            # 确定两张表的共同列和独有列
            validation_cols = set(validation_df.columns)
            available_cols = set(available_units_df.columns)
            common_cols = validation_cols & available_cols
            validation_only_cols = validation_cols - available_cols
            available_only_cols = available_cols - validation_cols
            
            merged_list = []
            
            # 从 validation_df 开始，因为它包含了 source lot 信息
            for _, val_row in validation_df.iterrows():
                unit_name = str(val_row[unit_name_col_validation]).strip() if unit_name_col_validation and pd.notna(val_row.get(unit_name_col_validation)) else ''
                source_lot_str = str(val_row.get('Source Lot', '')).strip() if pd.notna(val_row.get('Source Lot')) else ''
                
                if not unit_name:
                    continue
                
                # 如果 source lot 包含多个（逗号分隔），取第一个
                if ',' in source_lot_str:
                    source_lot_str = source_lot_str.split(',')[0].strip()
                
                # 创建合并行，包含两张表的所有列
                merged_row = {}
                
                # 处理共同列：优先使用 available_units_df 的值（如果存在），否则使用 validation_df 的值
                for col in common_cols:
                    if unit_name in available_units_dict:
                        avail_val = available_units_dict[unit_name].get(col)
                        if pd.notna(avail_val) and str(avail_val).strip():
                            merged_row[col] = avail_val
                        else:
                            merged_row[col] = val_row.get(col) if pd.notna(val_row.get(col)) else ''
                    else:
                        merged_row[col] = val_row.get(col) if pd.notna(val_row.get(col)) else ''
                
                # 添加 validation_df 独有的列
                for col in validation_only_cols:
                    merged_row[col] = val_row.get(col) if pd.notna(val_row.get(col)) else ''
                
                # 添加 available_units_df 独有的列（如果该 unit 存在）
                if unit_name in available_units_dict:
                    avail_data = available_units_dict[unit_name]
                    for col in available_only_cols:
                        merged_row[col] = avail_data.get(col) if pd.notna(avail_data.get(col)) else ''
                else:
                    # 如果 available_units_df 中没有该 unit，填充空值
                    for col in available_only_cols:
                        merged_row[col] = ''
                
                # 确保 Source Lot, Part Type, Quantity, MIR 列存在（这些是汇总信息）
                merged_row['Source Lot'] = source_lot_str if source_lot_str else 'N/A'
                merged_row['Part Type'] = source_lot_info_map.get(source_lot_str, {}).get('Part Type', '')
                merged_row['Quantity'] = source_lot_info_map.get(source_lot_str, {}).get('Quantity', '')
                merged_row['MIR'] = ''  # MIR将在提交后更新
                
                merged_list.append(merged_row)
            
            # 如果 available_units_df 中有额外的 units（不在 validation 中），也添加进去
            if unit_name_col_available:
                validation_units_set = set()
                if unit_name_col_validation:
                    for _, val_row in validation_df.iterrows():
                        unit = str(val_row[unit_name_col_validation]).strip() if pd.notna(val_row.get(unit_name_col_validation)) else ''
                        if unit:
                            validation_units_set.add(unit)
                
                for _, avail_row in available_units_df.iterrows():
                    unit_name = str(avail_row[unit_name_col_available]).strip() if pd.notna(avail_row.get(unit_name_col_available)) else ''
                    if unit_name and unit_name not in validation_units_set:
                        # 这个 unit 在 available 中但不在 validation 中，添加它
                        merged_row = {}
                        
                        # 处理共同列：使用 available_units_df 的值
                        for col in common_cols:
                            merged_row[col] = avail_row.get(col) if pd.notna(avail_row.get(col)) else ''
                        
                        # validation_df 独有的列填充空值
                        for col in validation_only_cols:
                            merged_row[col] = ''
                        
                        # 添加 available_units_df 独有的列
                        for col in available_only_cols:
                            merged_row[col] = avail_row.get(col) if pd.notna(avail_row.get(col)) else ''
                        
                        # 添加汇总信息
                        merged_row['Source Lot'] = 'N/A'
                        merged_row['Part Type'] = ''
                        merged_row['Quantity'] = ''
                        merged_row['MIR'] = ''
                        
                        merged_list.append(merged_row)
            
            # 创建汇总表（从 available_units_export 文件中读取信息）
            summary_list = []
            
            # 从 available_units_export 文件中查找 Source Lot, Part Type, Quantity 列
            source_lot_col_available = None
            part_type_col_available = None
            quantity_col_available = None
            
            for col in available_units_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCE', 'SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS'] and source_lot_col_available is None:
                    source_lot_col_available = col
                elif col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE'] and part_type_col_available is None:
                    part_type_col_available = col
                elif col_upper in ['QUANTITY', 'QTY', 'QTY.', 'COUNT'] and quantity_col_available is None:
                    quantity_col_available = col
            
            LOGGER.info(f"从 available_units_export 文件中找到列: Source Lot={source_lot_col_available}, Part Type={part_type_col_available}, Quantity={quantity_col_available}")
            
            # 从 available_units_export 文件中按 Source Lot 分组提取信息
            source_lot_info_from_export = {}
            if source_lot_col_available:
                grouped = available_units_df.groupby(source_lot_col_available)
                for source_lot, group_df in grouped:
                    if pd.isna(source_lot) or str(source_lot).strip() == '' or str(source_lot).strip() == 'N/A':
                        continue
                    
                    source_lot_str = str(source_lot).strip()
                    # 如果 source lot 包含多个（逗号分隔），取第一个
                    if ',' in source_lot_str:
                        source_lot_str = source_lot_str.split(',')[0].strip()
                    
                    # 从 group_df 中获取 Part Type（取第一个非空值）
                    part_type = ''
                    if part_type_col_available:
                        part_type_values = group_df[part_type_col_available].dropna()
                        if not part_type_values.empty:
                            part_type = str(part_type_values.iloc[0]).strip()
                    
                    # Quantity 是该 source lot 在 available_units_export 中的 units 数量（统计数量）
                    quantity = len(group_df)  # 统计该 source lot 有多少个可用的 units
                    
                    source_lot_info_from_export[source_lot_str] = {
                        'Part Type': part_type,
                        'Quantity': quantity,
                    }
            else:
                LOGGER.warning("在 available_units_export 文件中未找到 Source Lot 列，尝试从 validation_df 中获取...")
                # 如果 available_units_export 中没有 Source Lot 列，从 validation_df 中获取
                if not validation_df.empty and 'Source Lot' in validation_df.columns:
                    grouped = validation_df.groupby('Source Lot')
                    for source_lot, group_df in grouped:
                        if pd.isna(source_lot) or str(source_lot).strip() == '' or str(source_lot).strip() == 'N/A':
                            continue
                        
                        source_lot_str = str(source_lot).strip()
                        if ',' in source_lot_str:
                            source_lot_str = source_lot_str.split(',')[0].strip()
                        
                        # 从 source_lot_info_map 中获取 Part Type 和 Quantity（如果存在）
                        part_type = source_lot_info_map.get(source_lot_str, {}).get('Part Type', '')
                        quantity = source_lot_info_map.get(source_lot_str, {}).get('Quantity', '')
                        total_units = len(group_df)
                        
                        source_lot_info_from_export[source_lot_str] = {
                            'Part Type': part_type,
                            'Quantity': quantity,
                            'Total Units': total_units
                        }
            
            # 从 validation_df 中获取统计信息（按 Source Lot 分组）
            validation_stats = {}
            if not validation_df.empty and 'Source Lot' in validation_df.columns:
                grouped = validation_df.groupby('Source Lot')
                for source_lot, group_df in grouped:
                    if pd.isna(source_lot) or str(source_lot).strip() == '' or str(source_lot).strip() == 'N/A':
                        continue
                    
                    source_lot_str = str(source_lot).strip()
                    # 如果 source lot 包含多个（逗号分隔），取第一个
                    if ',' in source_lot_str:
                        source_lot_str = source_lot_str.split(',')[0].strip()
                    
                    # 统计 Units 数量
                    total_units = len(group_df)
                    matched_units = len(group_df[group_df.get('Status', '') == 'Matched']) if 'Status' in group_df.columns else total_units
                    
                    validation_stats[source_lot_str] = {
                        'Total Units': total_units,
                        'Matched Units': matched_units,
                        'Missing Units': total_units - matched_units
                    }
            
            # 生成汇总表：从 available_units_export 文件中获取的信息
            for source_lot_str, info in source_lot_info_from_export.items():
                summary_list.append({
                    'Source Lot': source_lot_str,
                    'Part Type': info.get('Part Type', ''),
                    'Quantity': info.get('Quantity', ''),
                    'MIR': '',  # MIR将在提交后更新
                })
            
            # 如果 validation_df 中有 Source Lot 不在 available_units_export 中，也添加进去
            for source_lot_str in validation_stats.keys():
                if source_lot_str not in source_lot_info_from_export:
                    # 尝试从 source_lot_info_map 获取信息
                    part_type = source_lot_info_map.get(source_lot_str, {}).get('Part Type', '')
                    quantity = source_lot_info_map.get(source_lot_str, {}).get('Quantity', '')
                    
                    summary_list.append({
                        'Source Lot': source_lot_str,
                        'Part Type': part_type,
                        'Quantity': quantity,
                        'MIR': '',  # MIR将在提交后更新
                    })
            
            summary_df = pd.DataFrame(summary_list)
            if not summary_df.empty:
                summary_df = summary_df.sort_values(by='Source Lot', ascending=True)
            
            # 创建单独的汇总表（只包含 Source Lot, Part Type, Quantity, MIR）
            summary_simple_df = pd.DataFrame()
            if not summary_df.empty:
                # 确保只包含这4列，如果列不存在则创建空列
                summary_simple_df = pd.DataFrame()
                summary_simple_df['Source Lot'] = summary_df['Source Lot'] if 'Source Lot' in summary_df.columns else ''
                summary_simple_df['Part Type'] = summary_df['Part Type'] if 'Part Type' in summary_df.columns else ''
                summary_simple_df['Quantity'] = summary_df['Quantity'] if 'Quantity' in summary_df.columns else ''
                summary_simple_df['MIR'] = summary_df['MIR'] if 'MIR' in summary_df.columns else ''
            
            # 保存到合并Excel文件（将两张表作为不同的工作表）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            merged_output_file = self.work_subdirs['mir'] / f"Merged_Validation_Table_{timestamp}.xlsx"
            
            # 保存汇总表文件路径
            summary_output_file = self.work_subdirs['mir'] / f"Summary_Table_{timestamp}.xlsx"
            self.summary_table_file = summary_output_file
            
            # 保存合并表路径，以便后续更新MIR
            self.merged_validation_file = merged_output_file
            
            try:
                # 保存合并表（包含两张原始表作为不同的工作表）
                with pd.ExcelWriter(merged_output_file, engine='openpyxl') as writer:
                    # 工作表1: available_units_export 表
                    available_units_df.to_excel(writer, sheet_name='Available Units Export', index=False)
                    LOGGER.info(f"  - 工作表1 'Available Units Export': {len(available_units_df)} 行")
                    
                    # 工作表2: Units_Validation_Comparison 表（All Units Comparison sheet）
                    validation_df.to_excel(writer, sheet_name='Units Validation Comparison', index=False)
                    LOGGER.info(f"  - 工作表2 'Units Validation Comparison': {len(validation_df)} 行")
                    
                    # 如果 Units_Validation_Comparison 文件有多个工作表，也添加进去
                    if self.units_validation_comparison_file and self.units_validation_comparison_file.suffix.lower() == '.xlsx':
                        try:
                            # 读取 Source Lot Summary sheet
                            source_lot_stats_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='Source Lot Summary')
                            source_lot_stats_df.to_excel(writer, sheet_name='Source Lot Summary', index=False)
                            LOGGER.info(f"  - 工作表3 'Source Lot Summary': {len(source_lot_stats_df)} 行")
                        except:
                            pass
                        
                        try:
                            # 读取 Missing Units sheet
                            missing_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='Missing Units')
                            if not missing_df.empty:
                                missing_df.to_excel(writer, sheet_name='Missing Units', index=False)
                                LOGGER.info(f"  - 工作表4 'Missing Units': {len(missing_df)} 行")
                        except:
                            pass
                        
                        try:
                            # 读取 Extra Units sheet
                            extra_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='Extra Units')
                            if not extra_df.empty:
                                extra_df.to_excel(writer, sheet_name='Extra Units', index=False)
                                LOGGER.info(f"  - 工作表5 'Extra Units': {len(extra_df)} 行")
                        except:
                            pass
                
                # 设置 Units Validation Comparison 工作表中 Status 列的背景颜色
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill
                    
                    wb = load_workbook(merged_output_file)
                    if 'Units Validation Comparison' in wb.sheetnames:
                        ws = wb['Units Validation Comparison']
                        
                        # 查找 Status 列的索引
                        status_col_idx = None
                        for col_idx, cell in enumerate(ws[1], 1):  # 第一行是标题行
                            if cell.value and str(cell.value).strip().upper() == 'STATUS':
                                status_col_idx = col_idx
                                break
                        
                        if status_col_idx:
                            # 定义颜色填充
                            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
                            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # 浅红色
                            
                            # 从第2行开始（跳过标题行）设置背景颜色
                            for row_idx in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row_idx, column=status_col_idx)
                                status_value = str(cell.value).strip() if cell.value else ''
                                
                                if status_value.upper() == 'MATCHED':
                                    cell.fill = green_fill
                                else:
                                    cell.fill = red_fill
                            
                            wb.save(merged_output_file)
                            LOGGER.info(f"  - 已为 'Units Validation Comparison' 工作表的 Status 列设置背景颜色（Matched=绿色，其他=红色）")
                except Exception as e:
                    LOGGER.warning(f"设置 Status 列背景颜色失败: {e}")
                
                LOGGER.info(f"✅ 合并表已保存到: {merged_output_file}")
                LOGGER.info(f"   包含多个工作表：Available Units Export, Units Validation Comparison 等")
                
                # 保存单独的汇总表（只包含 Source Lot, Part Type, Quantity, MIR）
                if not summary_simple_df.empty:
                    summary_simple_df.to_excel(summary_output_file, index=False, engine='openpyxl')
                    LOGGER.info(f"✅ 汇总表已保存到: {summary_output_file}")
                    LOGGER.info(f"   包含列: Source Lot, Part Type, Quantity, MIR ({len(summary_simple_df)} 行)")
                    LOGGER.info(f"   注意：MIR列将在提交MIR后更新")
                
                # 删除临时文件（因为数据已经合并到 Merged_Validation_Table 中）
                try:
                    if self.available_units_export_file and self.available_units_export_file.exists():
                        self.available_units_export_file.unlink()
                        LOGGER.debug(f"已删除临时文件: {self.available_units_export_file.name}")
                    if self.units_validation_comparison_file and self.units_validation_comparison_file.exists():
                        self.units_validation_comparison_file.unlink()
                        LOGGER.debug(f"已删除临时文件: {self.units_validation_comparison_file.name}")
                except Exception as e:
                    LOGGER.warning(f"删除临时文件时出错（可忽略）: {e}")
            except Exception as e:
                LOGGER.warning(f"保存Excel文件失败: {e}，尝试保存为CSV格式...")
                import traceback
                LOGGER.debug(traceback.format_exc())
                # 如果Excel保存失败，分别保存为CSV文件
                merged_output_file = self.work_subdirs['mir'] / f"Merged_Validation_Table_{timestamp}.csv"
                merged_df.to_csv(merged_output_file, index=False, encoding='utf-8-sig')
                LOGGER.info(f"✅ 合并表已保存到: {merged_output_file} (CSV格式)")
                LOGGER.warning("   注意：CSV格式只能保存合并表，原始表请查看原始文件")
            
        except Exception as e:
            LOGGER.error(f"立即合并验证表失败: {e}")
            import traceback
            LOGGER.error(traceback.format_exc())
    
    def _merge_validation_tables(self, source_lot_file_path: Path, mir_results: list) -> None:
        """
        合并 available_units_export 和 Units_Validation_Comparison 表
        生成包含 source lot, Part Type, quantity, MIR 的合并表
        
        Args:
            source_lot_file_path: Source Lot文件路径
            mir_results: MIR结果列表
        """
        try:
            
            # 检查文件是否存在
            if not self.available_units_export_file or not self.available_units_export_file.exists():
                LOGGER.warning(f"available_units_export 文件不存在: {self.available_units_export_file}")
                return
            
            if not self.units_validation_comparison_file or not self.units_validation_comparison_file.exists():
                LOGGER.warning(f"Units_Validation_Comparison 文件不存在: {self.units_validation_comparison_file}")
                return
            
            LOGGER.info(f"读取 available_units_export 文件: {self.available_units_export_file}")
            available_units_df = read_excel_file(self.available_units_export_file)
            LOGGER.info(f"  - 包含 {len(available_units_df)} 行数据")
            
            LOGGER.info(f"读取 Units_Validation_Comparison 文件: {self.units_validation_comparison_file}")
            # 如果是Excel文件，读取 'All Units Comparison' sheet
            if self.units_validation_comparison_file.suffix.lower() == '.xlsx':
                try:
                    validation_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='All Units Comparison')
                except:
                    # 如果没有该sheet，读取第一个sheet
                    validation_df = read_excel_file(self.units_validation_comparison_file)
            else:
                validation_df = read_excel_file(self.units_validation_comparison_file)
            LOGGER.info(f"  - 包含 {len(validation_df)} 行数据")
            
            # 读取 Source Lot 文件获取 Part Type 和 quantity
            LOGGER.info(f"读取 Source Lot 文件: {source_lot_file_path}")
            source_lot_df = read_excel_file(source_lot_file_path)
            LOGGER.info(f"  - 包含 {len(source_lot_df)} 行数据")
            
            # 查找列名
            # Source Lot 列
            source_lot_col = None
            for col in source_lot_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT']:
                    source_lot_col = col
                    break
            
            # Part Type 列
            part_type_col = None
            for col in source_lot_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE']:
                    part_type_col = col
                    break
            
            # Quantity 列
            quantity_col = None
            for col in source_lot_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['QUANTITY', 'QTY', 'QTY.', 'COUNT']:
                    quantity_col = col
                    break
            
            # Unit Name 列（在 available_units_export 和 validation 中）
            unit_name_col_available = None
            for col in available_units_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['UNIT NAME', 'UNITNAME', 'UNIT_NAME', 'UNIT', 'UNITS', 'UNIT ID', 'UNITID']:
                    unit_name_col_available = col
                    break
            
            unit_name_col_validation = None
            for col in validation_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['UNIT NAME', 'UNITNAME', 'UNIT_NAME', 'UNIT', 'UNITS', 'UNIT ID', 'UNITID']:
                    unit_name_col_validation = col
                    break
            
            # 创建 MIR 映射（source lot -> MIR）
            mir_map = {}
            for result in mir_results:
                source_lot = result.get('Source Lot', '')
                mir = result.get('MIR', '')
                if source_lot and mir:
                    mir_map[str(source_lot).strip()] = str(mir).strip()
            
            # 创建 Source Lot -> Part Type, Quantity 映射
            source_lot_info_map = {}
            if source_lot_col:
                for _, row in source_lot_df.iterrows():
                    sl = str(row[source_lot_col]).strip() if pd.notna(row[source_lot_col]) else ''
                    if sl:
                        info = {}
                        if part_type_col and pd.notna(row.get(part_type_col)):
                            info['Part Type'] = str(row[part_type_col]).strip()
                        if quantity_col and pd.notna(row.get(quantity_col)):
                            info['Quantity'] = str(row[quantity_col]).strip()
                        if info:
                            source_lot_info_map[sl] = info
            
            # 创建汇总表（从 available_units_export 文件中读取信息）
            summary_list = []
            
            # 从 available_units_export 文件中查找 Source Lot, Part Type, Quantity 列
            source_lot_col_available = None
            part_type_col_available = None
            quantity_col_available = None
            
            for col in available_units_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCE', 'SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS'] and source_lot_col_available is None:
                    source_lot_col_available = col
                elif col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE'] and part_type_col_available is None:
                    part_type_col_available = col
                elif col_upper in ['QUANTITY', 'QTY', 'QTY.', 'COUNT'] and quantity_col_available is None:
                    quantity_col_available = col
            
            LOGGER.info(f"从 available_units_export 文件中找到列: Source Lot={source_lot_col_available}, Part Type={part_type_col_available}, Quantity={quantity_col_available}")
            
            # 从 available_units_export 文件中按 Source Lot 分组提取信息
            source_lot_info_from_export = {}
            if source_lot_col_available:
                grouped = available_units_df.groupby(source_lot_col_available)
                for source_lot, group_df in grouped:
                    if pd.isna(source_lot) or str(source_lot).strip() == '' or str(source_lot).strip() == 'N/A':
                        continue
                    
                    source_lot_str = str(source_lot).strip()
                    # 如果 source lot 包含多个（逗号分隔），取第一个
                    if ',' in source_lot_str:
                        source_lot_str = source_lot_str.split(',')[0].strip()
                    
                    # 从 group_df 中获取 Part Type（取第一个非空值）
                    part_type = ''
                    if part_type_col_available:
                        part_type_values = group_df[part_type_col_available].dropna()
                        if not part_type_values.empty:
                            part_type = str(part_type_values.iloc[0]).strip()
                    
                    # Quantity 是该 source lot 在 available_units_export 中的 units 数量（统计数量）
                    quantity = len(group_df)  # 统计该 source lot 有多少个可用的 units
                    
                    source_lot_info_from_export[source_lot_str] = {
                        'Part Type': part_type,
                        'Quantity': quantity
                    }
            else:
                LOGGER.warning("在 available_units_export 文件中未找到 Source Lot 列，尝试从 source_lot_info_map 中获取...")
            
            # 从 validation_df 中获取统计信息（按 Source Lot 分组）
            validation_stats = {}
            if not validation_df.empty and 'Source Lot' in validation_df.columns:
                grouped = validation_df.groupby('Source Lot')
                for source_lot, group_df in grouped:
                    if pd.isna(source_lot) or str(source_lot).strip() == '' or str(source_lot).strip() == 'N/A':
                        continue
                    
                    source_lot_str = str(source_lot).strip()
                    # 如果 source lot 包含多个（逗号分隔），取第一个
                    if ',' in source_lot_str:
                        source_lot_str = source_lot_str.split(',')[0].strip()
                    
                    # 统计 Units 数量
                    total_units = len(group_df)
                    matched_units = len(group_df[group_df.get('Status', '') == 'Matched']) if 'Status' in group_df.columns else total_units
                    
                    # 获取 MIR
                    mir = mir_map.get(source_lot_str, '')
                    
                    validation_stats[source_lot_str] = {
                        'Total Units': total_units,
                        'Matched Units': matched_units,
                        'Missing Units': total_units - matched_units,
                        'MIR': mir
                    }
            
            # 生成汇总表：从 available_units_export 文件中获取的信息
            for source_lot_str, info in source_lot_info_from_export.items():
                stats = validation_stats.get(source_lot_str, {})
                summary_list.append({
                    'Source Lot': source_lot_str,
                    'Part Type': info.get('Part Type', ''),
                    'Quantity': info.get('Quantity', ''),
                    'MIR': stats.get('MIR', ''),
                })
            
            # 如果 validation_df 中有 Source Lot 不在 available_units_export 中，也添加进去
            for source_lot_str in validation_stats.keys():
                if source_lot_str not in source_lot_info_from_export:
                    # 尝试从 source_lot_info_map 获取信息
                    part_type = source_lot_info_map.get(source_lot_str, {}).get('Part Type', '')
                    quantity = source_lot_info_map.get(source_lot_str, {}).get('Quantity', '')
                    stats = validation_stats[source_lot_str]
                    
                    summary_list.append({
                        'Source Lot': source_lot_str,
                        'Part Type': part_type,
                        'Quantity': quantity,
                        'MIR': stats.get('MIR', ''),
                    })
            
            summary_df = pd.DataFrame(summary_list)
            if not summary_df.empty:
                summary_df = summary_df.sort_values(by='Source Lot', ascending=True)
            
            # 创建单独的汇总表（只包含 Source Lot, Part Type, Quantity, MIR）
            summary_simple_df = pd.DataFrame()
            if not summary_df.empty:
                summary_simple_df = summary_df[['Source Lot', 'Part Type', 'Quantity', 'MIR']].copy()
            
            # 更新合并Excel文件（如果已存在）或创建新文件
            if hasattr(self, 'merged_validation_file') and self.merged_validation_file and self.merged_validation_file.exists():
                # 更新现有文件
                merged_output_file = self.merged_validation_file
                LOGGER.info(f"更新现有合并表文件: {merged_output_file}")
            else:
                # 创建新文件
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                merged_output_file = self.work_subdirs['mir'] / f"Merged_Validation_Table_{timestamp}.xlsx"
                self.merged_validation_file = merged_output_file
            
            # 更新汇总表文件（如果已存在）或创建新文件
            if hasattr(self, 'summary_table_file') and self.summary_table_file and self.summary_table_file.exists():
                summary_output_file = self.summary_table_file
                LOGGER.info(f"更新现有汇总表文件: {summary_output_file}")
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                summary_output_file = self.work_subdirs['mir'] / f"Summary_Table_{timestamp}.xlsx"
                self.summary_table_file = summary_output_file
            
            try:
                # 更新合并表（将两张表作为不同的工作表）
                with pd.ExcelWriter(merged_output_file, engine='openpyxl') as writer:
                    # 工作表1: available_units_export 表
                    available_units_df.to_excel(writer, sheet_name='Available Units Export', index=False)
                    LOGGER.info(f"  - 工作表1 'Available Units Export': {len(available_units_df)} 行")
                    
                    # 工作表2: Units_Validation_Comparison 表（All Units Comparison sheet）
                    validation_df.to_excel(writer, sheet_name='Units Validation Comparison', index=False)
                    LOGGER.info(f"  - 工作表2 'Units Validation Comparison': {len(validation_df)} 行")
                    
                    # 如果 Units_Validation_Comparison 文件有多个工作表，也添加进去
                    if self.units_validation_comparison_file and self.units_validation_comparison_file.suffix.lower() == '.xlsx':
                        try:
                            # 读取 Source Lot Summary sheet
                            source_lot_stats_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='Source Lot Summary')
                            source_lot_stats_df.to_excel(writer, sheet_name='Source Lot Summary', index=False)
                            LOGGER.info(f"  - 工作表3 'Source Lot Summary': {len(source_lot_stats_df)} 行")
                        except:
                            pass
                        
                        try:
                            # 读取 Missing Units sheet
                            missing_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='Missing Units')
                            if not missing_df.empty:
                                missing_df.to_excel(writer, sheet_name='Missing Units', index=False)
                                LOGGER.info(f"  - 工作表4 'Missing Units': {len(missing_df)} 行")
                        except:
                            pass
                        
                        try:
                            # 读取 Extra Units sheet
                            extra_df = pd.read_excel(self.units_validation_comparison_file, sheet_name='Extra Units')
                            if not extra_df.empty:
                                extra_df.to_excel(writer, sheet_name='Extra Units', index=False)
                                LOGGER.info(f"  - 工作表5 'Extra Units': {len(extra_df)} 行")
                        except:
                            pass
                
                # 设置 Units Validation Comparison 工作表中 Status 列的背景颜色
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill
                    
                    wb = load_workbook(merged_output_file)
                    if 'Units Validation Comparison' in wb.sheetnames:
                        ws = wb['Units Validation Comparison']
                        
                        # 查找 Status 列的索引
                        status_col_idx = None
                        for col_idx, cell in enumerate(ws[1], 1):  # 第一行是标题行
                            if cell.value and str(cell.value).strip().upper() == 'STATUS':
                                status_col_idx = col_idx
                                break
                        
                        if status_col_idx:
                            # 定义颜色填充
                            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
                            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # 浅红色
                            
                            # 从第2行开始（跳过标题行）设置背景颜色
                            for row_idx in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row_idx, column=status_col_idx)
                                status_value = str(cell.value).strip() if cell.value else ''
                                
                                if status_value.upper() == 'MATCHED':
                                    cell.fill = green_fill
                                else:
                                    cell.fill = red_fill
                            
                            wb.save(merged_output_file)
                            LOGGER.info(f"  - 已为 'Units Validation Comparison' 工作表的 Status 列设置背景颜色（Matched=绿色，其他=红色）")
                except Exception as e:
                    LOGGER.warning(f"设置 Status 列背景颜色失败: {e}")
                
                LOGGER.info(f"✅ 合并表已更新: {merged_output_file}")
                LOGGER.info(f"   包含多个工作表：Available Units Export, Units Validation Comparison 等")
                
                # 更新汇总表（只包含 Source Lot, Part Type, Quantity, MIR）
                if not summary_simple_df.empty:
                    summary_simple_df.to_excel(summary_output_file, index=False, engine='openpyxl')
                    LOGGER.info(f"✅ 汇总表已更新: {summary_output_file}")
                    LOGGER.info(f"   包含列: Source Lot, Part Type, Quantity, MIR ({len(summary_simple_df)} 行)")
                    
                    # 在写入 MIR 后，将 Summary_Table 重命名为 Spark 会调用的表的名字
                    # 重命名为 MIR_Results_*.xlsx（使用当前时间戳），这样 Spark 步骤可以读取它
                    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                    spark_summary_file = self.work_subdirs['mir'] / f"MIR_Results_{date_str}.xlsx"
                    
                    try:
                        # 如果目标文件已存在，先删除
                        if spark_summary_file.exists():
                            spark_summary_file.unlink()
                        
                        # 重命名文件
                        summary_output_file.rename(spark_summary_file)
                        self.summary_table_file = spark_summary_file
                        LOGGER.info(f"✅ 汇总表已重命名为: {spark_summary_file.name}（供 Spark 使用）")
                    except Exception as e:
                        LOGGER.warning(f"重命名汇总表失败: {e}，保持原文件名: {summary_output_file.name}")
                
                # 删除临时文件（因为数据已经合并到 Merged_Validation_Table 中）
                try:
                    if self.available_units_export_file and self.available_units_export_file.exists():
                        self.available_units_export_file.unlink()
                        LOGGER.debug(f"已删除临时文件: {self.available_units_export_file.name}")
                    if self.units_validation_comparison_file and self.units_validation_comparison_file.exists():
                        self.units_validation_comparison_file.unlink()
                        LOGGER.debug(f"已删除临时文件: {self.units_validation_comparison_file.name}")
                except Exception as e:
                    LOGGER.warning(f"删除临时文件时出错（可忽略）: {e}")
            except Exception as e:
                LOGGER.warning(f"保存Excel文件失败: {e}，尝试保存为CSV格式...")
                import traceback
                LOGGER.debug(traceback.format_exc())
                # 如果Excel保存失败，分别保存为CSV文件
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                merged_output_file = self.work_subdirs['mir'] / f"Merged_Validation_Table_{timestamp}.csv"
                merged_df.to_csv(merged_output_file, index=False, encoding='utf-8-sig')
                LOGGER.info(f"✅ 合并表已保存到: {merged_output_file} (CSV格式)")
                LOGGER.warning("   注意：CSV格式只能保存合并表，原始表请查看原始文件")
            
        except Exception as e:
            LOGGER.error(f"合并验证表失败: {e}")
            import traceback
            LOGGER.error(traceback.format_exc())
    
    def _step_start_mole(self) -> None:
        """步骤0: 预先启动Mole工具"""
        try:
            LOGGER.info("正在启动Mole工具...")
            # 调用_ensure_application来启动Mole工具
            self.mole_submitter._ensure_application()
            LOGGER.info("✅ Mole工具已启动")
        except Exception as e:
            raise WorkflowError(f"启动Mole工具失败: {e}")
    
    def _close_mole(self) -> None:
        """关闭Mole工具"""
        try:
            if not self.mole_submitter._window:
                LOGGER.info("Mole窗口未连接，尝试查找并关闭...")
                # 尝试查找MOLE窗口
                if win32gui:
                    def find_mole_window(hwnd, windows):
                        try:
                            if not win32gui.IsWindowVisible(hwnd):
                                return True
                            window_text = win32gui.GetWindowText(hwnd)
                            if "MOLE" in window_text.upper() and "LOGIN" not in window_text.upper():
                                windows.append(hwnd)
                        except:
                            pass
                        return True
                    
                    mole_windows = []
                    win32gui.EnumWindows(find_mole_window, mole_windows)
                    if mole_windows:
                        LOGGER.info(f"找到 {len(mole_windows)} 个Mole窗口，尝试关闭...")
                        for hwnd in mole_windows:
                            try:
                                win32gui.PostMessage(hwnd, 0x0010, 0, 0)  # WM_CLOSE
                                LOGGER.info(f"已发送关闭消息到Mole窗口")
                            except:
                                pass
                        time.sleep(2.0)
                        return
            
            # 如果已连接窗口，尝试关闭
            if self.mole_submitter._window:
                try:
                    self.mole_submitter._window.close()
                    LOGGER.info("✅ 已关闭Mole窗口")
                    time.sleep(1.0)
                except Exception as e:
                    LOGGER.warning(f"关闭Mole窗口失败: {e}，尝试其他方法...")
                    # 尝试通过进程关闭
                    try:
                        if win32gui and self.mole_submitter._window:
                            hwnd = self.mole_submitter._window.handle
                            win32gui.PostMessage(hwnd, 0x0010, 0, 0)  # WM_CLOSE
                            time.sleep(1.0)
                    except:
                        pass
        except Exception as e:
            LOGGER.warning(f"关闭Mole工具时出错: {e}")
    
    def _step_read_excel(self, excel_file_path: Path) -> pd.DataFrame:
        """步骤1: 读取文件（Excel或CSV）"""
        try:
            df = read_excel_file(excel_file_path)
            file_type = "CSV文件" if excel_file_path.suffix.lower() == '.csv' else "Excel文件"
            LOGGER.info(f"✅ 成功读取{file_type}: {len(df)} 行，{len(df.columns)} 列")
            return df
        except Exception as e:
            raise WorkflowError(f"读取文件失败: {e}")
    
    def _step_submit_to_mole(self, df: pd.DataFrame, source_lot_file_path: Path, ui_config: dict = None) -> None:
        """步骤2: 提交MIR数据到Mole工具（循环处理所有行）
        
        Args:
            df: DataFrame
            source_lot_file_path: Source Lot文件路径（默认路径，可能被UI配置覆盖）
            ui_config: UI配置数据（可选），包含搜索方式和参数
        """
        try:
            # 获取搜索模式
            search_mode = ui_config.get('search_mode', 'vpos') if ui_config else 'vpos'
            
            # 根据搜索模式处理不同的数据源
            if search_mode == 'vpos':
                # VPOs 模式：从 Source Lot 文件读取
                # 使用 UI 配置的文件路径，如果没有则使用默认路径
                actual_source_lot_file_str = ui_config.get('source_lot_file', '') if ui_config else ''
                if not actual_source_lot_file_str:
                    actual_source_lot_file = source_lot_file_path
                else:
                    # 解析相对路径（基于 workflow_automation 目录的父目录）
                    actual_source_lot_file = Path(actual_source_lot_file_str)
                    if not actual_source_lot_file.is_absolute():
                        # 相对路径：基于 auto-vpo 根目录
                        base_dir = Path(__file__).parent.parent  # workflow_automation -> auto-vpo
                        actual_source_lot_file = (base_dir / actual_source_lot_file_str).resolve()
                
                # 检查文件是否存在
                if not actual_source_lot_file.exists():
                    # 尝试查找文件的其他可能位置
                    possible_paths = [
                        actual_source_lot_file,
                        Path(__file__).parent.parent / "input" / "Source Lot.csv",
                        Path(__file__).parent.parent / "Source Lot.csv",
                        source_lot_file_path,
                    ]
                    
                    found_file = None
                    for path in possible_paths:
                        if path.exists():
                            found_file = path
                            LOGGER.warning(f"原始路径不存在: {actual_source_lot_file}，使用找到的文件: {found_file}")
                            break
                    
                    if not found_file:
                        error_msg = (
                            f"Source Lot文件不存在: {actual_source_lot_file}\n"
                            f"已尝试以下路径:\n"
                            + "\n".join(f"  - {p}" for p in possible_paths)
                        )
                        raise WorkflowError(error_msg)
                    
                    actual_source_lot_file = found_file
                
                LOGGER.info(f"读取Source Lot文件: {actual_source_lot_file}")
                source_lot_df = read_excel_file(actual_source_lot_file)
                
                LOGGER.info(f"Source Lot文件列名: {source_lot_df.columns.tolist()}")
                LOGGER.info(f"Source Lot文件共有 {len(source_lot_df)} 行数据")
                
                # 查找SourceLot列（不区分大小写）
                source_lot_col = None
                for col in source_lot_df.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS']:
                        source_lot_col = col
                        LOGGER.info(f"找到SourceLot列: '{col}'")
                        break
                
                if source_lot_col is None:
                    raise WorkflowError(f"在Source Lot文件中未找到SourceLot列。可用列: {source_lot_df.columns.tolist()}")
                
                if source_lot_df.empty:
                    raise WorkflowError("Source Lot文件为空")
                
                # 存储所有MIR结果
                mir_results = []
                
                # 循环处理每一行
                for row_index, row in source_lot_df.iterrows():
                    source_lot_value = row[source_lot_col]
                    
                    if pd.isna(source_lot_value):
                        LOGGER.warning(f"第 {row_index + 1} 行的SourceLot值为空，跳过")
                        continue
                    
                    source_lot_value = str(source_lot_value).strip()
                    
                    LOGGER.info("=" * 80)
                    LOGGER.info(f"处理第 {row_index + 1}/{len(source_lot_df)} 行: SourceLot = {source_lot_value}")
                    LOGGER.info("=" * 80)
                    
                    try:
                        # 打开File菜单 -> New MIR Request
                        LOGGER.info("开始Mole工具操作流程...")
                        success = self.mole_submitter.submit_mir_data({})
                        
                        if success:
                            # Search By VPOs模式
                            LOGGER.info("使用Search By VPOs模式...")
                            # 点击Search By VPOs按钮
                            self.mole_submitter._click_search_by_vpos_button()
                            # 填写VPO搜索对话框
                            LOGGER.info("填写VPO搜索对话框...")
                            self.mole_submitter._fill_vpo_search_dialog(source_lot_value)
                            
                            # 检查搜索结果行状态并执行相应操作
                            # 注意：Search By VPOs模式使用Select Available Rows
                            LOGGER.info("检查搜索结果行状态...")
                            self.mole_submitter._check_row_status_and_select(ui_config, use_available_rows=True)
                            
                            # 点击Submit按钮
                            LOGGER.info("点击Submit按钮...")
                            self.mole_submitter._click_submit_button()
                            
                            # 处理最终成功对话框并获取MIR号码
                            LOGGER.info("处理最终成功对话框并获取MIR号码...")
                            
                            # 如果是最后一行，增加等待时间，确保copy MIR对话框完全弹出
                            is_last_row = (row_index == len(source_lot_df) - 1)
                            if is_last_row:
                                LOGGER.info("这是最后一行，等待更长时间确保copy MIR对话框完全弹出...")
                                time.sleep(3.0)  # 额外等待3秒
                            
                            mir_number = self.mole_submitter._handle_final_success_dialog_and_get_mir()
                            
                            # 如果是最后一行，再次等待确保对话框完全处理完成
                            if is_last_row:
                                LOGGER.info("最后一行处理完成，等待copy MIR对话框完全关闭...")
                                time.sleep(2.0)  # 再等待2秒确保对话框关闭
                            
                            if mir_number:
                                # 保存该行数据和MIR号码
                                result_row = row.to_dict()
                                result_row['MIR'] = mir_number
                                mir_results.append(result_row)
                                
                                LOGGER.info(f"✅ 第 {row_index + 1} 行处理成功: SourceLot={source_lot_value}, MIR={mir_number}")
                                
                                self.results.append({
                                    'row_index': row_index,
                                    'step': 'Mole',
                                    'status': 'success',
                                    'source_lot': source_lot_value,
                                    'mir': mir_number,
                                    'timestamp': datetime.now().isoformat()
                                })
                            else:
                                LOGGER.error(f"❌ 第 {row_index + 1} 行未能获取MIR号码")
                                self.errors.append({
                                    'row_index': row_index,
                                    'step': 'Mole',
                                    'error': '未能获取MIR号码',
                                    'source_lot': source_lot_value,
                                    'timestamp': datetime.now().isoformat()
                                })
                        else:
                            error_msg = f"第 {row_index + 1} 行Mole工具操作失败"
                            LOGGER.error(f"❌ {error_msg}")
                            self.errors.append({
                                'row_index': row_index,
                                'step': 'Mole',
                                'error': error_msg,
                                'source_lot': source_lot_value,
                                'timestamp': datetime.now().isoformat()
                            })
                    
                    except Exception as e:
                        error_msg = f"第 {row_index + 1} 行处理失败: {e}"
                        LOGGER.error(f"❌ {error_msg}")
                        LOGGER.error(traceback.format_exc())
                        self.errors.append({
                            'row_index': row_index,
                            'step': 'Mole',
                            'error': str(e),
                            'source_lot': source_lot_value,
                            'timestamp': datetime.now().isoformat()
                        })
                        # 继续处理下一行，不中断整个流程
                    
                    # 在处理下一行之前，等待一下，确保界面准备好
                    if row_index < len(source_lot_df) - 1:
                        LOGGER.info("等待2秒后处理下一行...")
                        time.sleep(2.0)
            
            elif search_mode == 'units':
                # Units 模式：使用粘贴的 units 信息
                units_info = ui_config.get('units_info', '') if ui_config else ''
                if not units_info:
                    raise WorkflowError("Units 模式下必须提供 units_info（请在配置UI中粘贴Units信息）")
                
                LOGGER.info("使用Search By Units模式...")
                LOGGER.info(f"Units信息: {units_info[:100]}...")
                
                # 存储所有MIR结果
                mir_results = []
                
                # 对于 Units 模式，通常只处理一次（不是循环）
                try:
                    # 打开File菜单 -> New MIR Request
                    LOGGER.info("开始Mole工具操作流程...")
                    success = self.mole_submitter.submit_mir_data({})
                    
                    if success:
                        # 点击Search By Units按钮
                        self.mole_submitter._click_search_by_units_button()
                        # 填写Units搜索对话框
                        LOGGER.info("填写Units搜索对话框...")
                        self.mole_submitter._fill_units_search_dialog(ui_config)
                        
                        # 检查搜索结果行状态并执行相应操作
                        # 注意：Units 模式使用 Select Visible Rows（不是 Select Available Rows）
                        LOGGER.info("检查搜索结果行状态...")
                        self.mole_submitter._check_row_status_and_select(ui_config, use_available_rows=False)
                        
                        # 点击Submit按钮
                        LOGGER.info("点击Submit按钮...")
                        self.mole_submitter._click_submit_button()
                        
                        # 处理最终成功对话框并获取MIR号码
                        LOGGER.info("处理最终成功对话框并获取MIR号码...")
                        # 增加等待时间，确保对话框完全弹出
                        LOGGER.info("等待copy MIR对话框完全弹出...")
                        time.sleep(4.0)  # 增加等待时间到4秒
                        mir_number = self.mole_submitter._handle_final_success_dialog_and_get_mir()
                        # 等待对话框关闭和剪贴板更新
                        time.sleep(2.0)  # 等待对话框关闭
                        
                        if mir_number:
                            mir_results.append({
                                'units_info': units_info[:50] + '...' if len(units_info) > 50 else units_info,
                                'MIR': mir_number
                            })
                            
                            LOGGER.info(f"✅ Units处理成功: MIR={mir_number}")
                            
                            self.results.append({
                                'step': 'Mole',
                                'status': 'success',
                                'units_info': units_info[:50] + '...' if len(units_info) > 50 else units_info,
                                'mir': mir_number,
                                'timestamp': datetime.now().isoformat()
                            })
                        else:
                            LOGGER.error("❌ 未能获取MIR号码")
                            self.errors.append({
                                'step': 'Mole',
                                'error': '未能获取MIR号码',
                                'timestamp': datetime.now().isoformat()
                            })
                    else:
                        error_msg = "Mole工具操作失败"
                        LOGGER.error(f"❌ {error_msg}")
                        self.errors.append({
                            'step': 'Mole',
                            'error': error_msg,
                            'timestamp': datetime.now().isoformat()
                        })
                
                except Exception as e:
                    error_msg = f"Units处理失败: {e}"
                    LOGGER.error(f"❌ {error_msg}")
                    LOGGER.error(traceback.format_exc())
                    self.errors.append({
                        'step': 'Mole',
                        'error': str(e),
                        'timestamp': datetime.now().isoformat()
                    })
                
                # 保存 Units 模式的 MIR 结果
                if mir_results:
                    LOGGER.info(f"保存Units模式的MIR结果...")
                    self._save_all_mir_results(source_lot_file_path, mir_results)
                    # 注意：_save_all_mir_results 方法已经设置了 self.last_mir_result_file，不需要重复设置
            
            elif search_mode == 'units_by_source_lot':
                # Units by Source Lot 模式：从文件读取 units，按 source lot 分组处理
                units_file_path = ui_config.get('units_file', '') if ui_config else ''
                if not units_file_path:
                    raise WorkflowError("Units by Source Lot 模式下必须提供 units_file（请选择包含 units 和 source lot 的文件）")
                
                units_file = Path(units_file_path)
                if not units_file.is_absolute():
                    base_dir = Path(__file__).parent.parent
                    units_file = (base_dir / units_file_path).resolve()
                
                if not units_file.exists():
                    raise WorkflowError(f"Units 文件不存在: {units_file}")
                
                LOGGER.info(f"读取 Units 文件: {units_file}")
                units_df = read_excel_file(units_file)
                
                LOGGER.info(f"Units 文件列名: {units_df.columns.tolist()}")
                LOGGER.info(f"Units 文件共有 {len(units_df)} 行数据")
                
                # 查找 Source Lot 列（支持 Source, SourceLot 等）
                source_lot_col = None
                for col in units_df.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper in ['SOURCE', 'SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT', 'SOURCELOTS', 'SOURCE LOTS']:
                        source_lot_col = col
                        LOGGER.info(f"找到 SourceLot 列: '{col}'")
                        break
                
                if source_lot_col is None:
                    raise WorkflowError(f"在 Units 文件中未找到 SourceLot 列。可用列: {units_df.columns.tolist()}")
                
                # 查找 Unit Name 列
                unit_name_col = None
                for col in units_df.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper in ['UNIT NAME', 'UNITNAME', 'UNIT_NAME', 'UNIT', 'UNITS', 'UNIT ID', 'UNITID']:
                        unit_name_col = col
                        LOGGER.info(f"找到 Unit Name 列: '{col}'")
                        break
                
                if unit_name_col is None:
                    raise WorkflowError(f"在 Units 文件中未找到 Unit Name 列。可用列: {units_df.columns.tolist()}")
                
                # 使用工作目录的验证和导出子目录
                # 使用 MIR 目录存储临时导出和验证文件
                temp_export_dir = self.work_subdirs['mir']
                validation_dir = self.work_subdirs['mir']
                temp_export_dir.mkdir(parents=True, exist_ok=True)
                validation_dir.mkdir(parents=True, exist_ok=True)
                LOGGER.info(f"导出文件将保存到: {temp_export_dir}")
                LOGGER.info(f"验证结果将保存到: {validation_dir}")
                
                # 第一步：收集所有units（不按source lot分组），验证它们是否都available
                LOGGER.info("=" * 80)
                LOGGER.info("第一步：验证所有units是否available（不按source lot分组）")
                LOGGER.info("=" * 80)
                
                # 收集所有units（去重）
                all_units_list = units_df[unit_name_col].dropna().astype(str).tolist()
                all_units_list = [u.strip() for u in all_units_list if u.strip()]
                all_units_list = list(set(all_units_list))  # 去重
                
                LOGGER.info(f"共收集到 {len(all_units_list)} 个唯一的units（去重后）")
                
                # 打开Mole工具进行验证
                LOGGER.info("打开Mole工具进行验证...")
                success = self.mole_submitter.submit_mir_data({})
                
                if not success:
                    raise WorkflowError("无法打开Mole工具")
                
                # 点击Search By Units按钮
                self.mole_submitter._click_search_by_units_button()
                
                # 填写所有units进行验证
                temp_ui_config_validation = ui_config.copy() if ui_config else {}
                temp_ui_config_validation['units_info'] = '\n'.join(all_units_list)
                LOGGER.info(f"填写所有units进行验证 ({len(all_units_list)} 个)...")
                self.mole_submitter._fill_units_search_dialog(temp_ui_config_validation)
                
                # 等待搜索结果
                time.sleep(3.0)
                
                # 处理可能出现的 "Inactive Source Lots" 对话框
                LOGGER.info("等待并检查 'Inactive Source Lots' 对话框...")
                time.sleep(2.0)
                dialog_handled = self.mole_submitter._handle_inactive_source_lots_dialog(max_wait_time=8)
                if dialog_handled:
                    LOGGER.info("✅ 已处理 'Inactive Source Lots' 对话框")
                else:
                    LOGGER.debug("未检测到 'Inactive Source Lots' 对话框，继续执行")
                time.sleep(1.0)
                
                # 导出并获取实际可用的units（使用 Select Visible Rows）
                LOGGER.info("导出并获取实际可用的units（使用 Select Visible Rows 包含所有可见 units）...")
                validation_info = self.mole_submitter._get_actual_units_count_from_export(
                    expected_units=all_units_list,
                    temp_export_dir=temp_export_dir,
                    source_lot="ALL_UNITS_VALIDATION",
                    use_visible_rows=True  # Units模式只能使用 Select Visible Rows
                )
                
                # 获取实际可用的units列表（保存供后续使用，不再重复导出）
                available_units_list = validation_info.get('actual_units', [])
                available_units_set = set(str(u).strip() for u in available_units_list)
                validation_count = validation_info.get('actual_count', 0)
                missing_units_validation = validation_info.get('missing_units', [])
                
                # 保存第一步验证的结果供后续使用
                self.available_units_set = available_units_set
                
                # 保存第一步验证的导出文件路径，并立即写入合并Excel文件
                export_file = validation_info.get('export_file')
                if export_file and Path(export_file).exists():
                    self.available_units_export_file = Path(export_file)
                    LOGGER.debug(f"临时文件（将合并到 Merged_Validation_Table）: {self.available_units_export_file.name}")
                    
                    # 不再立即写入，等最后合并时一起写入
                
                # 创建第一步验证的units比较结果（包含source lot信息）
                validation_comparison_list = []
                all_units_set = set(str(u).strip() for u in all_units_list)
                
                # 创建一个字典，记录每个unit对应的source lots（一个unit可能属于多个source lot）
                unit_to_source_lots = {}
                for _, row in units_df.iterrows():
                    unit_str = str(row[unit_name_col]).strip() if pd.notna(row[unit_name_col]) else ''
                    source_lot_str = str(row[source_lot_col]).strip() if pd.notna(row[source_lot_col]) else ''
                    if unit_str:
                        if unit_str not in unit_to_source_lots:
                            unit_to_source_lots[unit_str] = []
                        if source_lot_str and source_lot_str not in unit_to_source_lots[unit_str]:
                            unit_to_source_lots[unit_str].append(source_lot_str)
                
                # 输入的units（期望的）
                for unit in all_units_list:
                    unit_str = str(unit).strip()
                    is_matched = unit_str in available_units_set
                    source_lots = ', '.join(unit_to_source_lots.get(unit_str, ['N/A']))
                    validation_comparison_list.append({
                        'Source Lot': source_lots,
                        'Unit Name': unit_str,
                        'Status': 'Matched' if is_matched else 'Missing',
                        'In_Input_File': 'Yes',
                        'In_Mole_Export': 'Yes' if is_matched else 'No'
                    })
                
                # Mole导出中有但输入文件中没有的units（额外发现的）
                extra_units_validation = available_units_set - all_units_set
                for unit in extra_units_validation:
                    validation_comparison_list.append({
                        'Source Lot': 'N/A',
                        'Unit Name': str(unit).strip(),
                        'Status': 'Extra',
                        'In_Input_File': 'No',
                        'In_Mole_Export': 'Yes'
                    })
                
                # 保存第一步验证的比较结果到文件
                if validation_comparison_list:
                    from datetime import datetime
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    validation_output_file = validation_dir / f"Units_Validation_Comparison_{timestamp}.xlsx"
                    
                    try:
                        validation_comparison_df = pd.DataFrame(validation_comparison_list)
                        # 按Source Lot, Status和Unit Name排序
                        validation_comparison_df = validation_comparison_df.sort_values(
                            by=['Source Lot', 'Status', 'Unit Name'],
                            ascending=[True, True, True]
                        )
                        
                        # 创建Source Lot统计表
                        source_lot_stats_list = []
                        grouped_by_source_lot = units_df.groupby(source_lot_col)
                        for source_lot_value, group_df in grouped_by_source_lot:
                            if pd.isna(source_lot_value):
                                continue
                            source_lot_value = str(source_lot_value).strip()
                            
                            # 获取该source lot的所有units
                            source_lot_units = group_df[unit_name_col].dropna().astype(str).tolist()
                            source_lot_units = [u.strip() for u in source_lot_units if u.strip()]
                            
                            # 统计匹配和缺失的units
                            matched_units = [u for u in source_lot_units if str(u).strip() in available_units_set]
                            missing_units = [u for u in source_lot_units if str(u).strip() not in available_units_set]
                            
                            source_lot_stats_list.append({
                                'Source Lot': source_lot_value,
                                'Total_Units': len(source_lot_units),
                                'Matched_Units': len(matched_units),
                                'Missing_Units': len(missing_units),
                                'Match_Rate': f"{len(matched_units)/len(source_lot_units)*100:.1f}%" if source_lot_units else "0.0%"
                            })
                        
                        source_lot_stats_df = pd.DataFrame(source_lot_stats_list)
                        # 按Source Lot排序
                        source_lot_stats_df = source_lot_stats_df.sort_values(by='Source Lot', ascending=True)
                        
                        # 不单独保存 Units_Validation_Comparison 文件，数据将直接用于合并到 Merged_Validation_Table
                        # 创建一个临时文件路径用于后续合并逻辑（合并后会被删除）
                        validation_output_file = validation_dir / f"Units_Validation_Comparison_{timestamp}.xlsx"
                        
                        # 将数据保存到临时文件（仅用于后续读取和合并，合并后会被删除）
                        try:
                            with pd.ExcelWriter(validation_output_file, engine='openpyxl') as writer:
                                source_lot_stats_df.to_excel(writer, sheet_name='Source Lot Summary', index=False)
                                validation_comparison_df.to_excel(writer, sheet_name='All Units Comparison', index=False)
                                missing_df = validation_comparison_df[validation_comparison_df['Status'] == 'Missing'].copy()
                                if not missing_df.empty:
                                    missing_df.to_excel(writer, sheet_name='Missing Units', index=False)
                                extra_df = validation_comparison_df[validation_comparison_df['Status'] == 'Extra'].copy()
                                if not extra_df.empty:
                                    extra_df.to_excel(writer, sheet_name='Extra Units', index=False)
                            self.units_validation_comparison_file = validation_output_file
                            LOGGER.debug(f"临时文件（将合并到 Merged_Validation_Table）: {self.units_validation_comparison_file.name}")
                        except Exception as e:
                            LOGGER.warning(f"创建临时验证比较文件失败: {e}")
                        
                        # 立即合并两张表并生成汇总表（在try块外，确保即使保存失败也能执行）
                        if self.available_units_export_file and self.available_units_export_file.exists() and self.units_validation_comparison_file and self.units_validation_comparison_file.exists():
                            # 获取 source_lot_file_path（用于读取 Part Type 和 Quantity）
                            # 在 units_by_source_lot 模式下，需要从 units_file 或 source_lot_file_path 获取
                            actual_source_lot_file = source_lot_file_path
                            if not actual_source_lot_file or not actual_source_lot_file.exists():
                                # 尝试从配置中获取
                                base_dir = Path(__file__).parent.parent
                                possible_paths = [
                                    base_dir / "input" / "Source Lot.csv",
                                    base_dir / "Source Lot.csv",
                                    self.config.paths.source_lot_file if hasattr(self.config, 'paths') else None
                                ]
                                for path in possible_paths:
                                    if path and Path(path).exists():
                                        actual_source_lot_file = Path(path)
                                        break
                            
                            if actual_source_lot_file and actual_source_lot_file.exists():
                                LOGGER.info("=" * 80)
                                LOGGER.info("立即合并 available_units_export 和 Units_Validation_Comparison 表...")
                                LOGGER.info("=" * 80)
                                self._merge_validation_tables_immediately(actual_source_lot_file, units_df, source_lot_col, unit_name_col)
                            else:
                                LOGGER.warning(f"未找到 Source Lot 文件，无法合并表。尝试使用 units_df 中的信息...")
                                # 如果没有 source_lot_file，尝试使用 units_df 中的信息
                                self._merge_validation_tables_immediately(None, units_df, source_lot_col, unit_name_col)
                        LOGGER.info(f"   总计: {len(validation_comparison_list)} 个units")
                        matched_count = len([u for u in validation_comparison_list if u.get('Status') == 'Matched'])
                        missing_count = len([u for u in validation_comparison_list if u.get('Status') == 'Missing'])
                        extra_count = len([u for u in validation_comparison_list if u.get('Status') == 'Extra'])
                        LOGGER.info(f"     - 匹配: {matched_count} 个")
                        LOGGER.info(f"     - 缺失: {missing_count} 个")
                        LOGGER.info(f"     - 额外: {extra_count} 个")
                        LOGGER.info(f"   Source Lot统计: {len(source_lot_stats_df)} 个Source Lot")
                        total_units_by_source_lot = source_lot_stats_df['Total_Units'].sum()
                        total_matched_by_source_lot = source_lot_stats_df['Matched_Units'].sum()
                        LOGGER.info(f"     - 总计Units: {total_units_by_source_lot} 个")
                        LOGGER.info(f"     - 总计匹配: {total_matched_by_source_lot} 个")
                        LOGGER.info(f"     - 总计缺失: {total_units_by_source_lot - total_matched_by_source_lot} 个")
                    except Exception as e:
                        LOGGER.warning(f"保存第一步验证比较结果失败: {e}")
                        # 如果Excel保存失败，尝试保存为CSV
                        try:
                            csv_file = validation_dir / f"Units_Validation_Comparison_{timestamp}.csv"
                            validation_comparison_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
                            # 保存验证比较文件路径（用于后续合并）
                            self.units_validation_comparison_file = csv_file
                            LOGGER.info(f"✅ 第一步验证的Units比较结果已保存到: {csv_file} (CSV格式)")
                        except Exception as e2:
                            LOGGER.error(f"保存CSV文件也失败: {e2}")
                
                LOGGER.info("=" * 80)
                LOGGER.info("第一步验证结果:")
                LOGGER.info(f"  期望Units数量: {len(all_units_list)} 个")
                LOGGER.info(f"  实际可用Units数量: {validation_count} 个")
                if len(all_units_list) != validation_count:
                    LOGGER.warning(f"  差异: {len(all_units_list) - validation_count} 个units不可用")
                    if missing_units_validation:
                        LOGGER.warning(f"  缺失的units ({len(missing_units_validation)} 个):")
                        for missing in missing_units_validation[:10]:
                            LOGGER.warning(f"    - {missing}")
                        if len(missing_units_validation) > 10:
                            LOGGER.warning(f"    ... 还有 {len(missing_units_validation) - 10} 个")
                else:
                    LOGGER.info("  ✅ 所有units都可用")
                LOGGER.info("=" * 80)
                
                # 按 Source Lot 分组，使用原始units提交MIR
                grouped = units_df.groupby(source_lot_col)
                LOGGER.info(f"共找到 {len(grouped)} 个不同的 Source Lot")
                
                # 存储所有MIR结果
                mir_results = []
                
                # 初始化unit对比详情列表（用于保存详细的unit对比信息）
                self.unit_comparison_details = []
                
                # 循环处理每个 Source Lot 组
                for source_lot_value, group_df in grouped:
                    if pd.isna(source_lot_value):
                        LOGGER.warning(f"跳过 SourceLot 为空的组")
                        continue
                    
                    source_lot_value = str(source_lot_value).strip()
                    
                    # 提取该 Source Lot 的原始units
                    original_units_list = group_df[unit_name_col].dropna().astype(str).tolist()
                    original_units_list = [u.strip() for u in original_units_list if u.strip()]
                    
                    if not original_units_list:
                        LOGGER.warning(f"SourceLot {source_lot_value} 没有有效的units，跳过")
                        continue
                    
                    # 从第一步验证的available units中筛选出该source lot中可用的units
                    available_units_for_source_lot = [u for u in original_units_list if str(u).strip() in self.available_units_set]
                    
                    if not available_units_for_source_lot:
                        LOGGER.warning(f"SourceLot {source_lot_value} 没有可用的units（所有units都不可用），跳过")
                        continue
                    
                    LOGGER.info("=" * 80)
                    LOGGER.info(f"处理 SourceLot: {source_lot_value}")
                    LOGGER.info(f"  原始Units数量: {len(original_units_list)} 个")
                    LOGGER.info(f"  可用Units数量: {len(available_units_for_source_lot)} 个")
                    if len(original_units_list) != len(available_units_for_source_lot):
                        missing_count = len(original_units_list) - len(available_units_for_source_lot)
                        LOGGER.warning(f"  跳过不可用的units: {missing_count} 个")
                    LOGGER.info("=" * 80)
                    
                    try:
                        # 使用available的units搜索并提交MIR（直接使用已打开的Mole工具，不再打开新的MIR请求）
                        LOGGER.info("使用available的units搜索并提交MIR...")
                        LOGGER.info("直接使用已打开的Mole工具，跳过打开新MIR请求的步骤")
                        
                        # 直接点击Search By Units按钮（Mole工具已经在第一步验证时打开）
                        self.mole_submitter._click_search_by_units_button()
                        
                        # 填写该source lot的available units（只使用可用的units）
                        temp_ui_config_actual = ui_config.copy() if ui_config else {}
                        temp_ui_config_actual['units_info'] = '\n'.join(available_units_for_source_lot)
                        LOGGER.info(f"填写SourceLot {source_lot_value} 的available units ({len(available_units_for_source_lot)} 个)...")
                        self.mole_submitter._fill_units_search_dialog(temp_ui_config_actual)
                        
                        # 等待搜索结果
                        time.sleep(3.0)
                        
                        # 处理可能出现的 "Inactive Source Lots" 对话框
                        LOGGER.info("等待并检查 'Inactive Source Lots' 对话框...")
                        time.sleep(2.0)
                        dialog_handled = self.mole_submitter._handle_inactive_source_lots_dialog(max_wait_time=8)
                        if dialog_handled:
                            LOGGER.info("✅ 已处理 'Inactive Source Lots' 对话框")
                        else:
                            LOGGER.debug("未检测到 'Inactive Source Lots' 对话框，继续执行")
                        time.sleep(1.0)
                        
                        # 使用第一步验证的结果，不再重复导出
                        # available_units_for_source_lot 已经在上面筛选出来了
                        expected_set = set(str(u).strip() for u in original_units_list)
                        actual_set = set(str(u).strip() for u in available_units_for_source_lot)
                        actual_count = len(available_units_for_source_lot)
                        missing_units = list(expected_set - actual_set)
                        
                        LOGGER.info(f"  该Source Lot中可用units数量: {actual_count} / {len(original_units_list)}")
                        if missing_units:
                            LOGGER.warning(f"  缺失的units ({len(missing_units)} 个): {', '.join(missing_units[:5])}{'...' if len(missing_units) > 5 else ''}")
                        
                        # 按Unit Name对比，创建详细的units对比信息
                        # 为每个输入的unit创建对比记录
                        unit_comparison_list = []
                        
                        # 输入的units（期望的）
                        for unit in original_units_list:
                            unit_str = str(unit).strip()
                            is_matched = unit_str in actual_set
                            unit_comparison_list.append({
                                'Source Lot': source_lot_value,
                                'Unit Name': unit_str,
                                'Status': 'Matched' if is_matched else 'Missing',
                                'In_Input': 'Yes',
                                'In_Export': 'Yes' if is_matched else 'No'
                            })
                        
                        # 注意：不再检查额外units，因为第一步验证已经包含了所有units
                        
                        # 保存详细的units对比信息（用于后续汇总）
                        if not hasattr(self, 'unit_comparison_details'):
                            self.unit_comparison_details = []
                        self.unit_comparison_details.extend(unit_comparison_list)
                        
                        # 检查搜索结果行状态并执行相应操作
                        LOGGER.info("检查搜索结果行状态...")
                        self.mole_submitter._check_row_status_and_select(
                            temp_ui_config_actual, 
                            use_available_rows=False  # Units 模式使用 Select Visible Rows
                        )
                        
                        # 点击Submit按钮
                        LOGGER.info("点击Submit按钮...")
                        self.mole_submitter._click_submit_button()
                        
                        # 获取MIR号码（内部会轮询等待对话框出现）
                        LOGGER.info("等待对话框出现并获取MIR号码...")
                        mir_number = self.mole_submitter._handle_final_success_dialog_and_get_mir()
                        
                        # 等待对话框完全关闭和剪贴板更新
                        time.sleep(2.0)
                        
                        if mir_number:
                            # 保存结果，只包含统计信息（配置字段在Spark阶段从For Spark.csv读取）
                            result_row = {
                                'Source Lot': source_lot_value,
                                'Units_Count_Expected': len(original_units_list),
                                'Units_Count_Actual': actual_count,
                                'Units_Count_Diff': len(original_units_list) - actual_count,
                                'Count_Match': 'Yes' if len(original_units_list) == actual_count else 'No',
                                'Missing_Units_Count': len(missing_units),
                                'Missing_Units': ', '.join(missing_units[:5]) + ('...' if len(missing_units) > 5 else '') if missing_units else '',
                                'Units_Sample': ', '.join(original_units_list[:5]) + ('...' if len(original_units_list) > 5 else ''),
                                'MIR': mir_number,
                                'Export_File': ''  # 不再导出，使用第一步验证的结果
                            }
                            mir_results.append(result_row)
                            
                            LOGGER.info(f"✅ SourceLot {source_lot_value} 处理成功，MIR号码已保存:")
                            LOGGER.info(f"   期望Units数量: {len(original_units_list)} 个")
                            LOGGER.info(f"   实际可用Units数量: {actual_count} 个")
                            if len(original_units_list) != actual_count:
                                LOGGER.warning(f"   差异: {len(original_units_list) - actual_count} 个units不可用")
                            LOGGER.info(f"   MIR: {mir_number}")
                            LOGGER.info(f"   MIR号码已保存到结果列表，准备处理下一个Source Lot...")
                            
                            self.results.append({
                                'step': 'Mole',
                                'status': 'success',
                                'source_lot': source_lot_value,
                                'units_count_expected': len(original_units_list),
                                'units_count_actual': actual_count,
                                'mir': mir_number,
                                'timestamp': datetime.now().isoformat()
                            })
                        else:
                            LOGGER.error(f"❌ SourceLot {source_lot_value} 未能获取MIR号码")
                            self.errors.append({
                                'step': 'Mole',
                                'error': '未能获取MIR号码',
                                'source_lot': source_lot_value,
                                'timestamp': datetime.now().isoformat()
                            })
                    
                    except Exception as e:
                        error_msg = f"SourceLot {source_lot_value} 处理失败: {e}"
                        LOGGER.error(f"❌ {error_msg}")
                        LOGGER.error(traceback.format_exc())
                        self.errors.append({
                            'step': 'Mole',
                            'error': str(e),
                            'source_lot': source_lot_value,
                            'timestamp': datetime.now().isoformat()
                        })
                    
                    # 在处理下一个 Source Lot 之前，等待一下
                    if source_lot_value != list(grouped.groups.keys())[-1]:
                        LOGGER.info("等待2秒后处理下一个 Source Lot...")
                        time.sleep(2.0)
                
                # 保存 Units by Source Lot 模式的 MIR 结果
                if mir_results:
                    LOGGER.info(f"保存 Units by Source Lot 模式的MIR结果...")
                    self._save_all_mir_results(source_lot_file_path, mir_results)
                    
                    # 合并 available_units_export 和 Units_Validation_Comparison 表
                    if self.available_units_export_file and self.units_validation_comparison_file:
                        LOGGER.info("=" * 80)
                        LOGGER.info("合并 available_units_export 和 Units_Validation_Comparison 表...")
                        LOGGER.info("=" * 80)
                        self._merge_validation_tables(source_lot_file_path, mir_results)
            
            else:
                raise WorkflowError(f"不支持的搜索模式: {search_mode}")
            
            # 保存 VPOs 模式的 MIR 结果（如果有）
            if search_mode == 'vpos' and mir_results:
                LOGGER.info(f"保存VPOs模式的MIR结果...")
                self._save_all_mir_results(source_lot_file_path, mir_results)
                # 注意：_save_all_mir_results 方法已经设置了 self.last_mir_result_file，不需要重复设置
                
                # VPOs 模式：直接生成汇总表（source lot, Part Type, quantity, MIR）
                LOGGER.info("=" * 80)
                LOGGER.info("生成 VPOs 模式的汇总表...")
                LOGGER.info("=" * 80)
                self._generate_summary_table_for_vpos(source_lot_file_path, mir_results)
            
            # 显示处理结果统计
            LOGGER.info("=" * 80)
            LOGGER.info("Mole处理完成")
            LOGGER.info("=" * 80)
            if search_mode == 'vpos':
                LOGGER.info(f"  成功处理: {len(mir_results)} 行")
                LOGGER.info(f"  失败: {len(self.errors)} 行")
            elif search_mode == 'units_by_source_lot':
                LOGGER.info(f"  Units by Source Lot 模式处理完成")
                LOGGER.info(f"  成功处理: {len(mir_results)} 个 Source Lot")
                LOGGER.info(f"  失败: {len(self.errors)} 个 Source Lot")
            else:
                LOGGER.info(f"  Units模式处理完成")
                if mir_results:
                    LOGGER.info(f"  成功: 1 个MIR")
                else:
                    LOGGER.info(f"  失败: 未能获取MIR")
            
            if mir_results:
                LOGGER.info(f"  结果文件: {self.last_mir_result_file}")
            LOGGER.info("=" * 80)
            
            # 所有行处理完后，等待所有对话框关闭，然后关闭MOLE
            LOGGER.info("=" * 80)
            LOGGER.info("所有MIR提交完成，等待所有对话框关闭...")
            LOGGER.info("=" * 80)
            
            # 等待所有成功对话框关闭（最多等待10秒）
            if win32gui:
                max_wait = 10
                for i in range(max_wait):
                    try:
                        def check_dialog(hwnd, dialogs):
                            try:
                                if not win32gui.IsWindowVisible(hwnd):
                                    return True
                                window_text = win32gui.GetWindowText(hwnd)
                                if window_text == "Submit MIR Request":
                                    dialogs.append(hwnd)
                            except:
                                pass
                            return True
                        
                        remaining_dialogs = []
                        win32gui.EnumWindows(check_dialog, remaining_dialogs)
                        if not remaining_dialogs:
                            LOGGER.info("✅ 所有对话框已关闭")
                            break
                        else:
                            if i % 2 == 0:
                                LOGGER.info(f"等待对话框关闭... ({i+1}/{max_wait}秒，还有{len(remaining_dialogs)}个对话框)")
                            time.sleep(1.0)
                    except:
                        time.sleep(1.0)
            
            # 关闭MOLE工具（已禁用，用户要求不关闭Mole）
            # LOGGER.info("=" * 80)
            # LOGGER.info("关闭MOLE工具...")
            # LOGGER.info("=" * 80)
            # try:
            #     self._close_mole()
            #     LOGGER.info("✅ MOLE工具已关闭")
            # except Exception as e:
            #     LOGGER.warning(f"⚠️ 关闭MOLE工具时出错: {e}，继续执行...")
            LOGGER.info("⚠️ 已跳过关闭MOLE工具（用户要求保持Mole打开）")
            
            # 注意：MIR结果已经在第686行（Units模式）或第695行（VPOs模式）保存过了，不需要重复保存
            
            # 输出汇总信息
            LOGGER.info("=" * 80)
            if search_mode == 'vpos':
                LOGGER.info(f"处理汇总:")
                LOGGER.info(f"  总行数: {len(source_lot_df) if 'source_lot_df' in locals() else 'N/A'}")
                LOGGER.info(f"  成功: {len(mir_results)}")
                LOGGER.info(f"  失败: {len(self.errors)}")
            else:
                LOGGER.info(f"处理汇总:")
                LOGGER.info(f"  成功: {len(mir_results)}")
                LOGGER.info(f"  失败: {len(self.errors)}")
            LOGGER.info("=" * 80)
                
        except Exception as e:
            raise WorkflowError(f"提交MIR数据到Mole工具失败: {e}")
    
    def _step_submit_to_spark(self, df: pd.DataFrame) -> None:
        """步骤3: 提交VPO数据到Spark网页（从MIR结果文件读取数据）"""
        try:
            # 优先查找本次运行工作目录中的MIR结果文件，如果没有则查找output目录
            LOGGER.info("查找MIR结果文件...")
            
            # 首先在工作目录的mir_results子目录中查找（新结构：01_MIR）
            mir_results_dir = self.work_subdirs.get('mir', self.work_dir / '01_MIR')
            
            # 优先查找Excel文件，如果没有则查找CSV文件
            mir_excel_files = sorted(mir_results_dir.glob("MIR_Results_*.xlsx"), reverse=True)
            mir_csv_files = sorted(mir_results_dir.glob("MIR_Results_*.csv"), reverse=True)
            
            # 如果工作目录中没有，则在output/01_MIR目录中查找（新结构）
            if not mir_excel_files and not mir_csv_files:
                output_dir = self.config.paths.output_dir
                mir_dir = output_dir / "01_MIR"
                if mir_dir.exists():
                    mir_excel_files = sorted(mir_dir.glob("MIR_Results_*.xlsx"), reverse=True)
                    mir_csv_files = sorted(mir_dir.glob("MIR_Results_*.csv"), reverse=True)
            
            # 向后兼容：在output根目录中查找（旧格式）
            if not mir_excel_files and not mir_csv_files:
                output_dir = self.config.paths.output_dir
                if output_dir.exists():
                    mir_excel_files = sorted(output_dir.glob("MIR_Results_*.xlsx"), reverse=True)
                    mir_csv_files = sorted(output_dir.glob("MIR_Results_*.csv"), reverse=True)
            
            if not mir_excel_files and not mir_csv_files:
                raise WorkflowError(f"未找到MIR结果文件。请先完成Mole步骤。\n已检查目录: {mir_results_dir}\n支持格式: .xlsx, .csv")
            
            if mir_excel_files:
                selected_file = mir_excel_files[0]
                LOGGER.info(f"使用MIR结果文件（Excel格式）: {selected_file.name}")
            elif mir_csv_files:
                selected_file = mir_csv_files[0]
                LOGGER.info(f"使用MIR结果文件（CSV格式）: {selected_file.name}")
            else:
                raise WorkflowError(f"未在output目录找到MIR结果文件，无法提交到Spark。请先完成Mole步骤。\n已检查目录: {output_dir}\n支持格式: .xlsx, .csv")
            
            # 读取MIR结果文件
            mir_df = read_excel_file(selected_file)
            if mir_df.empty:
                raise WorkflowError("MIR结果文件为空")
            
            LOGGER.info(f"成功读取MIR数据：{len(mir_df)} 行")
            LOGGER.info(f"MIR文件列名: {mir_df.columns.tolist()}")
            
            # 读取 For Spark.csv 文件，建立 Source Lot -> 配置的映射
            spark_config_df = None
            spark_config_map = {}  # {source_lot: {operation, eng_id, ...}}
            
            # 查找 For Spark.csv 文件（在input目录或父目录）
            base_dir = Path(__file__).parent.parent
            possible_spark_config_paths = [
                base_dir / "input" / "For Spark.csv",
                base_dir / "For Spark.csv",
                self.config.paths.input_dir / "For Spark.csv"
            ]
            
            spark_config_file = None
            for path in possible_spark_config_paths:
                if path.exists():
                    spark_config_file = path
                    break
            
            # 合并 MIR 结果和 For Spark.csv，生成汇总文件
            LOGGER.info("=" * 80)
            LOGGER.info("合并 MIR 结果和 For Spark.csv...")
            LOGGER.info("=" * 80)
            
            merged_df = self._merge_mir_with_spark_config(mir_df, spark_config_file)
            
            # 保存汇总文件到 output 根目录（总是生成新的，确保是最新的合并结果）
            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            merged_file = self.work_subdirs['spark'] / f"MIR_Results_For_Spark_{date_str}.xlsx"
            
            # 确保 output 目录存在
            self.config.paths.output_dir.mkdir(parents=True, exist_ok=True)
            
            try:
                merged_df.to_excel(merged_file, index=False, engine='openpyxl')
                LOGGER.info(f"✅ 已生成合并文件: {merged_file}")
                LOGGER.info(f"   完整路径: {merged_file.absolute()}")
                LOGGER.info(f"   包含 {len(merged_df)} 行数据")
                LOGGER.info(f"   列: {merged_df.columns.tolist()}")
                
                # 验证文件是否真的被创建
                if merged_file.exists():
                    file_size = merged_file.stat().st_size
                    LOGGER.info(f"   文件大小: {file_size} 字节")
                else:
                    LOGGER.error(f"❌ 文件未成功创建: {merged_file}")
            except Exception as e:
                # 如果Excel保存失败，尝试保存为CSV
                LOGGER.warning(f"保存Excel文件失败: {e}，尝试保存为CSV格式...")
                import traceback
                LOGGER.debug(traceback.format_exc())
                
                merged_file = self.work_subdirs['spark'] / f"MIR_Results_For_Spark_{date_str}.csv"
                try:
                    merged_df.to_csv(merged_file, index=False, encoding='utf-8-sig')
                    LOGGER.info(f"✅ 已生成合并文件: {merged_file}")
                    LOGGER.info(f"   完整路径: {merged_file.absolute()}")
                    LOGGER.info(f"   包含 {len(merged_df)} 行数据")
                    
                    # 验证文件是否真的被创建
                    if merged_file.exists():
                        file_size = merged_file.stat().st_size
                        LOGGER.info(f"   文件大小: {file_size} 字节")
                    else:
                        LOGGER.error(f"❌ 文件未成功创建: {merged_file}")
                except Exception as e2:
                    LOGGER.error(f"❌ 保存CSV文件也失败: {e2}")
                    LOGGER.error(traceback.format_exc())
                    raise WorkflowError(f"无法保存合并文件: {e2}")
            
            # 使用汇总后的DataFrame继续处理
            mir_df = merged_df
            
            if spark_config_file:
                try:
                    LOGGER.info(f"读取 Spark 配置文件: {spark_config_file}")
                    spark_config_df = read_excel_file(spark_config_file)
                    
                    # 查找 SourceLot 列
                    spark_source_lot_col = None
                    for col in spark_config_df.columns:
                        col_upper = str(col).strip().upper()
                        if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT']:
                            spark_source_lot_col = col
                            break
                    
                    if spark_source_lot_col:
                        # 建立映射
                        for _, row in spark_config_df.iterrows():
                            source_lot = str(row[spark_source_lot_col]).strip() if pd.notna(row[spark_source_lot_col]) else None
                            if source_lot:
                                config = {}
                                
                                # 读取各个配置字段
                                for col in spark_config_df.columns:
                                    if col == spark_source_lot_col:
                                        continue
                                    col_upper = str(col).strip().upper()
                                    value = row[col]
                                    
                                    # 处理空值
                                    if pd.isna(value) or str(value).strip() == '':
                                        continue
                                    
                                    value = str(value).strip()
                                    
                                    # 映射到标准字段名
                                    if col_upper in ['OPERATION', 'OP', 'OPN']:
                                        config['operation'] = value
                                    elif col_upper in ['ENG ID', 'ENGID', 'ENG_ID', 'ENGINEERING ID', 'ENGINEERING_ID']:
                                        config['eng_id'] = value
                                    elif col_upper in ['UNIT TEST TIME', 'UNIT_TEST_TIME', 'TEST TIME']:
                                        config['unit_test_time'] = value
                                    elif col_upper in ['RETEST RATE', 'RETEST_RATE', 'RATE']:
                                        config['retest_rate'] = value
                                    elif col_upper in ['HRI / MRV', 'HRI_MRV', 'HRI', 'MRV']:
                                        config['hri_mrv'] = value
                                    elif col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE']:
                                        config['part_type'] = value
                                
                                spark_config_map[source_lot] = config
                                LOGGER.info(f"  加载配置: SourceLot={source_lot}, Operation={config.get('operation', 'N/A')}, EngID={config.get('eng_id', 'N/A')}")
                        
                        LOGGER.info(f"✅ 成功加载 {len(spark_config_map)} 个 Source Lot 的 Spark 配置")
                    else:
                        LOGGER.warning(f"⚠️ 在 For Spark.csv 中未找到 SourceLot 列，跳过加载配置")
                except Exception as e:
                    LOGGER.warning(f"⚠️ 读取 For Spark.csv 失败: {e}，将使用MIR结果文件或UI配置")
            else:
                LOGGER.info("未找到 For Spark.csv 文件，将使用MIR结果文件或UI配置中的值")
            
            # 使用上下文管理器确保WebDriver正确关闭
            with self.spark_submitter:
                # 初始化并导航到页面（只需要一次）
                LOGGER.info("初始化Spark网页...")
                self.spark_submitter._init_driver()
                self.spark_submitter._navigate_to_page()
                
                # 循环处理每一行MIR结果
                # 使用enumerate确保行号从0开始，避免DataFrame索引问题
                for row_num, (idx, row) in enumerate(mir_df.iterrows()):
                    LOGGER.info("=" * 80)
                    LOGGER.info(f"处理第 {row_num + 1}/{len(mir_df)} 行MIR数据 (DataFrame索引: {idx})")
                    LOGGER.info("=" * 80)
                    
                    try:
                        # 提取数据（支持多种列名格式）
                        LOGGER.info(f"行数据: {row.to_dict()}")
                        
                        # 查找SourceLot列
                        source_lot = None
                        for col in row.index:
                            col_upper = str(col).strip().upper()
                            if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT']:
                                source_lot = str(row[col]).strip() if pd.notna(row[col]) else ''
                                LOGGER.info(f"找到SourceLot列: '{col}' = '{source_lot}'")
                                break
                        
                        if not source_lot:
                            LOGGER.warning(f"第 {row_num + 1} 行SourceLot值为空，跳过")
                            LOGGER.warning(f"可用列: {row.index.tolist()}")
                            continue
                        
                        # 查找Part Type列（优先级：For Spark.csv > MIR结果文件）
                        part_type = None
                        
                        # 优先从 For Spark.csv 读取
                        if source_lot in spark_config_map and 'part_type' in spark_config_map[source_lot]:
                            part_type = spark_config_map[source_lot]['part_type']
                            LOGGER.info(f"从 For Spark.csv 读取 Part Type: {part_type}")
                        else:
                            # 从MIR结果文件读取
                            for col in row.index:
                                col_upper = str(col).strip().upper()
                                if col_upper in ['PART TYPE', 'PARTTYPE', 'PART_TYPE']:
                                    if pd.notna(row[col]) and str(row[col]).strip():
                                        part_type = str(row[col]).strip()
                                        LOGGER.info(f"从MIR结果文件读取 Part Type: {part_type}")
                                    break
                        
                        if not part_type:
                            LOGGER.warning(f"第 {row_num + 1} 行Part Type值为空，跳过")
                            continue
                        
                        # 查找Operation列（优先级：For Spark.csv > MIR结果文件 > UI配置）
                        operation = None
                        
                        # 优先级1: 从 For Spark.csv 读取
                        if source_lot in spark_config_map and 'operation' in spark_config_map[source_lot]:
                            operation = spark_config_map[source_lot]['operation']
                            LOGGER.info(f"从 For Spark.csv 读取 Operation: {operation}")
                        else:
                            # 优先级2: 从MIR结果文件读取
                            for col in row.index:
                                col_upper = str(col).strip().upper()
                                if col_upper in ['OPERATION', 'OP', 'OPN']:
                                    if pd.notna(row[col]) and str(row[col]).strip():
                                        operation = str(row[col]).strip()
                                        LOGGER.info(f"从MIR结果文件读取 Operation: {operation}")
                                    break
                            
                            # 优先级3: 使用UI配置或配置文件的值
                            if not operation:
                                if hasattr(self, 'ui_config') and self.ui_config:
                                    operation = self.ui_config.get('operation', '').strip() or None
                                    if operation:
                                        LOGGER.info(f"从UI配置读取 Operation: {operation}")
                                else:
                                    # 从配置文件读取
                                    try:
                                        config_path = Path(__file__).parent / "config.yaml"
                                        if config_path.exists():
                                            import yaml
                                            with open(config_path, 'r', encoding='utf-8') as f:
                                                config_data = yaml.safe_load(f) or {}
                                                mole_history = config_data.get('mole_history', {})
                                                operation = mole_history.get('operation', '').strip() or None
                                                if operation:
                                                    LOGGER.info(f"从配置文件读取 Operation: {operation}")
                                    except Exception:
                                        pass
                        
                        # 查找Eng ID列（优先级：For Spark.csv > MIR结果文件 > UI配置）
                        eng_id = None
                        
                        # 优先级1: 从 For Spark.csv 读取
                        if source_lot in spark_config_map and 'eng_id' in spark_config_map[source_lot]:
                            eng_id = spark_config_map[source_lot]['eng_id']
                            LOGGER.info(f"从 For Spark.csv 读取 EngID: {eng_id}")
                        else:
                            # 优先级2: 从MIR结果文件读取
                            for col in row.index:
                                col_upper = str(col).strip().upper()
                                if col_upper in ['ENG ID', 'ENGID', 'ENG_ID', 'ENGINEERING ID', 'ENGINEERING_ID']:
                                    if pd.notna(row[col]) and str(row[col]).strip():
                                        eng_id = str(row[col]).strip()
                                        LOGGER.info(f"从MIR结果文件读取 EngID: {eng_id}")
                                    break
                            
                            # 优先级3: 使用UI配置或配置文件的值
                            if not eng_id:
                                if hasattr(self, 'ui_config') and self.ui_config:
                                    eng_id = self.ui_config.get('engid', '').strip() or None
                                    if eng_id:
                                        LOGGER.info(f"从UI配置读取 EngID: {eng_id}")
                                else:
                                    # 从配置文件读取
                                    try:
                                        config_path = Path(__file__).parent / "config.yaml"
                                        if config_path.exists():
                                            import yaml
                                            with open(config_path, 'r', encoding='utf-8') as f:
                                                config_data = yaml.safe_load(f) or {}
                                                mole_history = config_data.get('mole_history', {})
                                                eng_id = mole_history.get('engid', '').strip() or None
                                                if eng_id:
                                                    LOGGER.info(f"从配置文件读取 EngID: {eng_id}")
                                    except Exception:
                                        pass
                        
                        # 处理More options字段（优先级：For Spark.csv > MIR结果文件 > UI配置）
                        unit_test_time = None
                        retest_rate = None
                        hri_mrv = None
                        
                        # 优先级1: 从 For Spark.csv 读取
                        if source_lot in spark_config_map:
                            spark_config = spark_config_map[source_lot]
                            if 'unit_test_time' in spark_config:
                                unit_test_time = spark_config['unit_test_time']
                                LOGGER.info(f"从 For Spark.csv 读取 Unit test time: {unit_test_time}")
                            if 'retest_rate' in spark_config:
                                retest_rate = spark_config['retest_rate']
                                LOGGER.info(f"从 For Spark.csv 读取 Retest rate: {retest_rate}")
                            if 'hri_mrv' in spark_config:
                                hri_mrv = spark_config['hri_mrv']
                                LOGGER.info(f"从 For Spark.csv 读取 HRI / MRV: {hri_mrv}")
                        
                        # 优先级2: 从MIR结果文件读取（如果For Spark.csv中没有）
                        if not unit_test_time:
                            unit_test_time = row.get('Unit test time', None)
                            if pd.notna(unit_test_time) and str(unit_test_time).strip():
                                unit_test_time = str(unit_test_time).strip()
                                LOGGER.info(f"从MIR结果文件读取 Unit test time: {unit_test_time}")
                        
                        if not retest_rate:
                            retest_rate = row.get('Retest rate', None)
                            if pd.notna(retest_rate) and str(retest_rate).strip():
                                retest_rate = str(retest_rate).strip()
                                LOGGER.info(f"从MIR结果文件读取 Retest rate: {retest_rate}")
                        
                        if not hri_mrv:
                            hri_mrv = row.get('HRI / MRV:', None)
                            if pd.notna(hri_mrv) and str(hri_mrv).strip():
                                hri_mrv = str(hri_mrv).strip()
                                LOGGER.info(f"从MIR结果文件读取 HRI / MRV: {hri_mrv}")
                        
                        # 优先级3: 使用UI配置或配置文件的值（如果前面都没有）
                        if not unit_test_time:
                            if hasattr(self, 'ui_config') and self.ui_config:
                                ui_value = self.ui_config.get('unit_test_time', '').strip()
                                if ui_value:
                                    unit_test_time = ui_value
                                    LOGGER.info(f"从UI配置读取 Unit test time: {unit_test_time}")
                            else:
                                # 从配置文件读取
                                try:
                                    config_path = Path(__file__).parent / "config.yaml"
                                    if config_path.exists():
                                        import yaml
                                        with open(config_path, 'r', encoding='utf-8') as f:
                                            config_data = yaml.safe_load(f) or {}
                                            mole_history = config_data.get('mole_history', {})
                                            ui_value = mole_history.get('unit_test_time', '').strip()
                                            if ui_value:
                                                unit_test_time = ui_value
                                                LOGGER.info(f"从配置文件读取 Unit test time: {unit_test_time}")
                                except Exception:
                                    pass
                        
                        if not retest_rate:
                            if hasattr(self, 'ui_config') and self.ui_config:
                                ui_value = self.ui_config.get('retest_rate', '').strip()
                                if ui_value:
                                    retest_rate = ui_value
                                    LOGGER.info(f"从UI配置读取 Retest rate: {retest_rate}")
                            else:
                                # 从配置文件读取
                                try:
                                    config_path = Path(__file__).parent / "config.yaml"
                                    if config_path.exists():
                                        import yaml
                                        with open(config_path, 'r', encoding='utf-8') as f:
                                            config_data = yaml.safe_load(f) or {}
                                            mole_history = config_data.get('mole_history', {})
                                            ui_value = mole_history.get('retest_rate', '').strip()
                                            if ui_value:
                                                retest_rate = ui_value
                                                LOGGER.info(f"从配置文件读取 Retest rate: {retest_rate}")
                                except Exception:
                                    pass
                        
                        if not hri_mrv:
                            if hasattr(self, 'ui_config') and self.ui_config:
                                ui_value = self.ui_config.get('hri_mrv', '').strip()
                                if ui_value:
                                    hri_mrv = ui_value
                                    LOGGER.info(f"从UI配置读取 HRI / MRV: {hri_mrv}")
                            else:
                                # 从配置文件读取
                                try:
                                    config_path = Path(__file__).parent / "config.yaml"
                                    if config_path.exists():
                                        import yaml
                                        with open(config_path, 'r', encoding='utf-8') as f:
                                            config_data = yaml.safe_load(f) or {}
                                            mole_history = config_data.get('mole_history', {})
                                            ui_value = mole_history.get('hri_mrv', '').strip()
                                            if ui_value:
                                                hri_mrv = ui_value
                                                LOGGER.info(f"从配置文件读取 HRI / MRV: {hri_mrv}")
                                except Exception:
                                    pass
                        
                        # 执行Spark提交流程
                        # 注意：第一行需要点击Add New，后续行在上一行Roll后已经点击了Add New
                        if row_num == 0:
                            LOGGER.info("步骤 1/13: 点击Add New...")
                            if not self.spark_submitter._click_add_new_button():
                                raise WorkflowError("点击Add New按钮失败")
                        else:
                            LOGGER.info("步骤 1/13: 已点击Add New（上一行Roll后已点击）")
                        
                        LOGGER.info("步骤 2/13: 填写TP路径...")
                        if not self.spark_submitter._fill_test_program_path(self.config.paths.tp_path):
                            raise WorkflowError("填写TP路径失败")
                        
                        LOGGER.info("步骤 3/13: 点击Add New Experiment...")
                        if not self.spark_submitter._click_add_new_experiment():
                            raise WorkflowError("点击Add New Experiment失败")
                        
                        LOGGER.info("步骤 4/13: 选择VPO类别...")
                        if not self.spark_submitter._select_vpo_category(self.config.spark.vpo_category):
                            raise WorkflowError("选择VPO类别失败")
                        
                        LOGGER.info("步骤 5/13: 填写实验信息...")
                        if not self.spark_submitter._fill_experiment_info(self.config.spark.step, self.config.spark.tags):
                            raise WorkflowError("填写实验信息失败")
                        
                        LOGGER.info("步骤 6/13: 添加Lot name...")
                        # 查找Quantity列（用于设置units数量）
                        quantity = None
                        for col in row.index:
                            col_upper = str(col).strip().upper()
                            if col_upper in ['QUANTITY', 'QTY', 'UNITS', 'UNIT COUNT', 'COUNT']:
                                if pd.notna(row[col]) and str(row[col]).strip():
                                    try:
                                        quantity = int(float(str(row[col]).strip()))
                                        LOGGER.info(f"从数据中读取Quantity: {quantity}")
                                    except (ValueError, TypeError):
                                        LOGGER.warning(f"Quantity值无效: {row[col]}，跳过设置units数量")
                                break
                        
                        if not self.spark_submitter._add_lot_name(source_lot, quantity):
                            raise WorkflowError("添加Lot name失败")
                        
                        LOGGER.info("步骤 7/13: 选择Part Type...")
                        if not self.spark_submitter._select_parttype(part_type):
                            raise WorkflowError("选择Part Type失败")
                        
                        LOGGER.info("步骤 8/13: 点击Flow标签...")
                        if not self.spark_submitter._click_flow_tab():
                            raise WorkflowError("点击Flow标签失败")
                        
                        # Operation是可选的，但如果存在则必须成功选择
                        if operation:
                            LOGGER.info("步骤 9/13: 选择Operation...")
                            if not self.spark_submitter._select_operation(operation):
                                raise WorkflowError("选择Operation失败")
                        else:
                            LOGGER.info("步骤 9/13: 跳过Operation（文件中未提供）")
                        
                        # Eng ID是可选的，但如果存在则必须成功选择
                        if eng_id:
                            LOGGER.info("步骤 10/13: 选择Eng ID...")
                            if not self.spark_submitter._select_eng_id(eng_id):
                                raise WorkflowError("选择Eng ID失败")
                        else:
                            LOGGER.info("步骤 10/13: 跳过Eng ID（文件中未提供）")
                        
                        LOGGER.info("步骤 11/13: 点击More options标签...")
                        if not self.spark_submitter._click_more_options_tab():
                            raise WorkflowError("点击More options标签失败")
                        
                        LOGGER.info("步骤 12/13: 填写More options字段...")
                        if not self.spark_submitter._fill_more_options(unit_test_time, retest_rate, hri_mrv):
                            raise WorkflowError("填写More options字段失败")
                        
                        LOGGER.info("步骤 13/13: 点击Roll按钮...")
                        if not self.spark_submitter._click_roll_button():
                            raise WorkflowError("点击Roll按钮失败")
                        
                        # 等待Roll提交完成
                        LOGGER.info("等待Roll提交完成...")
                        time.sleep(3.0)  # 等待提交响应
                        
                        LOGGER.info(f"✅ 第 {row_num + 1} 行数据提交成功")
                        self.results.append({
                            'row_index': idx,
                            'step': 'Spark',
                            'status': 'success',
                            'source_lot': source_lot,
                            'timestamp': datetime.now().isoformat()
                        })
                        
                        # 如果不是最后一行，点击Add New按钮开始下一行
                        if row_num < len(mir_df) - 1:
                            LOGGER.info("=" * 80)
                            LOGGER.info(f"准备处理下一行（第 {row_num + 2}/{len(mir_df)} 行）...")
                            LOGGER.info("点击Add New按钮开始新的提交...")
                            LOGGER.info("=" * 80)
                            
                            # 等待页面稳定
                            time.sleep(2.0)
                            
                            # 点击Add New按钮
                            if not self.spark_submitter._click_add_new_button():
                                raise WorkflowError("点击Add New按钮失败，无法继续处理下一行")
                            
                            # 等待Add New对话框或页面响应
                            time.sleep(2.0)
                            
                            LOGGER.info("✅ 已点击Add New按钮，准备处理下一行")
                        else:
                            LOGGER.info("=" * 80)
                            LOGGER.info("这是最后一行，不需要点击Add New按钮")
                            LOGGER.info("=" * 80)
                        
                    except Exception as e:
                        error_msg = f"第 {row_num + 1} 行数据提交失败: {e}"
                        LOGGER.error(f"❌ {error_msg}")
                        LOGGER.error(traceback.format_exc())
                        
                        source_lot_value = source_lot if 'source_lot' in locals() and source_lot else 'N/A'
                        self.errors.append({
                            'row_index': idx,
                            'step': 'Spark',
                            'error': str(e),
                            'source_lot': source_lot_value,
                            'timestamp': datetime.now().isoformat()
                        })
                        
                        # 继续处理下一行，不中断整个流程
                        LOGGER.warning(f"⚠️ 第 {row_num + 1} 行提交失败，但将继续处理下一行...")
                        
                        # 如果不是最后一行，尝试点击Add New按钮，为下一行做准备
                        if row_num < len(mir_df) - 1:
                            try:
                                LOGGER.info("尝试点击Add New按钮，为下一行做准备...")
                                time.sleep(2.0)  # 等待页面稳定
                                if self.spark_submitter._click_add_new_button():
                                    LOGGER.info("✅ 已点击Add New按钮，可以继续处理下一行")
                                    time.sleep(2.0)
                                else:
                                    LOGGER.warning("⚠️ 点击Add New按钮失败，下一行可能无法正常处理")
                            except Exception as e2:
                                LOGGER.warning(f"⚠️ 尝试点击Add New按钮时出错: {e2}，但将继续处理下一行")
                
                LOGGER.info("=" * 80)
                LOGGER.info("✅ 所有VPO数据提交请求已完成")
                LOGGER.info(f"   总行数: {len(mir_df)}")
                LOGGER.info(f"   成功: {len([r for r in self.results if r.get('step') == 'Spark' and r.get('status') == 'success'])}")
                LOGGER.info(f"   失败: {len([e for e in self.errors if e.get('step') == 'Spark'])}")
                LOGGER.info("=" * 80)

            # ------------------------------------------------------------------
            # 所有提交完成后，等待一段时间，从Dashboard收集VPO并写回CSV
            # ------------------------------------------------------------------
            try:
                expected_vpo_count = len(mir_df)
                LOGGER.info("开始从Spark Dashboard收集VPO编号，用于回写到MIR结果CSV...")
                vpo_list = self.spark_submitter.collect_recent_vpos_from_dashboard(expected_count=expected_vpo_count)

                if not vpo_list:
                    LOGGER.warning("未能从Dashboard收集到任何VPO编号，跳过生成包含VPO的新CSV。")
                    return

                # 页面上顺序：最新在前；MIR CSV顺序：最早在前
                # 需要将列表反向后按顺序对应到每一行
                LOGGER.info("开始将收集到的VPO编号与MIR结果按顺序匹配...")
                vpo_list_reversed = list(reversed(vpo_list))

                mir_with_vpo = mir_df.copy()
                vpo_col_name = "VPO"
                if vpo_col_name in mir_with_vpo.columns:
                    LOGGER.warning(f"检测到MIR结果中已存在列 '{vpo_col_name}'，将覆盖该列的值。")

                mir_with_vpo[vpo_col_name] = ""

                max_count = min(len(mir_with_vpo), len(vpo_list_reversed))
                for i in range(max_count):
                    mir_with_vpo.at[mir_with_vpo.index[i], vpo_col_name] = vpo_list_reversed[i]
                    LOGGER.info(f"第 {i+1} 行: SourceLot={mir_with_vpo.iloc[i].get('SourceLot', 'N/A')} , VPO={vpo_list_reversed[i]}")

                # 生成新的带VPO的Excel文件（保存到MIR结果目录）
                date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = self.work_subdirs['mir'] / f"MIR_Results_with_VPO_{date_str}.xlsx"
                try:
                    mir_with_vpo.to_excel(output_file, index=False, engine='openpyxl')
                    LOGGER.info(f"✅ 已生成包含VPO的新Excel文件: {output_file}")
                except Exception as e:
                    # 如果Excel保存失败，尝试保存为CSV
                    LOGGER.warning(f"保存Excel文件失败: {e}，尝试保存为CSV格式...")
                    csv_file = self.work_subdirs['mir'] / f"MIR_Results_with_VPO_{date_str}.csv"
                    mir_with_vpo.to_csv(csv_file, index=False, encoding="utf-8-sig")
                    output_file = csv_file
                    LOGGER.info(f"✅ 已生成包含VPO的新CSV文件: {output_file}")

                LOGGER.info(f"✅ 已生成包含VPO的新CSV文件: {output_file}")
                LOGGER.info(f"   共写入 {max_count} 条VPO记录（总行数: {len(mir_with_vpo)}）")
            except Exception as e:
                LOGGER.error(f"从Dashboard收集VPO并写回CSV时出错: {e}")
                LOGGER.error(traceback.format_exc())
                
        except Exception as e:
            raise WorkflowError(f"提交VPO数据到Spark网页失败: {e}")
    
    def _step_generate_gts_file(self) -> None:
        """步骤4: 生成GTS填充文件"""
        try:
            LOGGER.info("正在生成GTS填充文件...")
            
            # 导入GTS填充模块
            from .gts_excel_filler import fill_gts_template_from_csv
            
            # 查找最新的 MIR_Results_with_VPO_* 文件（支持Excel和CSV格式）
            # 优先在工作目录中查找，如果没有则在output目录中查找（向后兼容）
            mir_results_dir = self.work_subdirs.get('mir', self.work_dir / '01_MIR')
            vpo_excel_files = sorted(mir_results_dir.glob("MIR_Results_with_VPO_*.xlsx"), reverse=True)
            vpo_csv_files = sorted(mir_results_dir.glob("MIR_Results_with_VPO_*.csv"), reverse=True)
            
            # 如果工作目录中没有，则在output目录中查找
            if not vpo_excel_files and not vpo_csv_files:
                output_dir = self.config.paths.output_dir
                if output_dir.exists():
                    vpo_excel_files = sorted(output_dir.glob("MIR_Results_with_VPO_*.xlsx"), reverse=True)
                    vpo_csv_files = sorted(output_dir.glob("MIR_Results_with_VPO_*.csv"), reverse=True)
            
            if vpo_excel_files:
                vpo_files = vpo_excel_files
            elif vpo_csv_files:
                vpo_files = vpo_csv_files
            else:
                raise WorkflowError("未找到 MIR_Results_with_VPO_* 文件，请先完成 Spark 步骤\n支持格式: .xlsx, .csv")
            
            input_file = vpo_files[0]
            LOGGER.info(f"使用输入文件: {input_file.name}")
            
            # 查找GTS模板文件
            base_dir = Path(__file__).parent.parent
            template_path = base_dir / "input" / "GTS_Submit.xlsx"
            if not template_path.exists():
                # 尝试其他可能的位置
                possible_paths = [
                    base_dir / "input" / "GTS_Submit.xlsx",
                    self.config.paths.input_dir / "GTS_Submit.xlsx",
                ]
                for path in possible_paths:
                    if path.exists():
                        template_path = path
                        break
                else:
                    raise WorkflowError(f"未找到GTS模板文件。已检查: {possible_paths}")
            
            # 调用填充函数，输出到GTS文件目录
            output_file = fill_gts_template_from_csv(
                input_file,
                template_path,
                self.work_subdirs['gts']
            )
            
            if output_file and output_file.exists():
                LOGGER.info(f"✅ GTS填充文件已生成: {output_file.name}")
            else:
                raise WorkflowError("生成GTS填充文件失败")
                
        except Exception as e:
            raise WorkflowError(f"生成GTS填充文件失败: {e}")
    
    def _step_submit_to_gts(self) -> None:
        """步骤5: 自动填充并提交GTS"""
        try:
            LOGGER.info("正在打开GTS页面并自动填充...")
            
            # 更新GTS submitter的输出目录为工作目录的GTS文件目录
            self.gts_submitter.config.output_dir = self.work_subdirs['gts']
            
            # 调用新的自动填充逻辑
            self.gts_submitter.fill_ticket_with_latest_output()
            
            LOGGER.info("✅ GTS 填充和提交流程已完成")
            
        except Exception as e:
            LOGGER.error(f"GTS自动填充失败: {e}")
            LOGGER.error(traceback.format_exc())
            raise WorkflowError(f"提交GTS失败: {e}")
    
    def _merge_mir_with_spark_config(self, mir_df: pd.DataFrame, spark_config_file: Path | None) -> pd.DataFrame:
        """
        合并 MIR 结果文件和 For Spark.csv，生成汇总文件
        
        Args:
            mir_df: MIR结果DataFrame
            spark_config_file: For Spark.csv 文件路径（可选）
        
        Returns:
            合并后的DataFrame
        """
        merged_df = mir_df.copy()
        
        # 首先处理 Units_Count_Expected 到 Quantity 的映射（无论是否有 For Spark.csv）
        if 'Units_Count_Expected' in merged_df.columns:
            LOGGER.info("发现 Units_Count_Expected 列，将其映射到 Quantity 列...")
            # 如果 Quantity 列不存在，创建它
            if 'Quantity' not in merged_df.columns:
                merged_df['Quantity'] = None
            
            # 将 Units_Count_Expected 的值复制到 Quantity（如果 Quantity 为空或不存在）
            mask_quantity_empty = merged_df['Quantity'].isna() | (merged_df['Quantity'] == '')
            merged_df.loc[mask_quantity_empty, 'Quantity'] = merged_df.loc[mask_quantity_empty, 'Units_Count_Expected']
            LOGGER.info(f"已将 {mask_quantity_empty.sum()} 行的 Units_Count_Expected 映射到 Quantity")
        elif 'Units_Count_Actual' in merged_df.columns:
            # 如果没有 Units_Count_Expected，尝试使用 Units_Count_Actual
            LOGGER.info("发现 Units_Count_Actual 列，将其映射到 Quantity 列...")
            if 'Quantity' not in merged_df.columns:
                merged_df['Quantity'] = None
            mask_quantity_empty = merged_df['Quantity'].isna() | (merged_df['Quantity'] == '')
            merged_df.loc[mask_quantity_empty, 'Quantity'] = merged_df.loc[mask_quantity_empty, 'Units_Count_Actual']
            LOGGER.info(f"已将 {mask_quantity_empty.sum()} 行的 Units_Count_Actual 映射到 Quantity")
        
        if not spark_config_file or not spark_config_file.exists():
            LOGGER.info("未找到 For Spark.csv 文件，跳过合并（但已处理 Units_Count_Expected 映射）")
            return merged_df
        
        try:
            LOGGER.info(f"读取 For Spark.csv 文件: {spark_config_file}")
            spark_config_df = read_excel_file(spark_config_file)
            
            if spark_config_df.empty:
                LOGGER.warning("For Spark.csv 文件为空，跳过合并")
                return merged_df
            
            # 查找 SourceLot 列（在 MIR 结果中）
            mir_source_lot_col = None
            for col in merged_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT']:
                    mir_source_lot_col = col
                    break
            
            if not mir_source_lot_col:
                LOGGER.warning("MIR结果文件中未找到 SourceLot 列，跳过合并")
                return merged_df
            
            # 查找 SourceLot 列（在 For Spark.csv 中）
            spark_source_lot_col = None
            for col in spark_config_df.columns:
                col_upper = str(col).strip().upper()
                if col_upper in ['SOURCELOT', 'SOURCE LOT', 'SOURCE_LOT']:
                    spark_source_lot_col = col
                    break
            
            if not spark_source_lot_col:
                LOGGER.warning("For Spark.csv 中未找到 SourceLot 列，跳过合并")
                return merged_df
            
            # 标准化列名（统一为 Source Lot）
            if mir_source_lot_col != 'Source Lot':
                merged_df = merged_df.rename(columns={mir_source_lot_col: 'Source Lot'})
                mir_source_lot_col = 'Source Lot'
            
            if spark_source_lot_col != 'Source Lot':
                spark_config_df = spark_config_df.rename(columns={spark_source_lot_col: 'Source Lot'})
                spark_source_lot_col = 'Source Lot'
            
            # 标准化 Source Lot 值（去除空格，转为字符串）
            merged_df['Source Lot'] = merged_df['Source Lot'].astype(str).str.strip()
            spark_config_df['Source Lot'] = spark_config_df['Source Lot'].astype(str).str.strip()
            
            # 建立 For Spark.csv 的映射（以 Source Lot 为键）
            spark_config_dict = {}
            for _, row in spark_config_df.iterrows():
                source_lot = str(row['Source Lot']).strip()
                if source_lot and source_lot != 'nan':
                    spark_config_dict[source_lot] = row.to_dict()
            
            LOGGER.info(f"从 For Spark.csv 加载了 {len(spark_config_dict)} 个 Source Lot 的配置")
            
            # 合并数据：将 For Spark.csv 中的列添加到 MIR 结果中
            # 如果 MIR 结果中已有该列，则优先使用 For Spark.csv 的值（如果非空）
            # 特别处理：Part Type 字段始终以 For Spark.csv 为准（即使 MIR 结果中有值）
            part_type_cols = ['Part Type', 'PartType', 'PART_TYPE', 'Part_Type']  # 可能的 Part Type 列名
            
            for source_lot, config_row in spark_config_dict.items():
                # 找到 MIR 结果中匹配的行
                mask = merged_df['Source Lot'] == source_lot
                if mask.any():
                    # 对于每个配置列，如果 MIR 结果中没有或为空，则使用 For Spark.csv 的值
                    for col, value in config_row.items():
                        if col == 'Source Lot':
                            continue
                        if pd.notna(value) and str(value).strip():
                            # 如果列不存在，先创建
                            if col not in merged_df.columns:
                                merged_df[col] = None
                            
                            # 检查是否是 Part Type 相关的列
                            col_upper = str(col).strip().upper()
                            is_part_type = any(pt_col.upper() == col_upper for pt_col in part_type_cols)
                            
                            if is_part_type:
                                # Part Type 字段：始终以 For Spark.csv 为准，直接覆盖
                                merged_df.loc[mask, col] = value
                                LOGGER.debug(f"Source Lot '{source_lot}': Part Type 以 For Spark.csv 为准，值='{value}'")
                            else:
                                # 其他字段：如果 MIR 结果中该列为空，则填充；否则直接覆盖（For Spark.csv 优先）
                                merged_df.loc[mask, col] = merged_df.loc[mask, col].fillna(value)
                                merged_df.loc[mask, col] = value
            
            # 注意：Units_Count_Expected 到 Quantity 的映射已经在函数开头处理了
            
            LOGGER.info(f"✅ 合并完成：MIR结果 ({len(merged_df)} 行) + For Spark.csv ({len(spark_config_dict)} 个配置)")
            LOGGER.info(f"   合并后的列: {merged_df.columns.tolist()}")
            
        except Exception as e:
            LOGGER.warning(f"合并 For Spark.csv 时出错: {e}，将使用原始 MIR 结果")
            import traceback
            LOGGER.debug(traceback.format_exc())
        
        return merged_df
    
    def _step_save_results(self, df: pd.DataFrame) -> Path:
        """保存处理结果"""
        try:
            # 添加处理结果信息到数据框
            result_df = df.copy()
            
            # 添加处理状态列
            if self.results:
                results_df = pd.DataFrame(self.results)
                # 可以根据需要合并结果信息到原始数据框
                # 这里简化处理，直接保存原始数据和处理结果
            
            # 添加错误信息列
            if self.errors:
                errors_df = pd.DataFrame(self.errors)
                # 可以将错误信息合并到结果中
            
            # 生成日期字符串
            date_str = datetime.now().strftime("%Y%m%d")
            
            # 保存结果到工作目录
            output_path = save_result_excel(
                result_df,
                self.work_dir,
                date_str
            )
            
            # 如果存在错误，也保存错误日志到工作目录
            if self.errors:
                error_log_path = self.work_dir / f"workflow_errors_{date_str}.csv"
                errors_df = pd.DataFrame(self.errors)
                errors_df.to_csv(error_log_path, index=False, encoding='utf-8-sig')
                LOGGER.warning(f"存在 {len(self.errors)} 个错误，已保存到: {error_log_path}")
            
            LOGGER.info(f"✅ 结果已保存到: {output_path}")
            return output_path
            
        except Exception as e:
            raise WorkflowError(f"保存结果失败: {e}")

