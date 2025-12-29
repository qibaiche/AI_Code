"""GTS 自动填充模块 - 全新重写版本

核心功能：
1. 打开 GTS 新建 ticket 页面
2. 等待用户手动登录（约30秒内页面就绪）
3. 自动检测 Title 和 Description 输入框出现
4. 清空并填充 Title（从 config.yaml 读取）
5. 清空并填充 Description（从最新 Excel 转 HTML 表格，在 iframe 里）
6. 不自动提交，等待用户手动检查
"""

import logging
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Color
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

try:
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    WEBDRIVER_AVAILABLE = True
except ImportError:
    WEBDRIVER_AVAILABLE = False

from .utils.screenshot_helper import log_error_with_screenshot, capture_debug_screenshot

LOGGER = logging.getLogger(__name__)


# ============================================================================
# 配置类
# ============================================================================

@dataclass
class GTSConfig:
    """GTS 配置"""
    url: str
    title_text: str
    title_selector: str = "input.ui-inputtext[pinputtext][type='text']"
    description_iframe_selector: str = "iframe.fr-iframe"
    description_body_selector: str = "body.fr-view[contenteditable='true']"
    output_dir: Optional[Path] = None
    timeout: int = 60
    retry_count: int = 3
    retry_delay: int = 2
    headless: bool = False
    implicit_wait: int = 5
    explicit_wait: int = 20


# ============================================================================
# 辅助函数
# ============================================================================

def find_latest_excel(output_dir: Path) -> Path:
    """查找最新的 GTS_Submit_filled_*.xlsx 文件
    
    搜索策略：
    1. 首先在给定的 output_dir 中查找（通常是 03_GTS 目录）
    2. 如果未找到，查找基础 output 目录
    3. 在 03_GTS/ 子目录中搜索（新结构）
    4. 在所有 run_*/03_GTS/ 子目录中搜索（向后兼容旧结构）
    5. 在所有 run_*/02_GTS_Files/ 子目录中搜索（向后兼容）
    6. 返回按修改时间排序的最新文件
    """
    files = []
    
    # 1. 首先在给定目录中查找
    files.extend(output_dir.glob("GTS_Submit_filled_*.xlsx"))
    
    # 2. 如果未找到，查找基础 output 目录并搜索
    if not files:
        # 找到基础 output 目录
        current = output_dir
        base_output = None
        
        # 向上查找，直到找到 output 目录（最多查找3层）
        for _ in range(3):
            # 如果当前目录是 03_GTS 或 02_GTS_Files，则父级是 output
            if current.name in ["03_GTS", "GTS", "02_GTS_Files"]:
                parent = current.parent
                if parent.exists():
                    base_output = parent
                    break
            
            # 检查当前目录是否包含 03_GTS 或 01_MIR 子目录（说明是 output 目录）
            try:
                if any(d.is_dir() and d.name in ["01_MIR", "02_SPARK", "03_GTS"] for d in current.iterdir()):
                    base_output = current
                    break
            except (OSError, PermissionError):
                pass
            
            # 向上移动一层
            parent = current.parent
            if not parent or parent == current:  # 到达根目录
                break
            current = parent
        
        # 如果找到了基础 output 目录，进行搜索
        if base_output and base_output.exists():
            try:
                # 3. 在 03_GTS/ 子目录中搜索（新结构）
                gts_dir = base_output / "03_GTS"
                if gts_dir.exists() and gts_dir.is_dir() and gts_dir != output_dir:
                    found_files = list(gts_dir.glob("GTS_Submit_filled_*.xlsx"))
                    files.extend(found_files)
                
                # 4. 向后兼容：在所有 run_*/03_GTS/ 子目录中搜索（旧结构）
                for run_dir in base_output.glob("run_*/03_GTS"):
                    if run_dir.is_dir() and run_dir != output_dir:
                        found_files = list(run_dir.glob("GTS_Submit_filled_*.xlsx"))
                        files.extend(found_files)
                # 向后兼容：也搜索 run_*/GTS/ 目录
                for run_dir in base_output.glob("run_*/GTS"):
                    if run_dir.is_dir() and run_dir != output_dir:
                        found_files = list(run_dir.glob("GTS_Submit_filled_*.xlsx"))
                        files.extend(found_files)
                
                # 5. 向后兼容：在所有 run_*/02_GTS_Files/ 子目录中搜索
                for run_dir in base_output.glob("run_*/02_GTS_Files"):
                    if run_dir.is_dir() and run_dir != output_dir:
                        found_files = list(run_dir.glob("GTS_Submit_filled_*.xlsx"))
                        files.extend(found_files)
            except (OSError, PermissionError):
                pass
    
    if not files:
        raise FileNotFoundError(
            f"未找到文件: 在 {output_dir} 及其父级 output 目录中未找到 GTS_Submit_filled_*.xlsx"
        )
    
    # 按修改时间排序，返回最新的文件
    files.sort(key=lambda f: f.stat().st_mtime)
    return files[-1]


def excel_to_html_table(excel_path: Path) -> str:
    """将 Excel 转换为 HTML 表格（完整保留格式：合并单元格、颜色、字体等）"""
    # 使用 openpyxl 读取 Excel（保留样式）
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 获取合并单元格信息
    merged_cells_dict = {}  # {(row, col): (rowspan, colspan)}
    skip_cells = set()  # 被合并的单元格，需要跳过
    
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        merged_cells_dict[(min_row, min_col)] = (rowspan, colspan)
        
        # 标记被合并的单元格（除了左上角）
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r != min_row or c != min_col:
                    skip_cells.add((r, c))
    
    # 辅助函数：将 openpyxl 的颜色转为 CSS
    def get_color(color_obj):
        if color_obj is None:
            return None
        
        try:
            # 方法1：直接读取 rgb 属性
            if hasattr(color_obj, 'rgb') and color_obj.rgb:
                rgb = color_obj.rgb
                
                # 如果是字符串
                if isinstance(rgb, str):
                    # AARRGGBB (8位) -> RRGGBB (去掉透明度)
                    if len(rgb) == 8:
                        return f"#{rgb[2:]}"
                    # RRGGBB (6位)
                    elif len(rgb) == 6:
                        return f"#{rgb}"
                
                # 如果是对象（如 RGB），尝试获取其字符串表示
                else:
                    # 某些版本的 openpyxl 使用 RGB 类，有 __str__ 方法
                    try:
                        rgb_str = format(rgb, 'x') if hasattr(rgb, '__format__') else str(rgb)
                    except:
                        rgb_str = str(rgb)
                    
                    # 清理字符串（去掉非十六进制字符）
                    rgb_clean = ''.join(c for c in rgb_str if c in '0123456789ABCDEFabcdef')
                    
                    if len(rgb_clean) >= 6:
                        # 取最后6位
                        return f"#{rgb_clean[-6:]}"
            
            # 方法2：尝试 index 属性（主题颜色）
            if hasattr(color_obj, 'index') and color_obj.index:
                # index 是索引，无法直接转换，跳过
                pass
            
            # 方法3：尝试 type 和 value
            if hasattr(color_obj, 'type') and color_obj.type == 'rgb':
                if hasattr(color_obj, 'value') and color_obj.value:
                    val = str(color_obj.value)
                    val_clean = ''.join(c for c in val if c in '0123456789ABCDEFabcdef')
                    if len(val_clean) >= 6:
                        return f"#{val_clean[-6:]}"
        
        except Exception:
            # 如果所有方法都失败，返回 None
            pass
        
        return None
    
    # 表格基础样式
    table_style = "border-collapse: collapse; font-family: Calibri, sans-serif; font-size: 15px;"
    
    # 构建 HTML 表格
    html = f'<table style="{table_style}">'
    html += '<tbody>'
    
    # 遍历所有行
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        html += '<tr>'
        
        # 遍历该行的所有单元格
        for col_idx, cell in enumerate(row, start=1):
            # 检查是否被合并（需要跳过）
            if (row_idx, col_idx) in skip_cells:
                continue
            # 获取单元格的值
            value = cell.value if cell.value is not None else ""
            
            # 检查是否是合并单元格的起始位置
            rowspan, colspan = merged_cells_dict.get((row_idx, col_idx), (1, 1))
            
            # 构建单元格样式
            cell_styles = []
            
            # 背景色：读取 Excel 的真实背景色
            if cell.fill and cell.fill.start_color:
                bg_color = get_color(cell.fill.start_color)
                # FFFFFF00 = 黄色, 00000000 = 黑色/透明(显示为白色)
                if bg_color and bg_color != "#000000":
                    cell_styles.append(f"background-color: {bg_color}")
                else:
                    # 00000000 显示为白色
                    cell_styles.append("background-color: white")
            else:
                cell_styles.append("background-color: white")
            
            # 字体颜色：数据行（第3行起）强制黑色，前两行保留原色
            if row_idx >= 3:
                # 数据行：强制黑色
                cell_styles.append("color: black")
            else:
                # 表头：保留原始颜色
                try:
                    if cell.font and cell.font.color:
                        font_color = get_color(cell.font.color)
                        if font_color and font_color != "#000000":
                            cell_styles.append(f"color: {font_color}")
                        else:
                            cell_styles.append("color: black")
                    else:
                        cell_styles.append("color: black")
                except:
                    cell_styles.append("color: black")
            
            # 字体加粗
            if cell.font and cell.font.bold:
                cell_styles.append("font-weight: bold")
            else:
                cell_styles.append("font-weight: 400")
            
            # 字体斜体
            if cell.font and cell.font.italic:
                cell_styles.append("font-style: italic")
            
            # 对齐方式
            if cell.alignment:
                if cell.alignment.horizontal:
                    h_align = cell.alignment.horizontal
                    if h_align == 'center':
                        cell_styles.append("text-align: center")
                    elif h_align == 'right':
                        cell_styles.append("text-align: right")
                    elif h_align == 'left':
                        cell_styles.append("text-align: left")
                else:
                    cell_styles.append("text-align: center")
                
                if cell.alignment.vertical:
                    v_align = cell.alignment.vertical
                    if v_align == 'center':
                        cell_styles.append("vertical-align: middle")
                    elif v_align == 'top':
                        cell_styles.append("vertical-align: top")
                    elif v_align == 'bottom':
                        cell_styles.append("vertical-align: bottom")
                else:
                    cell_styles.append("vertical-align: middle")
            else:
                cell_styles.append("text-align: center")
                cell_styles.append("vertical-align: middle")
            
            # 边框
            cell_styles.append("border: 1px solid black")
            
            # 内边距
            cell_styles.append("padding: 8px")
            
            # 组合样式
            style_str = "; ".join(cell_styles)
            
            # 构建单元格标签（添加 rowspan 和 colspan）
            td_attrs = [f'style="{style_str}"']
            if rowspan > 1:
                td_attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                td_attrs.append(f'colspan="{colspan}"')
            
            # 输出单元格
            html += f'<td {" ".join(td_attrs)}>{value}</td>'
        
        html += '</tr>'
    
    html += '</tbody></table>'
    
    # 在表格前添加提示信息
    message = '<p style="font-size: 14px; color: black; margin-bottom: 10px;"><strong>Please return units to IC room ASAP after test is done.</strong></p>'
    
    # 将提示信息放在表格前面
    final_html = message + html
    
    return final_html


# ============================================================================
# GTS 提交器类
# ============================================================================

class GTSSubmitter:
    """GTS 自动填充器（全新重写）"""
    
    def __init__(self, config: GTSConfig, debug_dir: Optional[Path] = None):
        self.config = config
        self.driver: Optional[webdriver.Chrome] = None
        self._keep_browser_open = False  # 标志：是否保持浏览器打开（用户取消时）
        self.debug_dir = debug_dir or Path.cwd() / "output" / "05_Debug"
    
    def _log_error_with_screenshot(self, error_message: str, exception: Optional[Exception] = None, prefix: str = "gts_error") -> None:
        """记录错误并自动截图"""
        if self.driver:
            log_error_with_screenshot(self.driver, error_message, self.debug_dir, exception, prefix)
        else:
            LOGGER.error(f"❌ {error_message}")
            if exception:
                import traceback
                LOGGER.error(f"异常详情: {str(exception)}")
                LOGGER.debug(traceback.format_exc())
    
    # ------------------------------------------------------------------------
    # 浏览器管理
    # ------------------------------------------------------------------------
    
    def _init_browser(self):
        """初始化 Chrome 浏览器"""
        if self.driver:
            return
        
        print("正在启动 Chrome 浏览器...")
        options = webdriver.ChromeOptions()
        if self.config.headless:
            options.add_argument("--headless")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--no-sandbox")
        # 添加 detach 选项，让浏览器在 Python 程序退出后保持打开
        options.add_experimental_option("detach", True)
        
        try:
            if WEBDRIVER_AVAILABLE:
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
            else:
                self.driver = webdriver.Chrome(options=options)
            self.driver.implicitly_wait(self.config.implicit_wait)
            print("✅ Chrome 浏览器已启动")
        except Exception as e:
            raise RuntimeError(f"❌ 无法启动 Chrome: {e}")
    
    def _close_browser(self):
        """关闭浏览器（保留供外部调用）"""
        # 如果用户取消了，不关闭浏览器
        if self._keep_browser_open:
            print("💡 浏览器将保持打开（用户取消了自动提交）")
            return
        
        if self.driver:
            try:
                self.driver.quit()
                print("已关闭浏览器")
            except:
                pass
            finally:
                self.driver = None
    
    # ------------------------------------------------------------------------
    # 页面导航与等待
    # ------------------------------------------------------------------------
    
    def _open_gts_page(self):
        """打开 GTS 页面"""
        print("=" * 80)
        print(f"正在打开 GTS 页面...")
        print(f"URL: {self.config.url[:100]}...")
        print("=" * 80)
        
        self.driver.get(self.config.url)
        
        # 等待页面基础加载
        try:
            WebDriverWait(self.driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except:
            pass
        
        print()
        print("🌐 页面已打开")
        print("⏳ 如果需要登录/SSO，请在浏览器中手动完成...")
        print("⏳ 脚本将自动检测页面就绪（通常30秒内完成）...")
        print()
    
    def _wait_for_elements(self, max_wait: int = 180):
        """等待 Title 和 Description 输入框出现"""
        print("🔍 正在等待 Title 和 Description 输入框出现...")
        
        start_time = time.time()
        last_log_time = start_time
        
        while time.time() - start_time < max_wait:
            # 每10秒打印一次状态
            if time.time() - last_log_time > 10:
                elapsed = int(time.time() - start_time)
                print(f"   ⏳ 已等待 {elapsed} 秒...")
                last_log_time = time.time()
            
            try:
                # 检查 Title
                title_exists = len(self.driver.find_elements(By.CSS_SELECTOR, self.config.title_selector)) > 0
                
                # 检查 Description iframe
                iframe_exists = len(self.driver.find_elements(By.CSS_SELECTOR, self.config.description_iframe_selector)) > 0
                
                if title_exists and iframe_exists:
                    print("✅ Title 和 Description 输入框已就绪！")
                    print()
                    return True
            except:
                pass
            
            time.sleep(2)
        
        raise TimeoutError("❌ 超时：未检测到 Title 或 Description 输入框\n请确认已完成登录且页面已加载完成")
    
    # ------------------------------------------------------------------------
    # 填充 Title
    # ------------------------------------------------------------------------
    
    def _fill_title(self, text: str):
        """清空并填充 Title"""
        print("📝 步骤 1/2: 填充 Title")
        print(f"   目标文本: {text}")
        
        try:
            element = self.driver.find_element(By.CSS_SELECTOR, self.config.title_selector)
            
            # 1. 清空（使用多种方法确保清空）
            print("   🗑️  清空原有内容...")
            self.driver.execute_script("""
                arguments[0].value = '';
                arguments[0].setAttribute('value', '');
                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('change', {bubbles: true}));
            """, element)
            element.clear()
            time.sleep(0.5)
            
            # 2. 填充新内容
            print("   ✍️  填充新内容...")
            self.driver.execute_script("""
                arguments[0].value = arguments[1];
                arguments[0].setAttribute('value', arguments[1]);
                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('change', {bubbles: true}));
            """, element, text)
            time.sleep(0.3)
            
            # 3. 验证
            current = self.driver.execute_script("return arguments[0].value;", element)
            if current == text:
                print("   ✅ Title 填充成功")
                print()
            else:
                print(f"   ⚠️  Title 可能未完全填充（当前值: {current[:30]}...）")
                print()
        
        except Exception as e:
            raise RuntimeError(f"❌ 填充 Title 失败: {e}")
    
    # ------------------------------------------------------------------------
    # 填充 Description (在 iframe 里)
    # ------------------------------------------------------------------------
    
    def _fill_description(self, html_content: str):
        """清空并填充 Description（在 iframe 的富文本编辑器里）"""
        print("📝 步骤 2/2: 填充 Description")
        print("   目标: HTML 表格（黄色表头）")
        
        original_window = self.driver.current_window_handle
        
        try:
            # 1. 切换到主文档
            self.driver.switch_to.default_content()
            
            # 2. 找到 iframe
            print("   🔍 查找 Description iframe...")
            iframe = self.driver.find_element(By.CSS_SELECTOR, self.config.description_iframe_selector)
            
            # 3. 切换到 iframe
            print("   🔄 切换到 iframe...")
            self.driver.switch_to.frame(iframe)
            
            # 4. 找到 body.fr-view
            print("   🔍 查找富文本编辑器 (body.fr-view)...")
            body = self.driver.find_element(By.CSS_SELECTOR, self.config.description_body_selector)
            
            # 5. 彻底清空原有内容（多次确保完全清空）
            print("   🗑️  彻底清空原有内容...")
            
            # 第一轮：删除所有子节点
            self.driver.execute_script("""
                while (arguments[0].firstChild) {
                    arguments[0].removeChild(arguments[0].firstChild);
                }
            """, body)
            time.sleep(0.2)
            
            # 第二轮：清空所有文本和HTML
            self.driver.execute_script("""
                arguments[0].innerHTML = '';
                arguments[0].innerText = '';
                arguments[0].textContent = '';
            """, body)
            time.sleep(0.2)
            
            # 第三轮：再次检查并清空（确保像 <p> 这样的标签也被删除）
            self.driver.execute_script("""
                arguments[0].innerHTML = '';
                // 触发清空事件
                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
            """, body)
            time.sleep(0.3)
            
            # 6. 填充 HTML 表格（带内联样式）
            print("   ✍️  填充 HTML 表格（黄色表头 + 黑色数据）...")
            self.driver.execute_script("""
                // 直接设置 innerHTML
                arguments[0].innerHTML = arguments[1];
                
                // 触发输入事件通知编辑器内容已更改
                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('change', {bubbles: true}));
                
                // 确保富文本编辑器识别新内容
                if (arguments[0].focus) {
                    arguments[0].focus();
                }
            """, body, html_content)
            time.sleep(0.8)
            
            # 验证填充结果
            content_length = self.driver.execute_script("return arguments[0].innerHTML.length;", body)
            print(f"   ✅ 已填充 {content_length} 字符的 HTML 内容")
            
            # 切回主文档检查是否有 "Word Paste Detected" 对话框
            self.driver.switch_to.default_content()
            time.sleep(0.5)
            
            try:
                # 查找 "Keep" 按钮（对话框中的按钮）
                keep_buttons = self.driver.find_elements(By.XPATH, "//button[contains(text(), 'Keep')]")
                if keep_buttons:
                    print("   🔘 检测到 'Word Paste Detected' 对话框，自动点击 'Keep'...")
                    keep_buttons[0].click()
                    time.sleep(0.3)
                    print("   ✅ 已自动处理粘贴格式对话框")
            except Exception:
                # 没有对话框或点击失败，继续执行
                pass
            
            print("   ✅ Description 填充成功")
            print()
        
        except Exception as e:
            # 如果出错，确保切回主文档
            try:
                self.driver.switch_to.default_content()
            except:
                pass
            raise RuntimeError(f"❌ 填充 Description 失败: {e}")
    
    # ------------------------------------------------------------------------
    # 自动提交
    # ------------------------------------------------------------------------
    
    def _countdown_and_submit(self):
        """弹出桌面对话框让用户选择是否提交（显示实时倒计时）"""
        import tkinter as tk
        from tkinter import ttk
        
        print()
        print("=" * 80)
        print("⏰ 弹出确认对话框（10秒后自动提交）")
        print("=" * 80)
        
        # 用户选择结果
        user_choice = {'action': None}  # 'submit', 'cancel', 或 None (超时自动提交)
        
        # 创建对话框窗口
        dialog = tk.Tk()
        dialog.title("GTS 提交确认")
        dialog.geometry("500x250")
        dialog.resizable(False, False)
        
        # 置顶并居中
        dialog.attributes('-topmost', True)
        dialog.lift()
        dialog.focus_force()
        
        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f'500x250+{x}+{y}')
        
        # 倒计时变量
        countdown_seconds = [10]
        
        # 标题
        title_label = tk.Label(
            dialog, 
            text="✅ Title 和 Description 已填充完成！",
            font=("Arial", 12, "bold"),
            fg="green"
        )
        title_label.pack(pady=20)
        
        # 提示信息
        info_label = tk.Label(
            dialog,
            text="是否立即提交 GTS ticket？",
            font=("Arial", 10)
        )
        info_label.pack(pady=5)
        
        # 倒计时标签
        countdown_label = tk.Label(
            dialog,
            text=f"⏰ {countdown_seconds[0]} 秒后自动提交",
            font=("Arial", 14, "bold"),
            fg="red"
        )
        countdown_label.pack(pady=15)
        
        # 按钮框架
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        def on_submit():
            """用户点击提交"""
            user_choice['action'] = 'submit'
            dialog.destroy()
        
        def on_cancel():
            """用户点击取消"""
            user_choice['action'] = 'cancel'
            dialog.destroy()
        
        # 提交按钮（绿色）
        submit_btn = tk.Button(
            button_frame,
            text="✓ 立即提交",
            command=on_submit,
            width=15,
            height=2,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2"
        )
        submit_btn.pack(side=tk.LEFT, padx=10)
        
        # 取消按钮（红色）
        cancel_btn = tk.Button(
            button_frame,
            text="✕ 取消",
            command=on_cancel,
            width=15,
            height=2,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2"
        )
        cancel_btn.pack(side=tk.LEFT, padx=10)
        
        # 倒计时更新函数
        def update_countdown():
            if countdown_seconds[0] > 0 and user_choice['action'] is None:
                countdown_seconds[0] -= 1
                countdown_label.config(text=f"⏰ {countdown_seconds[0]} 秒后自动提交")
                dialog.after(1000, update_countdown)
            elif countdown_seconds[0] == 0 and user_choice['action'] is None:
                # 倒计时结束，自动提交
                user_choice['action'] = 'submit'
                dialog.destroy()
        
        # 启动倒计时
        dialog.after(1000, update_countdown)
        
        # 关闭窗口时视为取消
        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        
        # 显示对话框（阻塞）
        dialog.mainloop()
        
        # 根据用户选择执行操作
        if user_choice['action'] == 'submit':
            print("✅ 正在提交...")
            self._click_submit_button()
        else:
            print("❌ 用户取消了自动提交")
            print("请手动在浏览器中点击 Submit 按钮")
            print("💡 浏览器将保持打开，您可以手动操作")
            # 设置标志，防止自动关闭浏览器
            self._keep_browser_open = True
            # 等待用户确认，防止程序立即退出导致浏览器关闭
            print()
            print("=" * 80)
            print("⏸️  程序将等待，浏览器保持打开")
            print("   完成操作后，请关闭此窗口或按 Ctrl+C 退出")
            print("=" * 80)
            print()
            try:
                # 等待用户输入，保持程序运行
                input("按 Enter 键退出程序（浏览器将保持打开）...")
            except (KeyboardInterrupt, EOFError):
                print("\n程序退出，浏览器保持打开")
    
    def _click_submit_button(self):
        """点击页面右下角的 Submit 按钮"""
        try:
            # 切换回主文档
            self.driver.switch_to.default_content()
            
            # 查找Submit按钮（多种可能的选择器）
            print("   🔍 正在查找 Submit 按钮...")
            
            submit_selectors = [
                "//button[contains(text(), 'Submit')]",
                "button[type='submit']",
                "input[type='submit']",
                "button.submit-btn",
                "//input[@value='Submit']",
            ]
            
            submit_button = None
            for selector in submit_selectors:
                try:
                    if selector.startswith("//"):
                        # XPath
                        buttons = self.driver.find_elements(By.XPATH, selector)
                    else:
                        # CSS
                        buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    if buttons:
                        submit_button = buttons[0]
                        print(f"   ✅ 找到 Submit 按钮")
                        break
                except:
                    continue
            
            if not submit_button:
                raise RuntimeError("未找到 Submit 按钮，请手动点击提交")
            
            # 滚动到按钮位置
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_button)
            time.sleep(0.5)
            
            # 点击按钮
            print("   🖱️  正在点击 Submit 按钮...")
            submit_button.click()
            time.sleep(2)
            
            print()
            print("=" * 80)
            print("✅ GTS 已成功提交！")
            print("=" * 80)
            print()
            
        except Exception as e:
            print(f"   ❌ 自动提交失败: {e}")
            print("   请手动在浏览器中点击 Submit 按钮")
    
    # ------------------------------------------------------------------------
    # 主入口
    # ------------------------------------------------------------------------
    
    def fill_ticket_with_latest_output(self):
        """主流程：打开页面并自动填充 Title + Description"""
        
        # 1. 准备数据
        output_dir = self.config.output_dir or Path.cwd() / "output"
        excel_path = find_latest_excel(output_dir)
        html_table = excel_to_html_table(excel_path)
        title_text = self.config.title_text
        
        print()
        print("=" * 80)
        print("🚀 GTS 自动填充（全新版本）")
        print("=" * 80)
        print(f"📄 数据文件: {excel_path.name}")
        print(f"📝 Title: {title_text[:60]}...")
        print("=" * 80)
        print()
        
        # 2. 初始化浏览器
        self._init_browser()
        
        try:
            # 3. 打开页面
            self._open_gts_page()
            
            # 4. 等待元素就绪（自动检测登录完成）
            self._wait_for_elements()
            
            # 5. 填充 Title
            self._fill_title(title_text)
            
            # 6. 填充 Description
            self._fill_description(html_table)
            
            # 7. 完成
            print("=" * 80)
            print("✅ Title 和 Description 填充完成！")
            print("=" * 80)
            print("📋 请在浏览器中检查:")
            print("   - Title 是否正确")
            print("   - Description 表格格式是否正确")
            print("   - 数据内容是否准确")
            print()
            
            # 10秒倒计时，让用户选择是否提交
            self._countdown_and_submit()
            
            print("=" * 80)
            print()
        
        except Exception as e:
            print()
            print("=" * 80)
            print(f"❌ 错误: {e}")
            print("=" * 80)
            print()
            raise
        
        finally:
            # 如果用户取消了，不关闭浏览器
            if self._keep_browser_open:
                print("💡 浏览器保持打开，您可以手动操作或关闭")
            # 否则，浏览器会在 __exit__ 中关闭（如果使用了上下文管理器）
            # 但这里不使用上下文管理器，所以浏览器会保持打开
    
    # ------------------------------------------------------------------------
    # 兼容旧接口
    # ------------------------------------------------------------------------
    
    def submit_final_data(self, data: dict) -> bool:
        """兼容接口（不使用）"""
        return False
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self._close_browser()
