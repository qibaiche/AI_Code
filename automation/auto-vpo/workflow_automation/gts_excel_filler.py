"""填充 GTS_Submit 模板的辅助脚本

从最新的 MIR/VPO CSV 结果中提取关键信息，按照 `input/GTS_Submit.xlsx`
的列顺序生成新的 Excel，输出到 `output` 目录。
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from copy import copy

from .data_reader import read_excel_file


def _get_first_non_null(series: pd.Series, candidates: Iterable[str]) -> str | None:
    """在给定候选列名中按顺序返回首个非空值。"""
    for name in candidates:
        if name in series and pd.notna(series[name]):
            value = series[name]
            # 转为字符串以避免数字被保存成科学计数法
            return str(value).replace("\n", " ").strip()
    return None


def fill_gts_template_from_csv(
    csv_path: Path,
    template_path: Path,
    output_dir: Path,
) -> Path:
    """
    将 CSV 中的 VPO/MIR/SourceLot 信息填充到 GTS_Submit 模板并导出。

    Args:
        csv_path: MIR/VPO 结果 CSV 路径
        template_path: GTS_Submit.xlsx 模板路径
        output_dir: 输出目录

    Returns:
        新生成的 Excel 路径
    """
    csv_path = Path(csv_path)
    template_path = Path(template_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    csv_df = read_excel_file(csv_path)
    if csv_df.empty:
        raise ValueError(f"CSV 文件为空: {csv_path}")

    template_df = pd.read_excel(template_path)
    if template_df.empty:
        raise ValueError(f"模板为空: {template_path}")

    # 取模板中“非空字段最多”的行作为默认值来源，避免选到提示行
    non_empty_rows = template_df.dropna(how="all")
    default_row = non_empty_rows.loc[non_empty_rows.notna().sum(axis=1).idxmax()]
    default_values = {
        col: default_row[col]
        for col in template_df.columns
        if pd.notna(default_row[col])
    }

    result_rows: list[dict] = []
    for _, row in csv_df.iterrows():
        row_data = {col: default_values.get(col) for col in template_df.columns}

        row_data["VPO number"] = _get_first_non_null(
            row, ["VPO", "VPO number", "VPO#"]
        )
        row_data["Location code"] = _get_first_non_null(
            row,
            [
                "Operation",
                "OPERATION",
                "OP",
                "Location code",
                "Location",
                "LOC",
            ],
        )
        row_data["Source Lot/"] = _get_first_non_null(
            row, ["SourceLot", "Source Lot", "Source Lot/", "SOURCELOT", "SOURCE LOT"]
        )
        row_data["MIR"] = _get_first_non_null(row, ["MIR"])

        result_rows.append(row_data)

    if not result_rows:
        raise ValueError("未生成任何行，检查 CSV 内容是否包含需要的列")

    result_df = pd.DataFrame(result_rows, columns=template_df.columns)

    # 使用 openpyxl 保留模板格式（表头、填充颜色、合并单元格等）
    wb = load_workbook(template_path)
    ws = wb.active

    header_row_idx = 1
    merged_max_row = max((rng.max_row for rng in ws.merged_cells.ranges), default=header_row_idx)
    # 数据写入从合并区域之后开始，避免写入合并单元格
    data_start_row = max(header_row_idx + 1, merged_max_row + 1)

    # 找到模板里的“样板数据行”用于复制格式：取 data_start_row 开始第一个非全空行，否则用 data_start_row
    sample_row_idx = None
    for r in range(data_start_row, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, len(template_df.columns) + 1)):
            sample_row_idx = r
            break
    if sample_row_idx is None:
        sample_row_idx = data_start_row

    sample_cells = [
        ws.cell(row=sample_row_idx, column=c)
        for c in range(1, len(template_df.columns) + 1)
    ]

    # 仅清空旧数据行的值，不删除行，避免破坏合并单元格和格式
    for r in range(data_start_row, ws.max_row + 1):
        for c in range(1, len(template_df.columns) + 1):
            ws.cell(row=r, column=c, value=None)

    # 写入新数据，并应用样板行的格式
    for idx, row in result_df.iterrows():
        target_row_idx = data_start_row + idx
        for col_idx, col_name in enumerate(template_df.columns, start=1):
            target_cell = ws.cell(row=target_row_idx, column=col_idx)

            # 如目标单元格属于合并区域，则写入其左上角锚点单元格
            if isinstance(target_cell, MergedCell) or any(
                target_cell.coordinate in rng for rng in ws.merged_cells.ranges
            ):
                for rng in ws.merged_cells.ranges:
                    if target_cell.coordinate in rng:
                        anchor_cell = ws.cell(row=rng.min_row, column=rng.min_col)
                        anchor_cell.value = row[col_name]
                        target_cell = anchor_cell
                        break
            else:
                target_cell.value = row[col_name]

            # 复制样板格式
            sample_cell = sample_cells[col_idx - 1]
            target_cell._style = copy(sample_cell._style)
            target_cell.number_format = sample_cell.number_format

            # 数据行统一取消自动换行，保持单行显示
            align = copy(sample_cell.alignment)
            if align:
                align.wrap_text = False
                target_cell.alignment = align

    # ------------------------------------------------------------
    # 对相邻相同值进行纵向合并，并居中显示
    # ------------------------------------------------------------
    max_row = data_start_row + len(result_df) - 1
    if len(result_df) > 1:
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        for col_idx in range(1, len(template_df.columns) + 1):
            start = data_start_row
            prev = ws.cell(row=start, column=col_idx).value
            for r in range(data_start_row + 1, max_row + 2):  # 末尾多迭代一次以便收尾
                val = ws.cell(row=r, column=col_idx).value if r <= max_row else None
                if val != prev:
                    end = r - 1
                    if prev is not None and end > start:
                        ws.merge_cells(start_row=start, end_row=end, start_column=col_idx, end_column=col_idx)
                        anchor = ws.cell(row=start, column=col_idx)
                        anchor.alignment = center_align
                    start = r
                    prev = val

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"GTS_Submit_filled_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path


def get_latest_csv(output_dir: Path) -> Path:
    """在 output 目录下查找最新的 MIR 结果 CSV。"""
    output_dir = Path(output_dir)
    files = sorted(output_dir.glob("MIR_Results*.csv"))
    if not files:
        raise FileNotFoundError(f"未在 {output_dir} 找到 MIR 结果 CSV")
    return files[-1]


def get_latest_spark_file(output_dir: Path) -> Path:
    """在 output 目录下查找 SPARK 文件夹中最新的 MIR_Results_For_Spark 文件。
    
    支持以下目录结构：
    - output/02_SPARK/ (旧结构)
    - output/run_*/Spark/ (新结构)
    
    Returns:
        最新的文件路径（支持 .xlsx 和 .csv 格式）
    """
    output_dir = Path(output_dir)
    spark_files = []
    
    # 查找新结构：output/run_*/Spark/
    work_dirs = sorted(output_dir.glob("run_*"), reverse=True)
    for work_dir in work_dirs:
        spark_dir = work_dir / "Spark"
        if spark_dir.exists():
            found_files = list(spark_dir.glob("MIR_Results_For_Spark_*.xlsx"))
            found_files.extend(list(spark_dir.glob("MIR_Results_For_Spark_*.csv")))
            if found_files:
                spark_files.extend(found_files)
                break
    
    # 向后兼容：旧结构 output/02_SPARK/
    if not spark_files:
        spark_dir = output_dir / "02_SPARK"
        if spark_dir.exists():
            found_files = list(spark_dir.glob("MIR_Results_For_Spark_*.xlsx"))
            found_files.extend(list(spark_dir.glob("MIR_Results_For_Spark_*.csv")))
            if found_files:
                spark_files.extend(found_files)
    
    if not spark_files:
        raise FileNotFoundError(
            f"未在 {output_dir} 找到 SPARK 文件夹中的 MIR_Results_For_Spark 文件\n"
            f"请检查以下位置：\n"
            f"  - {output_dir / '02_SPARK'}\n"
            f"  - {output_dir / 'run_*/Spark'}"
        )
    
    # 按修改时间排序，返回最新的文件
    return max(spark_files, key=lambda p: p.stat().st_mtime)


if __name__ == "__main__":
    base_dir = Path(__file__).resolve().parent.parent
    output_dir = base_dir / "output"
    
    # 从 SPARK 文件夹读取最新文件
    print("=" * 80)
    print("📝 GTS 填充调试 - 使用 SPARK 文件夹最新文件")
    print("=" * 80)
    print()
    print(f"📁 查找 SPARK 文件夹中的最新文件...")
    
    try:
        latest_spark_file = get_latest_spark_file(output_dir)
        print(f"✅ 找到文件: {latest_spark_file.name}")
        print(f"   路径: {latest_spark_file}")
    except FileNotFoundError as e:
        print(f"❌ 错误: {e}")
        import sys
        sys.exit(1)
    
    # 模板文件
    template = base_dir / "input" / "GTS_Submit.xlsx"
    if not template.exists():
        print(f"❌ 错误: 模板文件不存在: {template}")
        import sys
        sys.exit(1)
    
    # 输出到 03_GTS 文件夹
    gts_output_dir = output_dir / "03_GTS"
    print()
    print(f"📁 输出目录: {gts_output_dir}")
    
    try:
        output_file = fill_gts_template_from_csv(
            latest_spark_file,
            template,
            gts_output_dir,
        )
        print()
        print(f"✅ 已生成: {output_file}")
        print(f"   完整路径: {output_file.absolute()}")
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        import sys
        sys.exit(1)

