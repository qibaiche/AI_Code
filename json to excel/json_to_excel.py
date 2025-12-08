import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_GENERAL
import os
import tkinter as tk
from tkinter import filedialog, messagebox

def format_number(value):
    """
    将数字格式化为普通格式（非科学计数法）
    """
    if isinstance(value, (int, float)):
        # 使用格式化字符串确保不使用科学计数法
        if abs(value) < 1e-10:  # 非常小的数字
            return f"{value:.15f}".rstrip("0").rstrip(".")
        elif abs(value) < 1:  # 小数
            return f"{value:.10f}".rstrip("0").rstrip(".")
        else:  # 大于1的数字
            return f"{value:.8f}".rstrip("0").rstrip(".")
    return value

def select_json_file():
    """
    弹出文件选择对话框选择JSON文件
    """
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    # 设置文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择JSON文件",
        filetypes=[
            ("JSON文件", "*.json"),
            ("DCLEAK文件", "*.dcleak.json"),
            ("所有文件", "*.*")
        ],
        initialdir=os.path.expanduser("~/Desktop")  # 默认打开桌面
    )
    
    root.destroy()
    return file_path

def process_sub_sheet_data(rows, test_type):
    """
    根据测试类型重新组织limit数据
    """
    processed_rows = []
    for row in rows:
        new_row = {
            "Configuration": row["Configuration"],
            "Pin_Group": row["Pin_Group"]
        }
        
        # 根据测试类型决定使用哪个前缀的limit
        if "HIGH" in test_type or "HI" in test_type:
            # High类型：使用Vss前缀，Test_Type只显示VSS
            new_row["LimitHigh"] = row["VssLimitHigh"]
            new_row["LimitLow"] = row["VssLimitLow"]
            new_row["Test_Type"] = "VSS"
        else:
            # Low类型：使用Vcc前缀，Test_Type只显示VCC
            new_row["LimitHigh"] = row["VccLimitHigh"]
            new_row["LimitLow"] = row["VccLimitLow"]
            new_row["Test_Type"] = "VCC"
        
        processed_rows.append(new_row)
    return processed_rows

def json_to_excel(json_file_path, excel_file_path=None):
    """
    将JSON文件转换为Excel文件
    
    Args:
        json_file_path (str): JSON文件路径
        excel_file_path (str): 输出的Excel文件路径，如果为None则自动生成
    """
    
    # 如果没有指定输出路径，则自动生成
    if excel_file_path is None:
        base_name = os.path.splitext(json_file_path)[0]
        excel_file_path = f"{base_name}.xlsx"
    
    # 读取JSON文件
    try:
        with open(json_file_path, "r", encoding="utf-8") as file:
            data = json.load(file)
    except FileNotFoundError:
        messagebox.showerror("错误", f"找不到文件：{json_file_path}")
        return False
    except json.JSONDecodeError as e:
        messagebox.showerror("错误", f"JSON文件格式错误：{e}")
        return False
    
    # 准备数据列表
    rows = []
    
    # 遍历配置集
    for config_set in data.get("ConfigurationSets", []):
        config_name = config_set.get("ConfigurationName", "")
        
        # 遍历设置
        for setting in config_set.get("Settings", []):
            limits = setting.get("LimitsSettings", {})
            leakage_elements = setting.get("LeakageElements", [])
            
            # 为每个LeakageElement创建一行
            for element in leakage_elements:
                # 处理LeakageElement，删除双冒号及其前面的内容
                processed_element = element
                if "::" in element:
                    processed_element = element.split("::", 1)[1]  # 取双冒号后的部分
                
                row = {
                    "Configuration": config_name,
                    "Pin_Group": processed_element,
                    "VssLimitHigh": format_number(limits.get("VssLimitHigh", "")),
                    "VssLimitLow": format_number(limits.get("VssLimitLow", "")),
                    "VccLimitHigh": format_number(limits.get("VccLimitHigh", "")),
                    "VccLimitLow": format_number(limits.get("VccLimitLow", ""))
                }
                rows.append(row)
    
    # 根据Configuration分类数据到四个子表
    print("正在根据Configuration分类数据...")
    
    # 创建四个空的列表来存储分类后的数据
    hot_leakhi_rows = []
    hot_leaklo_rows = []
    cold_leakhi_rows = []
    cold_leaklo_rows = []
    
    # 遍历原始数据，根据Configuration进行分类
    for row in rows:
        config_name = str(row.get("Configuration", "")).upper()
        
        # 根据Configuration进行分类（支持多种命名方式）
        # 检查是否包含HOT或COLD
        is_hot = "HOT" in config_name
        is_cold = "COLD" in config_name
        
        # 检查是否包含高/低标识（支持多种命名）
        is_high = any(keyword in config_name for keyword in ["LEAKHI", "HIGH", "HI"])
        is_low = any(keyword in config_name for keyword in ["LEAKLO", "LOW", "LO"])
        
        # 根据组合进行分类
        if is_hot and is_high:
            # HOT + High -> HOT_LEAKHI表
            hot_leakhi_rows.append(row)
        elif is_hot and is_low:
            # HOT + Low -> HOT_LEAKLO表
            hot_leaklo_rows.append(row)
        elif is_cold and is_high:
            # COLD + High -> COLD_LEAKHI表
            cold_leakhi_rows.append(row)
        elif is_cold and is_low:
            # COLD + Low -> COLD_LEAKLO表
            cold_leaklo_rows.append(row)
    
    # 处理各个子表的数据，根据测试类型重新组织limit结构
    hot_leakhi_processed = process_sub_sheet_data(hot_leakhi_rows, "HOT_LEAKHI")
    hot_leaklo_processed = process_sub_sheet_data(hot_leaklo_rows, "HOT_LEAKLO")
    cold_leakhi_processed = process_sub_sheet_data(cold_leakhi_rows, "COLD_LEAKHI")
    cold_leaklo_processed = process_sub_sheet_data(cold_leaklo_rows, "COLD_LEAKLO")
    
    # 创建新的工作簿，包含多个工作表
    wb_multi = Workbook()
    
    # 删除默认工作表
    wb_multi.remove(wb_multi.active)
    
    # 创建主工作表
    ws_main = wb_multi.create_sheet("Leakage Configuration")
    
    # 设置标题样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入主工作表标题行
    headers = ["Configuration", "Pin_Group", "VssLimitHigh", "VssLimitLow", "VccLimitHigh", "VccLimitLow"]
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入主工作表数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_main.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            if header in ["VssLimitHigh", "VssLimitLow", "VccLimitHigh", "VccLimitLow"]:
                cell.number_format = "@"
    
    # 创建四个子表
    sub_sheets = [
        ("HOT_LEAKHI", hot_leakhi_processed),
        ("HOT_LEAKLO", hot_leaklo_processed),
        ("COLD_LEAKHI", cold_leakhi_processed),
        ("COLD_LEAKLO", cold_leaklo_processed)
    ]
    
    for sheet_name, sheet_rows in sub_sheets:
        if sheet_rows:  # 只有当有数据时才创建工作表
            ws_sub = wb_multi.create_sheet(sheet_name)
            
            # 子表使用新的列结构
            sub_headers = ["Configuration", "Pin_Group", "LimitHigh", "LimitLow", "Test_Type"]
            for col, header in enumerate(sub_headers, 1):
                cell = ws_sub.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入子表数据
            for row_idx, row_data in enumerate(sheet_rows, 2):
                for col_idx, header in enumerate(sub_headers, 1):
                    cell = ws_sub.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                    if header in ["LimitHigh", "LimitLow"]:
                        cell.number_format = "@"
    
    # 自动调整所有工作表的列宽
    for ws in wb_multi.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存多工作表Excel文件
    try:
        # 生成新的文件名
        base_name = os.path.splitext(excel_file_path)[0]
        multi_sheet_file = f"{base_name}_MultiSheet.xlsx"
        
        wb_multi.save(multi_sheet_file)
        
        # 显示结果统计
        result_message = f"""转换成功！

Excel文件已保存为:
{multi_sheet_file}

总共处理了 {len(rows)} 行数据

工作表统计:
- 主表 (Leakage Configuration): {len(rows)} 行
- HOT_LEAKHI: {len(hot_leakhi_processed)} 行 (Test_Type: VSS)
- HOT_LEAKLO: {len(hot_leaklo_processed)} 行 (Test_Type: VCC)
- COLD_LEAKHI: {len(cold_leakhi_processed)} 行 (Test_Type: VSS)
- COLD_LEAKLO: {len(cold_leaklo_processed)} 行 (Test_Type: VCC)

文件包含 {len(wb_multi.worksheets)} 个工作表

注意：
- High类型测试使用Vss前缀的limits，Test_Type显示为VSS
- Low类型测试使用Vcc前缀的limits，Test_Type显示为VCC
- 子表列结构：Configuration, Pin_Group, LimitHigh, LimitLow, Test_Type
"""
        
        messagebox.showinfo("成功", result_message)
        return True
        
    except Exception as e:
        messagebox.showerror("错误", f"保存多工作表Excel文件时出错：{e}")
        return False

def main():
    # 弹出文件选择对话框
    json_file = select_json_file()
    
    if not json_file:
        print("未选择文件，程序退出")
        return
    
    # 检查文件是否存在
    if not os.path.exists(json_file):
        messagebox.showerror("错误", f"文件不存在：{json_file}")
        return
    
    # 转换文件
    success = json_to_excel(json_file)
    
    if success:
        print(f"转换完成：{json_file}")

if __name__ == "__main__":
    main()
