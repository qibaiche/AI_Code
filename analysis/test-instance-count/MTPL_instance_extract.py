#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
最终版MTPL提取工具 - 生成Excel文件包含两个子表

功能：
1. 支持两种测试类型：PrimeDcLeakageTestMethod 和 TraceAnalyticsDcLeakage
2. 根据Configuration内容智能分类：
   - 固定配置：直接字符串值 (如 "CDIE_Hot_Vmin_High_cdie_bscan_leakage_legacy_pins_10uA")
   - 规则配置：SIO_BSCAN_PCD_Rules.Temp函数调用
3. Configuration智能解析（直接提取 vs Temp函数四参数解析）
4. 生成一个Excel文件包含两个工作表，包含Test_Method列标识测试类型
"""

import re
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def parse_temp_function(config_text):
    """解析SIO_BSCAN_PCD_Rules.Temp函数的4个参数：HOT, COLD, PHMHOT, ALL"""
    result = {
        'primary_config': config_text,
        'hot_config': 'UNKNOWN',
        'cold_config': 'UNKNOWN',
        'phmhot_config': 'UNKNOWN',
        'all_config': 'UNKNOWN'
    }
    
    if 'SIO_BSCAN_PCD_Rules.Temp(' in config_text:
        params = re.findall(r'"([^"]+)"', config_text)
        if params:
            if len(params) >= 1:
                result['hot_config'] = params[0]
                result['primary_config'] = params[0]  # 使用HOT作为主配置
            if len(params) >= 2:
                result['cold_config'] = params[1]
            if len(params) >= 3:
                result['phmhot_config'] = params[2]
            if len(params) >= 4:
                result['all_config'] = params[3]
    else:
        # 不是Temp函数，直接提取
        result['primary_config'] = config_text.replace('"', '').strip()
    
    return result

# extract_field函数已被按行处理方法替代，不再需要

def process_mtpl_file(input_file):
    """处理MTPL文件"""
    print("🚀 开始处理MTPL文件...")
    print(f"📁 文件: {os.path.basename(input_file)}")
    
    # 读取文件
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            content = f.read()
        print(f"📄 文件读取成功")
    except Exception as e:
        print(f"❌ 文件读取失败: {e}")
        return None
    
    # 使用可靠的按行处理方法
    lines = content.splitlines()
    print(f"📄 文件共 {len(lines)} 行")
    
    # 找到所有DC泄漏测试实例的起始行（支持多种类型）
    test_lines = []
    prime_count = 0
    trace_count = 0
    
    for i, line in enumerate(lines):
        # 匹配 PrimeDcLeakageTestMethod
        if re.match(r'Test\s+PrimeDcLeakageTestMethod\s+\S+', line):
            test_lines.append((i, line, 'PrimeDcLeakageTestMethod'))
            prime_count += 1
        # 匹配 TraceAnalyticsDcLeakage
        elif re.match(r'Test\s+TraceAnalyticsDcLeakage\s+\S+', line):
            test_lines.append((i, line, 'TraceAnalyticsDcLeakage'))
            trace_count += 1
    
    print(f"🔍 找到测试实例总数: {len(test_lines)}")
    print(f"  • PrimeDcLeakageTestMethod: {prime_count} 个")
    print(f"  • TraceAnalyticsDcLeakage: {trace_count} 个")
    
    if not test_lines:
        print("❌ 未找到任何测试实例")
        return None
    
    # 处理实例
    fixed_instances = []  # BypassPort = 1
    rule_instances = []   # BypassPort = -1
    
    for idx, (line_num, line, test_type) in enumerate(test_lines, 1):
        if idx % 10 == 0:
            print(f"   处理进度：{idx}/{len(test_lines)}")
        
        # 提取实例名（根据测试类型使用不同的正则表达式）
        if test_type == 'PrimeDcLeakageTestMethod':
            name_match = re.search(r'Test\s+PrimeDcLeakageTestMethod\s+(\S+)', line)
        elif test_type == 'TraceAnalyticsDcLeakage':
            name_match = re.search(r'Test\s+TraceAnalyticsDcLeakage\s+(\S+)', line)
        else:
            continue
            
        if not name_match:
            continue
        name = name_match.group(1)
        
        # 从当前行开始向下查找字段（通常在接下来的30行内）
        test_type_field = 'UNKNOWN'  # TestType字段值
        bypass_port = 'UNKNOWN' 
        raw_config = 'UNKNOWN'
        
        for i in range(line_num + 1, min(line_num + 50, len(lines))):
            current_line = lines[i]
            
            # 提取TestType
            if 'TestType' in current_line and test_type_field == 'UNKNOWN':
                test_type_match = re.search(r'TestType\s*=\s*"([^"]+)"', current_line)
                if test_type_match:
                    test_type_field = test_type_match.group(1)
            
            # 提取BypassPort
            if 'BypassPort' in current_line and bypass_port == 'UNKNOWN':
                bypass_match = re.search(r'BypassPort\s*=\s*([^;]+);', current_line)
                if bypass_match:
                    bypass_port = bypass_match.group(1).strip()
            
            # 提取Configuration（可能跨多行）
            if 'Configuration' in current_line and raw_config == 'UNKNOWN':
                # 如果Configuration在同一行完成
                if '=' in current_line and ';' in current_line:
                    config_match = re.search(r'Configuration\s*=\s*([^;]+);', current_line)
                    if config_match:
                        raw_config = config_match.group(1).strip()
                else:
                    # Configuration可能跨多行，收集直到找到分号
                    config_lines = [current_line]
                    for j in range(i + 1, min(i + 10, len(lines))):
                        config_lines.append(lines[j])
                        if ';' in lines[j]:
                            break
                    
                    config_text = ' '.join(config_lines)
                    config_match = re.search(r'Configuration\s*=\s*([^;]+);', config_text, re.DOTALL)
                    if config_match:
                        raw_config = config_match.group(1).strip()
            
            # 如果遇到下一个Test，停止搜索
            if current_line.strip().startswith('Test ') and i > line_num + 1:
                break
        
        # 解析Configuration
        parsed_config = parse_temp_function(raw_config)
        
        instance = {
            'Test_Instance_Name': name,
            'Configuration': parsed_config['primary_config'],
            'TestType': test_type_field,  # 重命名避免与循环变量冲突
            'Test_Method': test_type,     # 新增：测试方法类型
            'BypassPort': bypass_port,
            'Hot_Config': parsed_config['hot_config'],
            'Cold_Config': parsed_config['cold_config'],
            'PhmHot_Config': parsed_config['phmhot_config'],
            'All_Config': parsed_config['all_config']
        }
        
        # 根据Configuration内容分类，而不是BypassPort
        if 'SIO_BSCAN_PCD_Rules.Temp(' in raw_config:
            # 包含Temp函数的是规则配置
            rule_instances.append(instance)
        else:
            # 直接字符串值的是固定配置
            fixed_instances.append(instance)
    
    print(f"✅ 处理完成")
    print(f"📊 固定配置 (直接Configuration值): {len(fixed_instances)} 个实例")
    print(f"📊 规则配置 (SIO_BSCAN_PCD_Rules.Temp函数): {len(rule_instances)} 个实例")
    
    return fixed_instances, rule_instances

def save_to_excel(fixed_data, rule_data, output_file):
    """保存到Excel文件，包含两个子表"""
    if not OPENPYXL_AVAILABLE:
        print("❌ 需要安装openpyxl库: pip install openpyxl")
        return False
    
    try:
        # 创建Excel工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认工作表
        
        # 原始表头定义（现在每个子表使用各自的表头）
        # headers = ['Test_Instance_Name', 'Configuration', 'TestType', 'BypassPort',
        #           'Hot_Config', 'Cold_Config', 'PhmHot_Config', 'All_Config']
        
        # 创建固定配置子表
        if fixed_data:
            ws_fixed = wb.create_sheet(title="Hardcoded_Configuration")
            
            # 固定配置的表头（不包含温度配置列）
            fixed_headers = ['Test_Instance_Name', 'Configuration', 'TestType', 'Test_Method', 'BypassPort']
            
            # 添加表头
            ws_fixed.append(fixed_headers)
            
            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for cell in ws_fixed[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 添加数据（只包含基本字段）
            for instance in fixed_data:
                row = [
                    instance['Test_Instance_Name'],
                    instance['Configuration'],
                    instance['TestType'],
                    instance['Test_Method'],
                    instance['BypassPort']
                ]
                ws_fixed.append(row)
            
            # 调整列宽
            for column in ws_fixed.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_fixed.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✅ 固定配置子表已创建: {len(fixed_data)} 行")
        
        # 创建规则配置子表
        if rule_data:
            ws_rule = wb.create_sheet(title="Rule_Configuration")
            
            # 规则配置的表头（不包含Configuration列）
            rule_headers = ['Test_Instance_Name', 'TestType', 'Test_Method', 'BypassPort',
                           'Hot_Config', 'Cold_Config', 'PhmHot_Config', 'All_Config']
            
            # 添加表头
            ws_rule.append(rule_headers)
            
            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")
            for cell in ws_rule[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 添加数据（不包含Configuration列）
            for instance in rule_data:
                row = [
                    instance['Test_Instance_Name'],
                    instance['TestType'],
                    instance['Test_Method'],
                    instance['BypassPort'],
                    instance['Hot_Config'],
                    instance['Cold_Config'],
                    instance['PhmHot_Config'],
                    instance['All_Config']
                ]
                ws_rule.append(row)
            
            # 调整列宽
            for column in ws_rule.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_rule.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✅ 规则配置子表已创建: {len(rule_data)} 行")
        
        # 保存Excel文件
        wb.save(output_file)
        print(f"💾 Excel文件已保存: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ 保存Excel失败: {e}")
        return False

def select_file():
    """文件选择对话框"""
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title="选择MTPL文件",
        filetypes=[
            ("MTPL文件", "*.mtpl"),
            ("所有文件", "*.*")
        ],
        initialdir=os.getcwd()
    )
    
    root.destroy()
    return file_path if file_path else None

def main():
    """主函数"""
    print("🚀 最终版MTPL提取工具")
    print("📊 生成Excel文件包含两个子表")
    print("=" * 50)
    
    # 文件选择
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        print(f"📁 命令行参数文件：{input_file}")
    else:
        print("📁 请选择MTPL文件...")
        input_file = select_file()
        
        if not input_file:
            print("❌ 未选择文件，程序退出")
            return
        
        print(f"📁 已选择文件：{input_file}")
    
    if not os.path.exists(input_file):
        error_msg = f"文件不存在：{input_file}"
        print(f"❌ {error_msg}")
        if 'tkinter' in sys.modules:
            messagebox.showerror("错误", error_msg)
        return
    
    # 处理文件
    result = process_mtpl_file(input_file)
    if not result:
        return
    
    fixed_data, rule_data = result
    
    if not fixed_data and not rule_data:
        print("⚠️  未找到有效的测试实例")
        return
    
    # 显示示例
    if fixed_data:
        print(f"\n🔍 固定配置示例:")
        inst = fixed_data[0]
        print(f"  {inst['Test_Instance_Name']}")
        print(f"  Configuration: {inst['Configuration'][:50]}...")
        print(f"  TestType: {inst['TestType']}")
    
    if rule_data:
        print(f"\n🔍 规则配置示例:")
        inst = rule_data[0]
        print(f"  {inst['Test_Instance_Name']}")
        print(f"  Configuration: {inst['Configuration'][:50]}...")
        print(f"  TestType: {inst['TestType']}")
        if inst['Hot_Config'] != 'UNKNOWN':
            print(f"  HOT配置: {inst['Hot_Config'][:40]}...")
    
    # 生成输出文件名
    base_name = os.path.splitext(input_file)[0]
    output_file = f"{base_name}.xlsx"
    
    # 保存到Excel
    if save_to_excel(fixed_data, rule_data, output_file):
        success_msg = f"""🎉 提取完成！

📊 处理统计：
• 固定配置实例：{len(fixed_data)} 个 (直接Configuration字符串值)
• 规则配置实例：{len(rule_data)} 个 (SIO_BSCAN_PCD_Rules.Temp函数)
• 总实例数：{len(fixed_data) + len(rule_data)} 个

📁 输出文件：
• {os.path.basename(output_file)}
• 包含两个子表：
  - Fixed_Configuration (5个字段)
  - Rule_Configuration (8个字段，包含温度配置)

🔍 功能特点：
• 智能识别两种测试类型 (PrimeDcLeakageTestMethod & TraceAnalyticsDcLeakage)
• Configuration内容智能分类 (固定值 vs Temp函数)
• Temp函数4参数解析 (HOT/COLD/PHMHOT/ALL)
• 按Configuration类型自动分类到不同子表
"""
        print(success_msg)
        if 'tkinter' in sys.modules:
            messagebox.showinfo("成功", success_msg)
    else:
        print("❌ 保存失败")

if __name__ == "__main__":
    main() 